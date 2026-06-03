"""Club View — full drill-in page for one club.

URL: /club_view?club_id=X

Layout (top → bottom):
  1. Header — display name, identity strip (league/formation/manager/agent prefs),
     conditional badges (Parachute Year N, Manager Change)
  2. 4 metric tiles — Selling Pressure, Transfer Budget, Wage Cap, Squad Size
  3. Selling Pressure breakdown — 5 components with what fired
  4. Seller-side section (hidden if no sellable players):
       Top 3 Likely to Move callout + Sellable Players table
  5. Buyer-side section (hidden if no buyer demand):
       Stated Requests table + Matches as Buyer table
  6. Sections never render placeholders — empty sections are removed entirely
"""

from __future__ import annotations

from datetime import datetime
import pandas as pd
import streamlit as st

import db
import labels
import components as ui

# config.DEMAND_MAPPED_LEAGUES — demand-side restriction applied below.
import sys as _sys
from pathlib import Path as _Path
_sys.path.insert(0, str(_Path(__file__).resolve().parent.parent.parent))
import config

st.set_page_config(page_title="Brokerage Engine", page_icon="🏟️", layout="wide")
ui.inject_css()
ui.render_global_search(db.get_player_search_options(), db.get_club_search_options())

# ─── Resolve query param ─────────────────────────────────────────────────────

raw_cid = st.query_params.get("club_id")

# ═════════════════════════════════════════════════════════════════════════════
# BROWSER VIEW (no query param) — pick-a-club table + sidebar filters
# ═════════════════════════════════════════════════════════════════════════════
if not raw_cid:
    st.title("Clubs")
    st.caption("Pick a club to see selling pressure, sellable players, and buyer demand.")

    # Pull all clubs with pressure + counts
    con = db.get_connection()
    club_rows = con.execute("""
        SELECT cp.club_id, cp.name AS official_name, cp.league_id,
               cp.total_pressure_score, cp.manager_change_flag,
               (SELECT COUNT(*) FROM player_universe pu
                WHERE pu.parent_club_id = cp.club_id
                  AND pu.sellability_status = 'sellable_now') AS sellable_count,
               (SELECT COUNT(*) FROM map_club_requests WHERE club_id = cp.club_id) AS requests_count,
               (SELECT COUNT(*) FROM matches WHERE buyer_club_id = cp.club_id) AS matches_as_buyer
        FROM club_pressure cp
        ORDER BY cp.total_pressure_score DESC NULLS LAST
    """).fetchall()
    cols = ["club_id","official_name","league_id","total_pressure_score",
            "manager_change_flag",
            "sellable_count","requests_count","matches_as_buyer"]

    # Parachute clubs — read from data/parachute_payments.xlsx (column not on club_pressure).
    # Cached in session so we don't re-open the workbook on every rerun.
    @st.cache_data(show_spinner=False)
    def _load_parachute_club_ids() -> set[int]:
        import openpyxl, pathlib
        p = pathlib.Path(__file__).parent.parent.parent / "data" / "parachute_payments.xlsx"
        if not p.exists():
            return set()
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        ws = wb.active
        ids: set[int] = set()
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:  # header row
                continue
            if row and row[0] is not None:
                try:
                    ids.add(int(row[0]))
                except (TypeError, ValueError):
                    pass
        return ids

    parachute_ids = _load_parachute_club_ids()

    rows = []
    for r in club_rows:
        d = dict(zip(cols, r))
        d["display_name"] = labels.club_display_name(d["club_id"], d["official_name"])
        d["pressure"] = float(d["total_pressure_score"] or 0.0)
        d["on_parachute"] = int(d["club_id"]) in parachute_ids
        d["has_mgr_change"] = d["manager_change_flag"] == 1
        # Matches as seller — count players parented at this club who appear in matches
        d["matches_as_seller"] = con.execute(
            "SELECT COUNT(*) FROM matches m JOIN player_universe pu ON pu.player_id=m.player_id "
            "WHERE pu.parent_club_id = ?", (d["club_id"],)
        ).fetchone()[0]
        d["total_matches"] = d["matches_as_buyer"] + d["matches_as_seller"]
        rows.append(d)
    df_all = pd.DataFrame(rows)

    # ─── Sidebar filters ─────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("### Filters")
        leagues = sorted(df_all["league_id"].dropna().unique().tolist())
        sel_leagues = st.multiselect(
            "League", options=leagues, default=leagues,
            format_func=labels.league_name,
        )
        p_min = 0
        p_max = int(df_all["pressure"].max()) + 1 if len(df_all) else 100
        sel_p = st.slider("Selling pressure", p_min, p_max, (p_min, p_max), 1)
        has_sellable = st.checkbox("Has sellable players")
        has_requests = st.checkbox("Has buyer requests")
        parachute = st.checkbox("On parachute payments")
        mgr_change = st.checkbox("Manager change detected")

    # ─── Apply filters ───────────────────────────────────────────────────
    filtered = df_all.copy()
    if sel_leagues:
        filtered = filtered[filtered["league_id"].isin(sel_leagues)]
    filtered = filtered[filtered["pressure"].between(sel_p[0], sel_p[1])]
    if has_sellable:
        filtered = filtered[filtered["sellable_count"] > 0]
    if has_requests:
        filtered = filtered[filtered["requests_count"] > 0]
    if parachute:
        filtered = filtered[filtered["on_parachute"]]
    if mgr_change:
        filtered = filtered[filtered["has_mgr_change"]]
    filtered = filtered.sort_values("pressure", ascending=False).reset_index(drop=True)

    # ─── Selector ────────────────────────────────────────────────────────
    options = ["— pick a club —"] + filtered["display_name"].tolist()

    def _on_select_club():
        sel = st.session_state.get("club_browser_select")
        if sel and sel != "— pick a club —":
            match = filtered[filtered["display_name"] == sel]
            if len(match):
                st.query_params["club_id"] = str(int(match.iloc[0]["club_id"]))

    st.selectbox(
        "Jump to club",
        options=options, index=0, key="club_browser_select",
        on_change=_on_select_club,
    )

    # ─── Stats row ───────────────────────────────────────────────────────
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Clubs shown", len(filtered))
    c2.metric("Total in universe", len(df_all))
    c3.metric("Median selling pressure", f"{filtered['pressure'].median():.1f}" if len(filtered) else "—")
    c4.metric("Top pressure", f"{filtered['pressure'].max():.1f}" if len(filtered) else "—")

    st.markdown("")

    # ─── Browser table ──────────────────────────────────────────────────
    if len(filtered) == 0:
        st.info("No clubs match the current filters.")
        st.stop()

    filtered["league_display"] = filtered["league_id"].apply(labels.league_name)
    filtered["Club_html"] = filtered.apply(
        lambda r: ui.club_link(int(r["club_id"]), str(r["display_name"])),
        axis=1,
    )

    # ─── Sort state from query params (drives clickable column headers) ──
    _CV_SORTABLE = {
        "Club":             "club",
        "League":           "league",
        "Selling Pressure": "pressure",
        "Sellable":         "sellable",
        "Requests":         "requests",
        "Total Matches":    "matches",
    }
    _CV_KEY_TO_COL = {
        "club":     "display_name",
        "league":   "league_display",
        "pressure": "pressure",
        "sellable": "sellable_count",
        "requests": "requests_count",
        "matches":  "total_matches",
    }
    cv_sort_key = st.query_params.get("sort", "pressure")
    cv_sort_dir = st.query_params.get("dir", "desc")
    if cv_sort_key not in _CV_KEY_TO_COL:
        cv_sort_key, cv_sort_dir = "pressure", "desc"
    cv_ascending = (cv_sort_dir == "asc")
    filtered = filtered.sort_values(
        _CV_KEY_TO_COL[cv_sort_key], ascending=cv_ascending,
        kind="stable", na_position="last",
    ).reset_index(drop=True)
    filtered.insert(0, "rank", range(1, len(filtered) + 1))

    display = filtered[[
        "rank", "Club_html", "league_display", "pressure",
        "sellable_count", "requests_count", "total_matches",
    ]].rename(columns={
        "rank":              "#",
        "Club_html":         "Club",
        "league_display":    "League",
        "pressure":          "Selling Pressure",
        "sellable_count":    "Sellable",
        "requests_count":    "Requests",
        "total_matches":     "Total Matches",
    })
    display["Selling Pressure"] = display["Selling Pressure"].apply(lambda v: f"{v:.1f}")

    def _bg_p(v):
        try: return ui.green_gradient(float(v), 0, 100)
        except (TypeError, ValueError): return ""

    styled = display.style.map(_bg_p, subset=["Selling Pressure"])
    ui.render_html_table(
        styled, max_height_px=640,
        sortable=_CV_SORTABLE,
        current_sort=(cv_sort_key, cv_sort_dir),
    )
    st.stop()

# ═════════════════════════════════════════════════════════════════════════════
# DETAIL VIEW (query param present) — full club profile
# ═════════════════════════════════════════════════════════════════════════════
try:
    cid = int(raw_cid)
except (TypeError, ValueError):
    st.error(f"Bad club_id: {raw_cid!r}")
    st.stop()

# Top spacer so the "← Back" button isn't visually clipped against the
# st.navigation tab bar at the very top of the viewport.
st.markdown("<div style='margin-top:0.5rem;'></div>", unsafe_allow_html=True)
if st.button("← Back to all clubs", key="back_to_clubs"):
    try:
        del st.query_params["club_id"]
    except KeyError:
        pass
    st.rerun()

data = db.get_club(cid)
if data["pressure"] is None:
    st.error(f"No club with club_id={cid}")
    st.stop()

pressure  = data["pressure"]
overview  = data["overview"]
sellable  = data["sellable_here"]
# Demand-side restriction: only show stated requests from clubs in mapped
# leagues, and only show matches-as-buyer rows where the buyer (= this club)
# is in a mapped league. Non-mapped league clubs (Saudi/MLS/Greek/etc.) get
# an empty section here — consistent with not surfacing untrusted demand on
# Position View, Targets, All Matches, and Player View.
requests  = data["requests"]
if "league" in requests.columns:
    requests = requests[requests["league"].isin(config.DEMAND_MAPPED_LEAGUES)].reset_index(drop=True)
buyer_mx  = data["matches_as_buyer"]
if "buyer_league_id" in buyer_mx.columns:
    buyer_mx = buyer_mx[buyer_mx["buyer_league_id"].isin(config.DEMAND_MAPPED_LEAGUES)].reset_index(drop=True)
snapshot  = db.get_snapshot_date()

club_name = labels.club_display_name(cid, pressure["name"])
league_full = labels.league_name(pressure.get("league_id"))

# ─── Header block ───────────────────────────────────────────────────────────

st.markdown(f"## {club_name}")

# Identity strip — country · league · formation · manager (with tenure)
identity_parts: list[str] = []
country_name, country_flag = labels.country_for_league(pressure.get("league_id"))
if country_name:
    identity_parts.append(f"{country_flag} {country_name}" if country_flag else country_name)
if league_full:
    identity_parts.append(league_full)
if overview is not None and pd.notna(overview.get("formation")):
    identity_parts.append(labels.format_formation(overview["formation"]))
mgr = (overview.get("manager") if overview is not None else None) or "—"
mgr_str = f"Manager: {mgr}" if mgr and mgr != "—" else None
if mgr_str: identity_parts.append(mgr_str)
identity = " · ".join(identity_parts)
st.markdown(f"<div style='color:#374151; font-size:0.95rem;'>{identity}</div>", unsafe_allow_html=True)

agent_prefs = overview.get("agent_preferences") if overview is not None else None
if agent_prefs and str(agent_prefs).strip() not in ("", "—", "N/A"):
    st.markdown(
        f"<div style='color:#6b7280; font-size:0.85rem; margin-top:2px;'>"
        f"Preferred agents to work with: {agent_prefs}</div>",
        unsafe_allow_html=True,
    )

# Badges row — every firing selling-pressure trigger surfaces as a pill.
# At-a-glance summary for "why this club is a selling target"; the breakdown
# panel further down has the full weights + components view.

# Parachute year — check parachute_payments.xlsx by club_id
_parachute_year: int | None = None
try:
    import openpyxl, pathlib
    _pp = pathlib.Path(__file__).parent.parent.parent / "data" / "parachute_payments.xlsx"
    if _pp.exists():
        wb = openpyxl.load_workbook(_pp, read_only=True, data_only=True)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row or row[0] is None:
                continue
            try:
                if int(row[0]) == cid and row[2] is not None:
                    _parachute_year = int(row[2])
                    break
            except (TypeError, ValueError):
                pass
except Exception:
    pass

# Per-league quartile reference values for the three score-based badges.
# Squad Overload uses the same top-quartile gate as Contract Leverage and
# Net Spend — a flat "≥3 positions oversupplied" rule fired for 64% of clubs
# across all leagues and 78% of Bundesliga clubs, which made the badge
# meaningless. Top-quartile + a 3-position floor surfaces the genuinely
# unusual cases.
_league_id = pressure.get("league_id")
_cl_p75 = _ns_p75 = _so_p75 = None
if _league_id:
    quartile_rows = db.get_connection().execute("""
        SELECT contract_leverage_score, net_spend_score, squad_oversupply_score
        FROM club_pressure
        WHERE league_id = ?
    """, (_league_id,)).fetchall()
    cl_vals = sorted([r[0] for r in quartile_rows if r[0] is not None])
    ns_vals = sorted([r[1] for r in quartile_rows if r[1] is not None])
    so_vals = sorted([r[2] for r in quartile_rows if r[2] is not None])
    if cl_vals:
        _cl_p75 = cl_vals[int(len(cl_vals) * 0.75)] if len(cl_vals) >= 4 else cl_vals[-1]
    if ns_vals:
        _ns_p75 = ns_vals[int(len(ns_vals) * 0.75)] if len(ns_vals) >= 4 else ns_vals[-1]
    if so_vals:
        _so_p75 = so_vals[int(len(so_vals) * 0.75)] if len(so_vals) >= 4 else so_vals[-1]

# Count oversupplied positions from scoring_basis text
_scoring_basis_str = pressure.get("scoring_basis") or ""
_oversupplied_count = 0
if isinstance(_scoring_basis_str, str) and "oversupplied:" in _scoring_basis_str:
    _tail = _scoring_basis_str.split("oversupplied:", 1)[1].strip()
    _oversupplied_count = len([p for p in _tail.split(",") if p.strip()])


def _badge(text: str, bg: str, fg: str) -> str:
    return (
        f"<span style='display:inline-block; padding:3px 10px; border-radius:10px; "
        f"background:{bg}; color:{fg}; font-weight:600; font-size:0.85rem;'>{text}</span>"
    )


badges: list[str] = []
# Recently relegated is the dominant seller-side signal — surface it first.
if pressure.get("recently_relegated") == 1:
    badges.append(_badge("⬇️ Recently Relegated", "#fecaca", "#7f1d1d"))
if _parachute_year is not None:
    badges.append(_badge(f"🪂 Parachute Year {_parachute_year}", "#fee2e2", "#7f1d1d"))
if pressure.get("public_must_sell_flag") == 1:
    badges.append(_badge("📢 Public Must-Sell", "#fee2e2", "#7f1d1d"))
if pressure.get("manager_change_flag") == 1:
    badges.append(_badge("🔄 Manager Change", "#fef3c7", "#78350f"))
_cl_score = pressure.get("contract_leverage_score") or 0
if _cl_p75 is not None and _cl_score >= _cl_p75 and _cl_score > 0:
    badges.append(_badge("📊 Contract Leverage", "#fef3c7", "#78350f"))
_so_score = pressure.get("squad_oversupply_score") or 0
if (_so_p75 is not None and _so_score >= _so_p75
        and _so_score > 0 and _oversupplied_count >= 3):
    badges.append(_badge(f"⚠️ Squad Overload ({_oversupplied_count} positions)", "#fef3c7", "#78350f"))

if badges:
    st.markdown(
        "<div style='margin-top:10px; display:flex; flex-wrap:wrap; gap:8px;'>"
        + "".join(badges) + "</div>",
        unsafe_allow_html=True,
    )

st.markdown("")

# ─── 4 metric tiles ─────────────────────────────────────────────────────────

total_pressure = pressure.get("total_pressure_score") or 0.0
max_fee = overview.get("highest_transfer_fee_2526_eur") if overview is not None else None
max_wage = overview.get("max_salary_pw_2526_eur") if overview is not None else None
# Squad size — from senior_roster
con = db.get_connection()
squad_size = con.execute(
    "SELECT COUNT(*) FROM senior_roster WHERE club_id = ? AND minutes_last_18m > 0",
    (cid,),
).fetchone()[0]
if squad_size == 0:
    # Fall back to total roster headcount (for external/second-tier clubs without minutes)
    squad_size = con.execute("SELECT COUNT(*) FROM senior_roster WHERE club_id = ?", (cid,)).fetchone()[0]

def _money(x):
    if x is None or pd.isna(x): return "—"
    x = float(x)
    if abs(x) >= 1_000_000: return f"€{x/1_000_000:.1f}m"
    if abs(x) >= 1_000:     return f"€{x/1_000:.0f}k"
    return f"€{int(x)}"

c1, c2, c3, c4 = st.columns(4)
c1.metric("Selling pressure", f"{total_pressure:.1f}",
          help="0-100 score across 5 weighted components: contract leverage (20%), squad oversupply (20%), net spend (20%), manager change (15%), public must-sell (25%).")
c2.metric("Transfer budget",  _money(max_fee),
          help="Highest transfer fee paid by this club in 25/26 — proxy for upper-bound transfer-fee capacity.")
c3.metric("Wage cap",         _money(max_wage),
          help="Max salary per week observed at this club in 25/26.")
c4.metric("Squad size",       squad_size,
          help="Active senior-roster players (minutes_last_18m > 0; falls back to total roster for second-tier clubs).")

st.markdown("")

# ─── Selling Pressure breakdown ─────────────────────────────────────────────

st.markdown("### Selling pressure breakdown")
_floor_note = (
    " · <span style='color:#b45309;'>recently relegated — pressure floor of 65 applies</span>"
    if pressure.get("recently_relegated") == 1 else ""
)
st.markdown(
    f"<div style='color:#6b7280; font-style:italic;'>"
    f"Total: <strong style='color:#111827;'>{total_pressure:.1f} / 100</strong>{_floor_note}</div>",
    unsafe_allow_html=True,
)

cl_s = pressure.get("contract_leverage_score") or 0
so_s = pressure.get("squad_oversupply_score") or 0
mc_f = pressure.get("manager_change_flag") == 1
ps_f = pressure.get("public_must_sell_flag") == 1
releg = pressure.get("recently_relegated") == 1

# Relegation component (tier-scaled; PL = 100 today) + weighted contributions.
releg_score = 100.0 if releg else 0.0
c_releg = 0.40 * releg_score
c_must  = 0.20 * (100.0 if ps_f else 0.0)
c_cl    = 0.15 * cl_s
c_so    = 0.15 * so_s
c_mgr   = 0.10 * (100.0 if mc_f else 0.0)

scoring_basis = pressure.get("scoring_basis") or ""
oversupply_text = ""
if isinstance(scoring_basis, str) and "oversupplied:" in scoring_basis:
    oversupply_text = scoring_basis.split("oversupplied:", 1)[1].strip()


def _contract_leverage_copy(score: float) -> str:
    """Band-based — describes THIS club's actual contract situation, not generic."""
    if score < 25:
        return ("Most of the squad is on longer contracts — little run-down pressure. "
                "Contract timing isn't forcing sales here.")
    if score < 50:
        return ("About a third of the first team (minutes-weighted) is in the final two "
                "years of contract — some sell-now pressure, but most of the squad is "
                "tied down on longer deals.")
    if score < 70:
        return ("Around half the first team is running down contracts — strong "
                "sell-now-or-lose-value pressure, and buyers gain leverage on price.")
    return ("A majority of the squad is in the final two years — acute Bosman risk; "
            "the club must move players or lose them for free.")


def _squad_overload_copy(score: float, buckets: str) -> str:
    if score <= 0:
        return "No positions overstocked."
    suffix = f" ({buckets})" if buckets else ""
    if score <= 20:
        return f"One to two positions overstocked — a minor logjam{suffix}."
    return (f"Several positions overstocked — significant surplus the club is "
            f"motivated to clear{suffix}.")


def _line(label: str, score_str: str, contribution: float, weight: str, body: str) -> str:
    return (
        f"<div style='padding:10px 0; border-bottom:1px solid #f3f4f6;'>"
        f"<div style='display:flex; justify-content:space-between; align-items:baseline;'>"
        f"<div style='font-weight:600;'>{label}</div>"
        f"<div style='color:#6b7280; font-size:0.85rem;'>{weight}</div></div>"
        f"<div style='font-size:0.9rem; color:#374151;'>"
        f"{score_str} · contributes <strong>{contribution:.1f}</strong> — {body}</div>"
        f"</div>"
    )


_releg_body = (
    "Relegated from the Premier League end of 25/26. The dominant seller-side "
    "signal — wage bill no longer matches Championship revenue, players push to "
    "stay at the top level, structurally forced to sell this window."
    if releg else
    "Not recently relegated — no relegation pressure on this club."
)

pressure_html = (
    f"<div style='background:#f9fafb; border:1px solid #e5e7eb; border-radius:6px; padding:12px;'>"
    + _line("Recent relegation", f"<strong>{releg_score:.0f} / 100</strong>", c_releg, "weight 40%",
            _releg_body)
    + _line("Known must-sell", "flag ON" if ps_f else "flag OFF", c_must, "weight 20%",
            "Publicly signalling a need to sell: PSR/FFP, parachute reset, or parent-company stress."
            if ps_f else "No public must-sell flag set.")
    + _line("Contract leverage", f"<strong>{cl_s:.0f} / 100</strong>", c_cl, "weight 15%",
            _contract_leverage_copy(cl_s))
    + _line("Squad overload", f"<strong>{so_s:.0f} / 100</strong>", c_so, "weight 15%",
            _squad_overload_copy(so_s, oversupply_text))
    + _line("Manager change", "flag ON" if mc_f else "flag OFF", c_mgr, "weight 10%",
            "Manager/sporting-director change confirmed — regime change typically triggers squad churn."
            if mc_f else "No managerial change flagged.")
    + "</div>"
)
st.markdown(pressure_html, unsafe_allow_html=True)
st.markdown("")

# ─── Seller-side section (hidden if no sellable players at this club) ─────

n_sellable = len(sellable)
if n_sellable > 0:
    st.markdown(f"## Sellable players ({n_sellable})")
    # Top 3 Likely to Move callout (from club_pressure column populated by script 09)
    top3 = pressure.get("top_3_likely_to_move") or ""
    if top3:
        # The text was already display-name-swapped by script 10; render inline
        st.markdown(
            f"<div style='background:#eff6ff; border:1px solid #bfdbfe; border-radius:6px; padding:10px; margin-bottom:12px;'>"
            f"<strong>Top 3 likely to move:</strong> {top3}"
            f"</div>",
            unsafe_allow_html=True,
        )

    # Build display
    sellable_df = sellable.copy()
    # Display name + URL for player
    player_map = db.get_player_search_options()
    pname_lookup = {p["player_id"]: p["display_name"] for p in player_map}
    sellable_df["player_display"] = sellable_df.apply(
        lambda r: pname_lookup.get(int(r["player_id"]), r["name"]), axis=1
    )
    sellable_df["bucket_tier"] = sellable_df["position_bucket"].apply(labels.super_bucket)
    sellable_df["mv_str"] = sellable_df["current_tm_value_eur"].apply(_money)

    def _yrs(ce):
        if not ce or pd.isna(ce): return None
        try:
            return round((datetime.fromisoformat(str(ce)).date() - snapshot).days / 365.25, 2)
        except (TypeError, ValueError):
            return None
    sellable_df["yrs_left"] = sellable_df["contract_end_date"].apply(_yrs)

    # Number of buyer matches per player
    con2 = db.get_connection()
    match_counts = dict(con2.execute(
        f"SELECT player_id, COUNT(*) FROM matches "
        f"WHERE player_id IN ({','.join(str(int(x)) for x in sellable_df['player_id'])}) "
        f"GROUP BY player_id"
    ).fetchall()) if len(sellable_df) else {}
    sellable_df["buyer_count"] = sellable_df["player_id"].apply(
        lambda pid: match_counts.get(int(pid), 0)
    )

    sellable_display = sellable_df.sort_values("sellability_score", ascending=False).reset_index(drop=True)
    sellable_display.insert(0, "rank", range(1, len(sellable_display) + 1))

    sellable_display["Player_html"] = sellable_display.apply(
        lambda r: ui.player_link(int(r["player_id"]), str(r["player_display"])),
        axis=1,
    )

    show_df = sellable_display[[
        "rank", "Player_html", "age", "position_bucket", "bucket_tier",
        "mv_str", "yrs_left", "sellability_score", "buyer_count",
    ]].rename(columns={
        "rank":               "#",
        "Player_html":        "Player",
        "age":                "Age",
        "position_bucket":    "Position",
        "bucket_tier":        "Bucket",
        "mv_str":             "Market Value",
        "yrs_left":           "Years Remaining",
        "sellability_score":  "Sellability ▼",
        "buyer_count":        "Buyer Matches",
    })
    show_df["Sellability ▼"]   = show_df["Sellability ▼"].apply(lambda v: f"{v:.1f}")
    show_df["Years Remaining"] = show_df["Years Remaining"].apply(lambda v: f"{v:.1f}" if pd.notna(v) else "—")
    show_df["Position"]        = show_df["Position"].apply(labels.display_bucket)

    def _bucket_bg(val):
        colour = labels.TIER_COLOURS.get(val, "")
        return f"background-color: {colour}; font-weight: 600;" if colour else ""

    def _bg_sell(v):
        try: return ui.green_gradient(float(v), 15, 100)
        except (TypeError, ValueError): return ""

    styled = (
        show_df.style
        .map(_bucket_bg, subset=["Bucket"])
        .map(_bg_sell, subset=["Sellability ▼"])
    )
    ui.render_html_table(styled, max_height_px=min(520, 60 + 38 * len(show_df)))
    st.markdown("")

# ─── Buyer-side section (hidden if no demand entries OR no matches as buyer) ─

n_requests = len(requests)
n_buyer_mx = len(buyer_mx)
buyer_side_has_content = (n_requests > 0) or (n_buyer_mx > 0)

if buyer_side_has_content:
    if n_sellable > 0:
        # Visual divider — only when we're splitting both sides
        st.markdown(
            "<hr style='border:none; border-top:2px solid #e5e7eb; margin:24px 0 16px 0;'>",
            unsafe_allow_html=True,
        )

    # Sub A — Stated Requests
    if n_requests > 0:
        st.markdown(f"## Buyer demand — stated requests ({n_requests})")
        req_df = requests.copy()
        # Get linked_shortlisted_player from map_club_requests for explicit rows
        # (the union view in db.py doesn't include it — pull directly)
        more_cols = con.execute(
            f"SELECT request_id, role_notes, validated_by, linked_shortlisted_player "
            f"FROM map_club_requests WHERE club_id = ?", (cid,)
        ).fetchall()
        req_extra = {r[0]: {"role_notes": r[1], "validated_by": r[2], "linked": r[3]} for r in more_cols}

        req_df["pos_display"] = req_df["position_bucket"].apply(labels.display_bucket)
        req_df["budget_str"]  = req_df["max_transfer_fee_eur"].apply(_money)
        req_df["wage_str"]    = req_df["max_wage_pw_eur"].apply(_money)

        # Drop rows where Side, Buyer Budget AND Wage Cap are all empty/—.
        # A row with only a Position cell carries no actionable signal.
        def _is_blank(v) -> bool:
            if v is None or pd.isna(v):
                return True
            s = str(v).strip()
            return s in ("", "—", "-", "N/A", "nan", "None")

        req_df = req_df[
            ~(req_df["preferred_side"].apply(_is_blank)
              & req_df["budget_str"].apply(_is_blank)
              & req_df["wage_str"].apply(_is_blank))
        ].reset_index(drop=True)

        if len(req_df):
            # Render — Streamlit table. Layer, Demand, Validated columns all
            # dropped: Layer is internal terminology; Demand was always
            # "intel-derived" with zero per-row signal; Validated added noise
            # without a clear actionable line for the user. Final shape:
            # Position · Side · Buyer Budget · Wage Cap.
            req_show = req_df[[
                "pos_display", "preferred_side", "budget_str", "wage_str",
            ]].rename(columns={
                "pos_display":    "Position",
                "preferred_side": "Side",
                "budget_str":     "Buyer Budget",
                "wage_str":       "Wage Cap",
            })
            # No explicit height — Streamlit auto-fits to the row count, so
            # there's no blank space below the last data row.
            st.dataframe(req_show, hide_index=True, use_container_width=True)
        else:
            st.caption("No stated requests carry buyer-budget, wage, or side data.")

    # Sub B — Matches as Buyer (filter Kill List)
    if n_buyer_mx > 0:
        excluded_ids = db.get_excluded_ids()
        mx_df = buyer_mx[~buyer_mx["player_id"].isin(excluded_ids)].copy()
        if len(mx_df):
            st.markdown(
                f"## Matches as buyer ({len(mx_df)}) {ui.level_fit_info_icon()}",
                unsafe_allow_html=True,
            )
            mx_df["player_display"] = mx_df.apply(
                lambda r: labels.player_display_name(int(r["player_id"]), r["player_name"]), axis=1
            )
            mx_df["current_club_display"] = mx_df.apply(
                lambda r: labels.club_display_name(int(r["current_club_id"]) if pd.notna(r.get("current_club_id")) else None,
                                                    r["current_club"]), axis=1
            )
            mx_df["parent_league_display"] = mx_df["parent_league"].apply(labels.league_name)
            mx_df["bucket_tier"] = mx_df["position_bucket"].apply(labels.super_bucket)
            mx_df["mv_str"] = mx_df["current_tm_value_eur"].apply(_money)
            mx_df["rationale"] = mx_df.apply(ui.build_rationale, axis=1)

            mx_sorted = mx_df.sort_values("match_score", ascending=False).reset_index(drop=True)
            mx_sorted["Player_html"] = mx_sorted.apply(
                lambda r: ui.player_link(int(r["player_id"]), str(r["player_display"])),
                axis=1,
            )
            mx_sorted["ca_str"]         = mx_sorted["player_ca"].apply(lambda v: f"{float(v):.0f}" if pd.notna(v) else "—")
            mx_sorted["pa_str"]         = mx_sorted["player_pa"].apply(lambda v: f"{float(v):.0f}" if pd.notna(v) else "—")
            mx_sorted["level_fit_pill"] = mx_sorted["level_fit"].apply(ui.level_fit_pill)
            mx_show = (
                mx_sorted
                .assign(rank=lambda d: range(1, len(d) + 1))
                [["rank", "Player_html", "age", "position_bucket", "bucket_tier",
                  "current_club_display", "parent_league_display", "mv_str",
                  "ca_str", "pa_str", "level_fit_pill",
                  "match_score", "sellability_score", "rationale"]]
                .rename(columns={
                    "rank":                  "#",
                    "Player_html":           "Player",
                    "age":                   "Age",
                    "position_bucket":       "Position",
                    "bucket_tier":           "Bucket",
                    "current_club_display":  "Current Club",
                    "parent_league_display": "Parent League",
                    "mv_str":                "Market Value",
                    "ca_str":                "CA",
                    "pa_str":                "PA",
                    "level_fit_pill":        "Level fit",
                    "match_score":           "Match Score ▼",
                    "sellability_score":     "Sellability",
                    "rationale":             "Rationale",
                })
            )
            mx_show["Match Score ▼"] = mx_show["Match Score ▼"].apply(lambda v: f"{v:.1f}")
            mx_show["Sellability"]   = mx_show["Sellability"].apply(lambda v: f"{v:.1f}")
            mx_show["Position"]      = mx_show["Position"].apply(labels.display_bucket)

            def _bb(val):
                colour = labels.TIER_COLOURS.get(val, "")
                return f"background-color: {colour}; font-weight: 600;" if colour else ""

            def _bg_m(v):
                try: return ui.green_gradient(float(v), 10, 100)
                except (TypeError, ValueError): return ""
            def _bg_s(v):
                try: return ui.green_gradient(float(v), 15, 100)
                except (TypeError, ValueError): return ""

            styled_mx = (
                mx_show.style
                .map(_bb, subset=["Bucket"])
                .map(_bg_m, subset=["Match Score ▼"])
                .map(_bg_s, subset=["Sellability"])
            )
            ui.render_html_table(styled_mx, max_height_px=min(640, 60 + 38 * len(mx_show)))
