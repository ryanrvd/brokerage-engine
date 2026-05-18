"""
Step 18 — Regenerate data/manual_flags.xlsx (the user-facing review checklist).

One row per club. 16 columns: hard data (read-only) + 2 user-input flags +
1 free-text notes column + 1 auto-stamped last_reviewed column.

After Day 4 (manager scrape): auto-derives manager_change_flag from the TM
Mitarbeiter data in club_manager (populated by scripts/20_scrape_managers.py).
Four rules, priority-ordered:

  1. caretaker_title         — title contains 'interim', 'caretaker', 'acting', or 'assistant'
  2. tenure_<=6mo            — manager_tenure_months ≤ 6
  3. contract_expiring_2026  — manager_contract_end_date ≤ 2026-06-30
  4. vacancy_inferred        — manager_name IS NULL AND base_pressure ≥ 35

manager_change_basis shows which rule fired (or 'manual_override' if the user's
flag value disagrees with the auto-derived value).

public_must_sell_flag is auto-derived from data/parachute_payments.csv (a
hand-curated registry of clubs in their post-relegation parachute-payment window —
PL → Championship only, since no other covered league has the equivalent
declining-payment structure). public_must_sell_basis values:

  • parachute_yr1 | parachute_yr2 | parachute_yr3 — rule fired
  • manual_override — user value disagrees with auto-derived value
  • '' (empty) — no signal

The parachute CSV is auto-seeded on first run from manual_league_overrides.csv
(GB1 → GB2 transitions = newly relegated → yr1). Add yr2/yr3 entries by hand.

Manual override semantics (both flags):
  • First refresh after scrape: every flag is auto-set. basis = rule name.
  • Subsequent refresh: if the user has flipped the flag in the xlsx (so user_flag
    ≠ auto_flag), basis becomes 'manual_override' and user_flag is preserved.
    The pipeline NEVER overwrites a manually-set flag value.

"Base pressure" sorts the sheet — highest structural pressure floats to the top:
    0.25 × contract_leverage_score + 0.20 × squad_oversupply_score + 0.20 × net_spend_score

Idempotent. Reads existing values from:
  1. data/manual_flags.xlsx if it exists (subsequent runs — preserves user edits)
  2. else data/manual_flags.csv  (first-time CSV → xlsx migration)
  3. else seeds every flag at 0 (no prior data)
"""

import csv
import shutil
import sqlite3
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import config

XLSX_PATH = Path("data/manual_flags.xlsx")
BAK_XLSX_PATH = Path("data/manual_flags.xlsx.bak")  # rolling backup, refreshed each successful run
STAGING_XLSX_PATH = Path("data/manual_flags.staging.xlsx")
LEGACY_CSV_PATH = Path("data/manual_flags.csv")
PARACHUTE_XLSX = Path("data/parachute_payments.xlsx")
PARACHUTE_CSV = Path("data/parachute_payments.csv")  # legacy fallback during migration
OVERRIDES_CSV = Path("data/manual_league_overrides.csv")
RECONCILIATION_LOG_DIR = Path("logs")  # logs/script_18_reconciliation_<ts>.log

MANAGER_CHANGE_PROMPT = (
    "Auto-derived by 4 rules (see manager_change_basis). "
    "Flip manually to override; pipeline preserves manual overrides."
)
PUBLIC_MUST_SELL_PROMPT = (
    "Auto-derived from data/parachute_payments.csv (see public_must_sell_basis). "
    "Flip manually to override; pipeline preserves manual overrides."
)

# Header layout (17 columns)
HEADERS = [
    "club_id",                         # 1  RO (key)
    "club",                            # 2  RO
    "league",                          # 3  RO
    "current_base_pressure_score",     # 4  RO
    "current_manager_name",            # 5  RO (from club_manager)
    "manager_title",                   # 6  RO
    "manager_appointment_date",        # 7  RO
    "manager_tenure_months",           # 8  RO (computed)
    "manager_contract_end_date",       # 9  RO
    "manager_change_check",            # 10 RO prompt
    "manager_change_flag",             # 11 user 0/1
    "manager_change_basis",            # 12 RO (rule fired or 'manual_override')
    "public_must_sell_check",          # 13 RO prompt
    "public_must_sell_flag",           # 14 user 0/1
    "public_must_sell_basis",          # 15 RO (NEW — 'parachute_yr1/2/3' or 'manual_override')
    "notes_source",                    # 16 user text
    "last_reviewed",                   # 17 auto-stamped
]
WIDTHS  = [10, 36, 7, 14, 26, 22, 14, 12, 14, 56, 12, 22, 70, 12, 22, 40, 14]

# Column indices (1-based) for convenience
COL_MANAGER_CHANGE_FLAG     = 11
COL_MANAGER_CHANGE_BASIS    = 12
COL_PUBLIC_MUST_SELL_FLAG   = 14
COL_PUBLIC_MUST_SELL_BASIS  = 15
COL_NOTES_SOURCE            = 16
COL_LAST_REVIEWED           = 17

# Auto-derivation thresholds
TENURE_MAX_MONTHS_FOR_FLAG  = 6                       # tenure ≤ 6mo → flag
CONTRACT_CUTOFF             = date(2026, 6, 30)       # end-of-season 25/26
VACANCY_PRESSURE_THRESHOLD  = 35.0                    # base_pressure ≥ 35
# "assistant" included because clubs running with an Assistant Manager in charge
# are de facto in transition (the senior role is unfilled or unconfirmed).
CARETAKER_KEYWORDS          = ("interim", "caretaker", "acting", "assistant")

# Parachute payments only apply to PL → Championship in our 19-league coverage.
# Other relegations (La Liga, Bundesliga, etc.) have no equivalent declining-payment
# structure. If that ever changes, extend the allowed (from, to) pair set below.
PARACHUTE_FROM_TO = ("GB1", "GB2")


# ─── Parachute registry (data/parachute_payments.csv) ──────────────────────────

def _seed_parachute_csv_if_missing() -> None:
    """First-run seed of the legacy CSV. Only used if neither xlsx nor csv exist.

    Year-1 cohort is derivable from manual_league_overrides.csv (PL → Championship
    transitions). Year-2 and year-3 cohorts need historical games data — when
    that's not available the user is reminded to fill them in.

    The xlsx (data/parachute_payments.xlsx) is the canonical source going forward;
    build it with scripts/_build_parachute_xlsx.py for a pre-populated registry.
    """
    if PARACHUTE_XLSX.exists() or PARACHUTE_CSV.exists():
        return
    PARACHUTE_CSV.parent.mkdir(parents=True, exist_ok=True)

    seed_rows: list[dict] = []
    if OVERRIDES_CSV.exists():
        with OVERRIDES_CSV.open(newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                if row.get("override_league_id") == PARACHUTE_FROM_TO[1]:
                    seed_rows.append({
                        "club_id": row["club_id"],
                        "club_name": row["club_name"],
                        "parachute_year": "1",
                        "season_relegated": "25/26",
                        "notes": "auto-seeded from manual_league_overrides.csv (newly relegated for 26/27)",
                    })

    with PARACHUTE_CSV.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "club_id", "club_name", "parachute_year", "season_relegated", "notes",
        ])
        writer.writeheader()
        for row in seed_rows:
            writer.writerow(row)
    print(f"  Seeded {PARACHUTE_CSV} with {len(seed_rows)} yr1 entry/entries "
          "(run scripts/_build_parachute_xlsx.py to get a pre-populated xlsx).")


def _load_parachutes() -> tuple[dict[str, int], str]:
    """Returns ({club_id: parachute_year}, source_label).

    Prefer xlsx (data/parachute_payments.xlsx); fall back to legacy CSV. Skips
    invalid rows quietly. Rows with needs_review=1 in the xlsx are still loaded;
    the flag is informational for the user, not gating.
    """
    if PARACHUTE_XLSX.exists():
        return _load_parachutes_from_xlsx(), "xlsx"
    if PARACHUTE_CSV.exists():
        return _load_parachutes_from_csv(), "csv"
    return {}, "(none)"


def _load_parachutes_from_xlsx() -> dict[str, int]:
    out: dict[str, int] = {}
    wb = load_workbook(PARACHUTE_XLSX, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return out
    headers = [str(h) if h is not None else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}
    if "club_id" not in idx or "parachute_year" not in idx:
        return out
    for row in rows[1:]:
        cid_raw = row[idx["club_id"]]
        if cid_raw is None or str(cid_raw).strip() == "":
            continue
        try:
            yr = int(row[idx["parachute_year"]])
        except (TypeError, ValueError):
            continue
        if yr in (1, 2, 3):
            out[str(cid_raw).strip()] = yr
    return out


def _load_parachutes_from_csv() -> dict[str, int]:
    out: dict[str, int] = {}
    with PARACHUTE_CSV.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            cid = (row.get("club_id") or "").strip()
            try:
                yr = int((row.get("parachute_year") or "").strip())
            except ValueError:
                continue
            if cid and yr in (1, 2, 3):
                out[cid] = yr
    return out


def derive_public_must_sell_auto(parachute_year: int | None) -> tuple[int, str]:
    """Returns (auto_flag, auto_basis). Currently only the parachute rule is wired."""
    if parachute_year in (1, 2, 3):
        return 1, f"parachute_yr{parachute_year}"
    return 0, ""


# ─── Read existing values ──────────────────────────────────────────────────────

def _load_existing_flags() -> tuple[dict[str, dict], bool, bool]:
    """Returns (club_id → prev-state dict, mgr_basis_present, ps_basis_present).

    The two booleans indicate whether the previous xlsx already had the
    manager_change_basis / public_must_sell_basis columns. False on the first
    auto-derivation run for that flag (i.e. when the xlsx existed pre-auto but
    didn't have the column yet) — used to skip override detection on that pass.
    """
    if XLSX_PATH.exists():
        return _load_from_xlsx()
    if LEGACY_CSV_PATH.exists():
        return _load_from_csv(), False, False
    return {}, False, False


def _load_from_xlsx() -> tuple[dict[str, dict], bool, bool]:
    out: dict[str, dict] = {}
    wb = load_workbook(XLSX_PATH, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return out, False, False
    headers = [str(h) if h is not None else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}
    has_mgr_basis = "manager_change_basis" in idx
    has_ps_basis = "public_must_sell_basis" in idx
    for row in rows[1:]:
        if not row or not row[idx.get("club_id", 0)]:
            continue
        cid = str(row[idx["club_id"]])
        out[cid] = {
            "manager_change_flag":    int(row[idx["manager_change_flag"]] or 0),
            "manager_change_basis":   (row[idx["manager_change_basis"]] if has_mgr_basis else "") or "",
            "public_must_sell_flag":  int(row[idx["public_must_sell_flag"]] or 0),
            "public_must_sell_basis": (row[idx["public_must_sell_basis"]] if has_ps_basis else "") or "",
            "notes_source":           row[idx.get("notes_source", -1)] if idx.get("notes_source", -1) >= 0 else "",
            "last_reviewed":          row[idx.get("last_reviewed", -1)] if idx.get("last_reviewed", -1) >= 0 else "",
        }
    wb.close()
    return out, has_mgr_basis, has_ps_basis


def _load_from_csv() -> dict[str, dict]:
    out: dict[str, dict] = {}
    with LEGACY_CSV_PATH.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            out[row["club_id"]] = {
                "manager_change_flag":    int(row.get("manager_change", "0") or "0"),
                "manager_change_basis":   "",
                "public_must_sell_flag":  int(row.get("public_must_sell", "0") or "0"),
                "public_must_sell_basis": "",
                "notes_source":           row.get("notes", "") or "",
                "last_reviewed":          "",
            }
    return out


# ─── Auto-derivation ───────────────────────────────────────────────────────────

def derive_manager_change_auto(
    manager_name: str | None,
    manager_title: str | None,
    tenure_months: int | None,
    contract_end: date | None,
    base_pressure: float,
) -> tuple[int, str]:
    """Returns (auto_flag, auto_basis). Priority: caretaker > tenure > contract > vacancy."""
    if manager_title:
        t = manager_title.lower()
        if any(kw in t for kw in CARETAKER_KEYWORDS):
            return 1, "caretaker_title"
    if tenure_months is not None and tenure_months <= TENURE_MAX_MONTHS_FOR_FLAG:
        return 1, "tenure_<=6mo"
    if contract_end is not None and contract_end <= CONTRACT_CUTOFF:
        return 1, "contract_expiring_2026"
    if (not manager_name) and base_pressure >= VACANCY_PRESSURE_THRESHOLD:
        return 1, "vacancy_inferred"
    return 0, ""


# ─── Build the rows ────────────────────────────────────────────────────────────

def _resolve_override(
    *,
    prev_flag: int,
    prev_basis: str,
    new_auto_flag: int,
    new_auto_basis: str,
    basis_col_present: bool,
) -> tuple[int, str]:
    """Return (final_flag, final_basis) given prior state + new auto value.

    Three cases:
      1. Prior xlsx didn't have the basis column yet (migration pass) → accept
         new auto. The prior 0/0 seeds carry no override information.
      2. Prior basis was 'manual_override' → user already overrode → preserve
         prev_flag and keep basis = 'manual_override'.
      3. Otherwise, the prior auto value is inferred from prev_basis (non-empty
         = auto fired = 1; empty = auto didn't fire = 0). If prev_flag matches
         that inferred prior auto, the user didn't touch it → accept new auto.
         If prev_flag disagrees, the user has edited it → preserve as override.
    """
    if not basis_col_present:
        return new_auto_flag, new_auto_basis
    if prev_basis == "manual_override":
        return prev_flag, "manual_override"
    prior_auto_flag = 1 if prev_basis else 0
    if prev_flag == prior_auto_flag:
        return new_auto_flag, new_auto_basis
    return prev_flag, "manual_override"


def _base_pressure(cl: float | None, so: float | None, ns: float | None) -> float:
    return 0.25 * (cl or 0.0) + 0.20 * (so or 0.0) + 0.20 * (ns or 0.0)


def _table_exists(con: sqlite3.Connection, name: str) -> bool:
    return con.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?", (name,)
    ).fetchone() is not None


def build_rows(
    con: sqlite3.Connection,
    existing: dict[str, dict],
    mgr_basis_col_present_in_prior: bool,
    ps_basis_col_present_in_prior: bool,
    parachutes: dict[str, int],
) -> tuple[list[dict], dict[str, int], dict[str, int]]:
    """Builds the full set of rows for the xlsx.

    Returns (rows, manager_rule_firings, parachute_rule_firings).
    """
    today_str = str(date.today())
    snapshot = config.SNAPSHOT_DATE

    have_managers = _table_exists(con, "club_manager")
    # Pull club_pressure + manager fields (from club_manager when available, else from
    # map_club_overview for the demand-mapped leagues, else NULL).
    rows = con.execute(f"""
        SELECT cp.club_id, cp.name, cp.league_id,
               cp.contract_leverage_score, cp.squad_oversupply_score, cp.net_spend_score,
               { 'cm.manager_name' if have_managers else 'NULL' } AS mgr_name,
               { 'cm.manager_title' if have_managers else 'NULL' } AS mgr_title,
               { 'cm.appointment_date' if have_managers else 'NULL' } AS mgr_appt,
               { 'cm.tenure_months' if have_managers else 'NULL' } AS mgr_tenure,
               { 'cm.contract_end_date' if have_managers else 'NULL' } AS mgr_contract,
               (SELECT manager FROM map_club_overview mco
                 WHERE mco.club_id = cp.club_id
                 ORDER BY snapshot_date DESC LIMIT 1) AS map_manager
        FROM club_pressure cp
        { 'LEFT JOIN club_manager cm ON cm.club_id = cp.club_id' if have_managers else '' }
    """).fetchall()

    manager_firings: dict[str, int] = {
        "caretaker_title": 0, "tenure_<=6mo": 0,
        "contract_expiring_2026": 0, "vacancy_inferred": 0,
        "manual_override": 0,
    }
    parachute_firings: dict[str, int] = {
        "parachute_yr1": 0, "parachute_yr2": 0, "parachute_yr3": 0,
        "manual_override": 0,
    }

    out = []
    for r in rows:
        cid, name, lid, cl, so, ns, mgr_name, mgr_title, mgr_appt, mgr_tenure, mgr_contract, map_manager = r

        # Manager fields: club_manager > map_club_overview > empty
        final_name = mgr_name or map_manager or ""
        final_title = mgr_title or ""
        appt = date.fromisoformat(mgr_appt) if mgr_appt else None
        contract_end = date.fromisoformat(mgr_contract) if mgr_contract else None
        tenure = mgr_tenure if mgr_tenure is not None else None

        base_p = _base_pressure(cl, so, ns)

        # ─── Manager change: auto-derive + override detection ───────────────
        mgr_auto_flag, mgr_auto_basis = derive_manager_change_auto(
            final_name or None, final_title or None,
            tenure, contract_end, base_p,
        )
        ps_auto_flag, ps_auto_basis = derive_public_must_sell_auto(parachutes.get(cid))

        prev = existing.get(cid)
        if prev is None:
            mgr_final_flag, mgr_final_basis = mgr_auto_flag, mgr_auto_basis
            ps_final_flag, ps_final_basis = ps_auto_flag, ps_auto_basis
            prev_notes = ""
            prev_last_reviewed = ""
        else:
            prev_notes = prev.get("notes_source", "") or ""
            prev_last_reviewed = prev.get("last_reviewed", "") or ""

            # Override detection compares the user's stored flag against the
            # *prior* auto-derived value (inferred from the prior basis), not
            # against the new auto value. This is important: when the auto rules
            # themselves change (rule tweaks, new rules), the auto value can flip
            # without any user edit — we mustn't read that as an override.
            mgr_final_flag, mgr_final_basis = _resolve_override(
                prev_flag=int(prev.get("manager_change_flag", 0) or 0),
                prev_basis=(prev.get("manager_change_basis") or "").strip(),
                new_auto_flag=mgr_auto_flag,
                new_auto_basis=mgr_auto_basis,
                basis_col_present=mgr_basis_col_present_in_prior,
            )
            ps_final_flag, ps_final_basis = _resolve_override(
                prev_flag=int(prev.get("public_must_sell_flag", 0) or 0),
                prev_basis=(prev.get("public_must_sell_basis") or "").strip(),
                new_auto_flag=ps_auto_flag,
                new_auto_basis=ps_auto_basis,
                basis_col_present=ps_basis_col_present_in_prior,
            )

        if mgr_final_basis:
            manager_firings[mgr_final_basis] = manager_firings.get(mgr_final_basis, 0) + 1
        if ps_final_basis:
            parachute_firings[ps_final_basis] = parachute_firings.get(ps_final_basis, 0) + 1

        out.append({
            "club_id":                    cid,
            "club":                       name,
            "league":                     lid,
            "current_base_pressure_score": round(base_p, 1),
            "current_manager_name":       final_name,
            "manager_title":              final_title,
            "manager_appointment_date":   mgr_appt or "",
            "manager_tenure_months":      tenure if tenure is not None else "",
            "manager_contract_end_date":  mgr_contract or "",
            "manager_change_check":       MANAGER_CHANGE_PROMPT,
            "manager_change_flag":        mgr_final_flag,
            "manager_change_basis":       mgr_final_basis,
            "public_must_sell_check":     PUBLIC_MUST_SELL_PROMPT,
            "public_must_sell_flag":      ps_final_flag,
            "public_must_sell_basis":     ps_final_basis,
            "notes_source":               prev_notes,
            "last_reviewed":              prev_last_reviewed,
        })

    out.sort(key=lambda r: -r["current_base_pressure_score"])

    # last_reviewed: stamp today if a flag value changed vs prior xlsx. Only
    # applies once both basis columns existed (the migration guard).
    if XLSX_PATH.exists() and mgr_basis_col_present_in_prior and ps_basis_col_present_in_prior:
        for r in out:
            prev = existing.get(r["club_id"])
            if not prev:
                r["last_reviewed"] = today_str
                continue
            if (r["manager_change_flag"] != int(prev.get("manager_change_flag", 0) or 0)
                    or r["public_must_sell_flag"] != int(prev.get("public_must_sell_flag", 0) or 0)):
                r["last_reviewed"] = today_str
    return out, manager_firings, parachute_firings


# ─── Write the workbook ────────────────────────────────────────────────────────

def write_xlsx(rows: list[dict], path: Path = XLSX_PATH) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "manual_flags"

    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F3864")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = bold
        cell.fill = fill
        cell.alignment = center

    # Header comments on input columns
    ws.cell(row=1, column=COL_MANAGER_CHANGE_FLAG).comment = Comment(
        "Auto-derived by 4 rules. Flip 0↔1 to override; pipeline preserves your value "
        "and sets manager_change_basis = 'manual_override' on next refresh.", "build"
    )
    ws.cell(row=1, column=COL_PUBLIC_MUST_SELL_FLAG).comment = Comment(
        "Auto-derived from data/parachute_payments.csv. Flip 0↔1 to override; "
        "pipeline preserves your value and sets public_must_sell_basis = "
        "'manual_override' on next refresh.", "build"
    )
    ws.cell(row=1, column=COL_NOTES_SOURCE).comment = Comment(
        "Free-text note on the source/evidence for any flag you set.", "build"
    )

    user_input_fill = PatternFill("solid", fgColor="FFF2CC")  # pale yellow
    readonly_fill   = PatternFill("solid", fgColor="F2F2F2")  # pale grey
    override_fill   = PatternFill("solid", fgColor="FCE4D6")  # pale orange — flag the override basis
    wrap = Alignment(wrap_text=True, vertical="top")

    for ri, r in enumerate(rows, start=2):
        ws.cell(row=ri, column=1,  value=r["club_id"]).fill = readonly_fill
        ws.cell(row=ri, column=2,  value=r["club"]).fill = readonly_fill
        ws.cell(row=ri, column=3,  value=r["league"]).fill = readonly_fill
        c4 = ws.cell(row=ri, column=4, value=r["current_base_pressure_score"])
        c4.fill = readonly_fill
        c4.number_format = "0.0"
        ws.cell(row=ri, column=5, value=r["current_manager_name"]).fill = readonly_fill
        ws.cell(row=ri, column=6, value=r["manager_title"]).fill = readonly_fill
        ws.cell(row=ri, column=7, value=r["manager_appointment_date"]).fill = readonly_fill
        ws.cell(row=ri, column=8, value=r["manager_tenure_months"]).fill = readonly_fill
        ws.cell(row=ri, column=9, value=r["manager_contract_end_date"]).fill = readonly_fill
        c10 = ws.cell(row=ri, column=10, value=r["manager_change_check"])
        c10.fill = readonly_fill
        c10.alignment = wrap
        ws.cell(row=ri, column=COL_MANAGER_CHANGE_FLAG, value=r["manager_change_flag"]).fill = user_input_fill
        c12 = ws.cell(row=ri, column=COL_MANAGER_CHANGE_BASIS, value=r["manager_change_basis"])
        c12.fill = override_fill if r["manager_change_basis"] == "manual_override" else readonly_fill
        c13 = ws.cell(row=ri, column=13, value=r["public_must_sell_check"])
        c13.fill = readonly_fill
        c13.alignment = wrap
        ws.cell(row=ri, column=COL_PUBLIC_MUST_SELL_FLAG, value=r["public_must_sell_flag"]).fill = user_input_fill
        c15 = ws.cell(row=ri, column=COL_PUBLIC_MUST_SELL_BASIS, value=r["public_must_sell_basis"])
        c15.fill = override_fill if r["public_must_sell_basis"] == "manual_override" else readonly_fill
        ws.cell(row=ri, column=COL_NOTES_SOURCE, value=r["notes_source"]).fill = user_input_fill
        ws.cell(row=ri, column=COL_LAST_REVIEWED, value=r["last_reviewed"]).fill = readonly_fill

    # Data validation on the flag columns
    dv = DataValidation(type="list", formula1='"0,1"', allow_blank=False)
    dv.error = "Enter 0 or 1."
    dv.errorTitle = "Invalid flag value"
    mc_letter = get_column_letter(COL_MANAGER_CHANGE_FLAG)
    ps_letter = get_column_letter(COL_PUBLIC_MUST_SELL_FLAG)
    dv.add(f"{mc_letter}2:{mc_letter}{len(rows) + 1}")
    dv.add(f"{ps_letter}2:{ps_letter}{len(rows) + 1}")
    ws.add_data_validation(dv)

    for ci, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "D2"
    last_col = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"
    ws.row_dimensions[1].height = 36

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ─── Reconciliation: staged xlsx vs club_pressure ──────────────────────────────

def _read_xlsx_flags(path: Path) -> dict[str, dict]:
    """Re-read a written xlsx and return {club_id: {mgr, ps, mgr_basis, ps_basis}}.
    Only the fields needed for reconciliation."""
    out: dict[str, dict] = {}
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return out
    headers = [str(h) if h is not None else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}
    for row in rows[1:]:
        if not row or not row[idx.get("club_id", 0)]:
            continue
        cid = str(row[idx["club_id"]])
        out[cid] = {
            "manager_change_flag":    int(row[idx["manager_change_flag"]] or 0),
            "manager_change_basis":   (row[idx.get("manager_change_basis", -1)] if idx.get("manager_change_basis", -1) >= 0 else "") or "",
            "public_must_sell_flag":  int(row[idx["public_must_sell_flag"]] or 0),
            "public_must_sell_basis": (row[idx.get("public_must_sell_basis", -1)] if idx.get("public_must_sell_basis", -1) >= 0 else "") or "",
            "club_name":              row[idx["club"]],
        }
    return out


def _reconcile_with_club_pressure(
    staging_path: Path,
    con: sqlite3.Connection,
) -> tuple[bool, list[dict], bool]:
    """Diff staged-xlsx flag values against club_pressure (canonical source of truth).

    Returns (matches, diff_rows, skipped). matches=True iff every flag value in the
    staged xlsx matches the corresponding club_pressure row. skipped=True iff the
    check couldn't run (e.g. club_pressure missing/empty — first run scenario).

    diff_rows entries: dict with keys club_id, club, field, xlsx_value, db_value.
    """
    # Pull club_pressure (canonical)
    cp_rows = con.execute(
        "SELECT club_id, name, manager_change_flag, public_must_sell_flag FROM club_pressure"
    ).fetchall()
    if not cp_rows:
        return True, [], True
    db_flags = {
        str(cid): {
            "name": name,
            "manager_change_flag": int(mc or 0),
            "public_must_sell_flag": int(ps or 0),
        }
        for cid, name, mc, ps in cp_rows
    }

    staged = _read_xlsx_flags(staging_path)
    diff: list[dict] = []
    for cid, st in staged.items():
        db = db_flags.get(cid)
        if db is None:
            # Club in xlsx but not in club_pressure — skip silently; this is
            # legitimate (e.g. xlsx still has retired clubs that 08 pruned).
            continue
        for field in ("manager_change_flag", "public_must_sell_flag"):
            if st[field] != db[field]:
                diff.append({
                    "club_id":    cid,
                    "club":       st["club_name"] or db["name"],
                    "field":      field,
                    "xlsx_value": st[field],
                    "db_value":   db[field],
                })
    return (len(diff) == 0), diff, False


def _reconcile_counts(
    staging_path: Path,
    con: sqlite3.Connection,
    manager_firings: dict[str, int],
    parachute_firings: dict[str, int],
) -> list[dict]:
    """Top-line count checks. Catches a different failure mode than per-row
    matching: confirms that what the script *intended* to write is what was
    *actually* written.

    Four checks (each returns a dict if it fails):
      1. count(xlsx mgr_flag=1)  vs  count(db mgr_flag=1)
      2. count(xlsx ps_flag=1)   vs  count(db ps_flag=1)
      3. count(xlsx mgr_basis='manual_override')  vs  build_rows manual_override firings
      4. count(xlsx ps_basis='manual_override')   vs  build_rows manual_override firings

    Returns a list of mismatches (empty list = all counts agree).
    """
    cp_row = con.execute(
        "SELECT COUNT(*) FILTER (WHERE manager_change_flag=1), "
        "       COUNT(*) FILTER (WHERE public_must_sell_flag=1) "
        "FROM club_pressure"
    ).fetchone()
    db_mgr_count, db_ps_count = (cp_row[0] or 0), (cp_row[1] or 0)

    staged = _read_xlsx_flags(staging_path)
    xlsx_mgr_count = sum(1 for v in staged.values() if v["manager_change_flag"] == 1)
    xlsx_ps_count  = sum(1 for v in staged.values() if v["public_must_sell_flag"] == 1)
    xlsx_mgr_overrides = sum(1 for v in staged.values() if v["manager_change_basis"] == "manual_override")
    xlsx_ps_overrides  = sum(1 for v in staged.values() if v["public_must_sell_basis"] == "manual_override")

    expected_mgr_overrides = manager_firings.get("manual_override", 0)
    expected_ps_overrides  = parachute_firings.get("manual_override", 0)

    mismatches: list[dict] = []
    if xlsx_mgr_count != db_mgr_count:
        mismatches.append({
            "check": "manager_change_flag=1 count",
            "xlsx":  xlsx_mgr_count,
            "expected": db_mgr_count,
            "source":   "club_pressure",
        })
    if xlsx_ps_count != db_ps_count:
        mismatches.append({
            "check": "public_must_sell_flag=1 count",
            "xlsx":  xlsx_ps_count,
            "expected": db_ps_count,
            "source":   "club_pressure",
        })
    if xlsx_mgr_overrides != expected_mgr_overrides:
        mismatches.append({
            "check": "manager_change_basis='manual_override' count",
            "xlsx":  xlsx_mgr_overrides,
            "expected": expected_mgr_overrides,
            "source":   "build_rows in-memory firings",
        })
    if xlsx_ps_overrides != expected_ps_overrides:
        mismatches.append({
            "check": "public_must_sell_basis='manual_override' count",
            "xlsx":  xlsx_ps_overrides,
            "expected": expected_ps_overrides,
            "source":   "build_rows in-memory firings",
        })
    return mismatches


def _write_divergence_log(diff: list[dict], count_mismatches: list[dict]) -> Path:
    from datetime import datetime
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = RECONCILIATION_LOG_DIR / f"script_18_reconciliation_{ts}.log"
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("w", encoding="utf-8") as f:
        f.write("# Script 18 reconciliation failure: staged manual_flags.xlsx vs club_pressure\n")
        f.write(f"# Timestamp: {datetime.now().isoformat(timespec='seconds')}\n")
        f.write("# club_pressure (in db/yatin.db) is the canonical source of truth.\n")
        f.write(f"# Existing {XLSX_PATH} was NOT overwritten — prior values are intact.\n")
        f.write(f"# Rolling backup at {BAK_XLSX_PATH} (refreshed at start of every run).\n")
        f.write(f"# Staging file kept at {STAGING_XLSX_PATH} for inspection.\n")
        f.write("#\n")

        if count_mismatches:
            f.write("## Top-line count mismatches\n")
            f.write(f"# {'check':50s}  {'xlsx':>6s}  {'expected':>9s}  source\n")
            for m in count_mismatches:
                f.write(
                    f"  {m['check']:50s}  {m['xlsx']:>6}  {m['expected']:>9}  {m['source']}\n"
                )
            f.write("\n")

        if diff:
            f.write(f"## Per-row flag divergences ({len(diff)} total)\n")
            f.write(f"# {'club_id':>7}  {'club':40s}  {'field':25s}  {'xlsx':>5s}  {'db':>5s}\n")
            # Sort: ps before mgr (parachute layer was the Day 4.5 bug source), then by club
            diff_sorted = sorted(diff, key=lambda d: (d["field"] != "public_must_sell_flag", d["club"] or ""))
            for d in diff_sorted:
                club = (d["club"] or "")[:40]
                f.write(
                    f"  {d['club_id']:>7}  {club:40s}  {d['field']:25s}  "
                    f"{d['xlsx_value']:>5}  {d['db_value']:>5}\n"
                )
    return log_path


# ─── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    if not Path(config.SQLITE_FILE).exists():
        sys.exit(f"Missing {config.SQLITE_FILE} — run 02→08 first.")

    _seed_parachute_csv_if_missing()
    parachutes, parachute_src = _load_parachutes()

    # .bak rotation: snapshot the current live xlsx at the start of the run so the
    # last-known-good state is preserved regardless of how the run terminates.
    if XLSX_PATH.exists():
        shutil.copy2(XLSX_PATH, BAK_XLSX_PATH)

    existing, mgr_basis_present, ps_basis_present = _load_existing_flags()
    with sqlite3.connect(config.SQLITE_FILE) as con:
        have_managers = _table_exists(con, "club_manager")
        rows, manager_firings, parachute_firings = build_rows(
            con, existing, mgr_basis_present, ps_basis_present, parachutes,
        )

        # Write to staging, then reconcile against club_pressure before committing.
        # club_pressure is the canonical source of truth — see CLAUDE.md "Known issues".
        write_xlsx(rows, STAGING_XLSX_PATH)
        matches, diff, skipped = _reconcile_with_club_pressure(STAGING_XLSX_PATH, con)
        count_mismatches = (
            [] if skipped
            else _reconcile_counts(STAGING_XLSX_PATH, con, manager_firings, parachute_firings)
        )

    if skipped:
        print("  Reconciliation skipped — club_pressure empty (first run / fresh DB).")
        # No .bak on first run — there is no prior good state to back up.
        STAGING_XLSX_PATH.replace(XLSX_PATH)
    elif not matches or count_mismatches:
        log_path = _write_divergence_log(diff, count_mismatches)
        bar = "=" * 78
        print()
        print(bar)
        print("\033[1;31m⚠  RECONCILIATION FAILED — staged xlsx diverges from club_pressure\033[0m")
        print(bar)
        if count_mismatches:
            print(f"   Count mismatches:    {len(count_mismatches)}")
            for m in count_mismatches:
                print(f"     • {m['check']}: xlsx={m['xlsx']}, expected={m['expected']} ({m['source']})")
        if diff:
            print(f"   Per-row divergences: {len(diff)}")
        print(f"   Log:              {log_path.resolve()}")
        print(f"   Live xlsx:        {XLSX_PATH.resolve()}  (UNCHANGED — your previous values are safe)")
        if BAK_XLSX_PATH.exists():
            print(f"   Rolling backup:   {BAK_XLSX_PATH.resolve()}  (last-known-good)")
        print(f"   Staging file:     {STAGING_XLSX_PATH.resolve()}  (kept for inspection)")
        print()
        print("   This guard catches the Day 4.5 bug where script 18 silently stripped 44")
        print("   user-edited manual_override rows. The live xlsx was NOT overwritten.")
        print()
        print("   Next steps:")
        print("     Inspect the log file, then re-run; the canonical source of truth is")
        print("     club_pressure in db/yatin.db. If divergence persists after a clean")
        print("     08 → 18 run, investigate before trusting either side.")
        print(bar)
        sys.exit(2)
    else:
        # All checks passed. .bak already captured the prior live state at the
        # start of the run; just atomically promote staging into place.
        STAGING_XLSX_PATH.replace(XLSX_PATH)

    mgr_flagged = sum(1 for r in rows if r["manager_change_flag"] == 1)
    ps_flagged  = sum(1 for r in rows if r["public_must_sell_flag"] == 1)

    print(f"Wrote {XLSX_PATH.resolve()}")
    print(f"  {len(rows)} clubs, sorted by base pressure descending")
    print(f"  manager source: {'club_manager (scraped)' if have_managers else 'map_club_overview only'}")
    src_path = PARACHUTE_XLSX if parachute_src == "xlsx" else PARACHUTE_CSV
    print(f"  parachute source: {src_path} [{parachute_src}] ({len(parachutes)} club(s))")
    print()
    print("Manager-change rule firings:")
    print(f"  {'caretaker_title':24s} : {manager_firings.get('caretaker_title', 0):>4d}")
    print(f"  {'tenure_<=6mo':24s} : {manager_firings.get('tenure_<=6mo', 0):>4d}")
    print(f"  {'contract_expiring_2026':24s} : {manager_firings.get('contract_expiring_2026', 0):>4d}")
    print(f"  {'vacancy_inferred':24s} : {manager_firings.get('vacancy_inferred', 0):>4d}")
    print(f"  {'manual_override':24s} : {manager_firings.get('manual_override', 0):>4d}  (preserved across refresh)")
    print(f"  {'TOTAL auto-flagged':24s} : {mgr_flagged} / {len(rows)} clubs")
    print()
    print("Public-must-sell (parachute) rule firings:")
    print(f"  {'parachute_yr1':24s} : {parachute_firings.get('parachute_yr1', 0):>4d}")
    print(f"  {'parachute_yr2':24s} : {parachute_firings.get('parachute_yr2', 0):>4d}")
    print(f"  {'parachute_yr3':24s} : {parachute_firings.get('parachute_yr3', 0):>4d}")
    print(f"  {'manual_override':24s} : {parachute_firings.get('manual_override', 0):>4d}  (preserved across refresh)")
    print(f"  {'TOTAL auto-flagged':24s} : {ps_flagged} / {len(rows)} clubs")
    print()
    print("Top 30 manager-change clubs by current_base_pressure_score:")
    print(f"  {'club':36s} {'lg':5s} {'base':>6s}  {'basis':22s}  {'manager':28s}  title")
    n_shown = 0
    for r in rows:
        if r["manager_change_flag"] != 1:
            continue
        print(f"  {(r['club'] or '')[:36]:36s} {r['league']:5s} "
              f"{r['current_base_pressure_score']:>6.1f}  "
              f"{(r['manager_change_basis'] or ''):22s}  "
              f"{(r['current_manager_name'] or '(vacant)')[:28]:28s}  "
              f"{r['manager_title']}")
        n_shown += 1
        if n_shown >= 30:
            break

    print()
    print("Top 20 public-must-sell (parachute-derived) clubs by current_base_pressure_score:")
    print(f"  {'club':36s} {'lg':5s} {'base':>6s}  {'basis':16s}")
    n_shown = 0
    for r in rows:
        if r["public_must_sell_flag"] != 1:
            continue
        print(f"  {(r['club'] or '')[:36]:36s} {r['league']:5s} "
              f"{r['current_base_pressure_score']:>6.1f}  "
              f"{(r['public_must_sell_basis'] or ''):16s}")
        n_shown += 1
        if n_shown >= 20:
            break

    if LEGACY_CSV_PATH.exists():
        print()
        print(f"Legacy CSV still at {LEGACY_CSV_PATH.resolve()} — safe to delete after verifying xlsx.")


if __name__ == "__main__":
    main()
