"""
Load `data/scisports_ratings.xlsx` into the SQLite `player_ratings` table.

Schema (player_ratings):
  tm_player_id      INTEGER PRIMARY KEY
  current_ability   REAL    -- NULL when status=pending and CA blank
  potential_ability REAL    -- NULL when status=pending and PA blank
  status            TEXT    -- pending / active / departed / killed
  last_updated      TEXT    -- ISO date

UI consumers filter on status — rows with departed/killed are kept in the
table so we don't lose historic ratings on a transient cohort drop-out, but
the Streamlit pages can choose whether to surface them.

Run order in the refresh pipeline:
  ... → 09_compute_sellability → 12_patch_parent_contract →
  reconcile_scisports → load_scisports_ratings → 22_match_engine → ...

Idempotent: safe to re-run. DROPs and recreates the player_ratings table on
every run so it always mirrors the workbook exactly.
"""

from __future__ import annotations

import sqlite3
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import config

FILE_PATH = Path(config.DATA_DIR) / "scisports_ratings.xlsx"


def _to_float(v) -> float | None:
    """Parse a cell value into a float; return None for blank / "None" /
    non-numeric strings (so pending rows with empty CA/PA load as NULL)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s.lower() in ("none", "nan", "n/a"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


# Bounds for the Sci Sports CA / PA scale (0-200 per current convention).
# Values outside the range, or PA < CA, get flagged status=invalid and the
# offending row is logged with its tm_player_id + name. The whole load
# doesn't abort — invalid rows just get sequestered so the rest still flows
# to downstream consumers.
CA_MIN, CA_MAX = 0.0, 200.0
PA_MIN, PA_MAX = 0.0, 200.0


def _validate(ca: float | None, pa: float | None) -> str | None:
    """Return a validation error string, or None if the row is valid."""
    if ca is not None and not (CA_MIN <= ca <= CA_MAX):
        return f"current_ability {ca} outside [{CA_MIN}, {CA_MAX}]"
    if pa is not None and not (PA_MIN <= pa <= PA_MAX):
        return f"potential_ability {pa} outside [{PA_MIN}, {PA_MAX}]"
    if ca is not None and pa is not None and pa < ca:
        return f"potential_ability {pa} < current_ability {ca}"
    return None


def read_workbook(file_path: Path) -> tuple[list[dict], list[dict]]:
    """Return (valid_rows, invalid_rows). Invalid rows fail bounds-check and
    get status=invalid; they're still loaded into the table but logged so
    the user can fix the source data."""
    if not file_path.exists():
        raise FileNotFoundError(
            f"{file_path} not found. Run "
            "`python scripts/reconcile_scisports.py` first to create it."
        )
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = list(rows[0])
    valid: list[dict] = []
    invalid: list[dict] = []
    for raw in rows[1:]:
        if not raw or raw[0] is None:
            continue
        d = dict(zip(headers, raw))
        try:
            pid = int(d.get("tm_player_id"))
        except (TypeError, ValueError):
            continue
        ca = _to_float(d.get("current_ability"))
        pa = _to_float(d.get("potential_ability"))
        status = (str(d.get("status") or "").strip().lower() or "pending")
        last_updated = str(d.get("last_updated") or "")
        player_name = str(d.get("player_name") or "")

        error = _validate(ca, pa)
        if error:
            invalid.append({
                "tm_player_id":      pid,
                "player_name":       player_name,
                "current_ability":   ca,
                "potential_ability": pa,
                "status":            "invalid",
                "last_updated":      last_updated,
                "error":             error,
            })
        else:
            valid.append({
                "tm_player_id":      pid,
                "current_ability":   ca,
                "potential_ability": pa,
                "status":            status,
                "last_updated":      last_updated,
            })
    return valid, invalid


def main() -> None:
    valid, invalid = read_workbook(FILE_PATH)
    combined = valid + [
        {
            "tm_player_id":      r["tm_player_id"],
            "current_ability":   r["current_ability"],
            "potential_ability": r["potential_ability"],
            "status":            "invalid",
            "last_updated":      r["last_updated"],
        }
        for r in invalid
    ]

    con = sqlite3.connect(config.SQLITE_FILE)
    try:
        con.execute("DROP TABLE IF EXISTS player_ratings")
        con.execute("""
            CREATE TABLE player_ratings (
                tm_player_id      INTEGER PRIMARY KEY,
                current_ability   REAL,
                potential_ability REAL,
                status            TEXT,
                last_updated      TEXT
            )
        """)
        con.executemany(
            "INSERT INTO player_ratings "
            "(tm_player_id, current_ability, potential_ability, status, last_updated) "
            "VALUES (?, ?, ?, ?, ?)",
            [
                (r["tm_player_id"], r["current_ability"], r["potential_ability"],
                 r["status"], r["last_updated"])
                for r in combined
            ],
        )
        con.commit()

        # Summary counts
        counts = dict(con.execute(
            "SELECT status, COUNT(*) FROM player_ratings GROUP BY status"
        ).fetchall())
        n_rated = con.execute(
            "SELECT COUNT(*) FROM player_ratings "
            "WHERE current_ability IS NOT NULL OR potential_ability IS NOT NULL"
        ).fetchone()[0]
    finally:
        con.close()

    print(
        f"Loaded player_ratings — {sum(counts.values())} total rows: "
        + ", ".join(f"{k}={v}" for k, v in sorted(counts.items()))
        + f"; {n_rated} rows with CA/PA filled in."
    )

    # Log invalid rows so the user can fix the source data
    if invalid:
        print()
        print(f"⚠️  {len(invalid)} row(s) failed bounds-check "
              f"(CA in [0,200], PA in [0,200], PA ≥ CA):")
        for r in invalid:
            print(f"  pid={r['tm_player_id']:>9d}  {r['player_name'][:30]:<30}  "
                  f"CA={r['current_ability']}  PA={r['potential_ability']}  → {r['error']}")


if __name__ == "__main__":
    main()
