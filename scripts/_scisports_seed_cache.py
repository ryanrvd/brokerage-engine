"""Seed the SciSports cache from existing manual CA/PA so Phase 3 doesn't
re-call the SciSkill endpoint for ~120 players Ryan has already rated by hand.

Pipeline:
  1) Read data/scisports_ratings.xlsx — every row with CA + PA filled in.
  2) Read player_universe from db/yatin.db for tm_player_id → (name, DOB).
  3) Read data/scisports_pl_player_roster.json for the bridge (name + birthDate
     → scisports_player_id).
  4) Fuzzy-match: each xlsx entry → tm row → scisports row.
  5) For each match, write a cache entry keyed by the exact GET call that
     script 30 will make in Phase 3:
        /api/v2/metrics/players/sciskill?PlayerIds=<sci_id>&Offset=0&Limit=1
     ...with source='manual_seed' and TTL=7 days.

The cache key must match what _scisports_client.get_player_sciskills(
PlayerIds=..., limit=1) produces, so script 30 sees a hit and skips network.
"""
from __future__ import annotations

import json
import re
import sqlite3
import sys
import unicodedata
from datetime import date
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import load_workbook

import config
from _scisports_cache import TTL_SCISKILL, cache_path, write_cached

XLSX_PATH = PROJECT_ROOT / "data" / "scisports_ratings.xlsx"
ROSTER_PATH = PROJECT_ROOT / "data" / "scisports_pl_player_roster.json"

SCISKILL_ENDPOINT = "/api/v2/metrics/players/sciskill"


def _normalize(s: str | None) -> str:
    if not s:
        return ""
    # Decompose accents, drop combining marks, lowercase, drop punctuation
    nfkd = unicodedata.normalize("NFKD", s)
    no_accents = "".join(c for c in nfkd if not unicodedata.combining(c))
    cleaned = re.sub(r"[^\w\s]", " ", no_accents.lower())
    return re.sub(r"\s+", " ", cleaned).strip()


def _parse_dob(v) -> str | None:
    """Parse various DOB formats into ISO YYYY-MM-DD."""
    if not v:
        return None
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    if "T" in s:
        s = s.split("T", 1)[0]
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        return s
    return None


def load_xlsx_entries() -> list[dict]:
    """Return [{tm_player_id, ca, pa, name, status}] for every rated row."""
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = list(rows[0])
    out: list[dict] = []
    ci = {h: i for i, h in enumerate(headers)}
    for r in rows[1:]:
        if not r or r[ci["tm_player_id"]] is None:
            continue
        ca = r[ci["current_ability"]]
        pa = r[ci["potential_ability"]]
        if ca in (None, "") or pa in (None, ""):
            continue
        try:
            ca = float(ca); pa = float(pa)
        except (TypeError, ValueError):
            continue
        try:
            pid = int(r[ci["tm_player_id"]])
        except (TypeError, ValueError):
            continue
        out.append({
            "tm_player_id": pid,
            "ca": ca,
            "pa": pa,
            "name": str(r[ci["player_name"]] or ""),
            "status": str(r[ci.get("status", -1)] or "").strip().lower() if "status" in ci else "",
        })
    return out


def load_player_universe_index(con: sqlite3.Connection) -> dict[int, dict]:
    """Return tm_player_id -> {name, dob, parent_club}."""
    out: dict[int, dict] = {}
    for pid, name, dob, parent in con.execute("""
        SELECT player_id, name, date_of_birth, parent_club FROM player_universe
    """):
        out[int(pid)] = {
            "name": name or "",
            "dob": dob or None,
            "parent_club": parent or "",
        }
    return out


def load_roster_bridge() -> list[dict]:
    """Return a flat list of {scisports_player_id, name_variants, birthDate, team}."""
    raw = json.loads(ROSTER_PATH.read_text(encoding="utf-8"))
    players = raw.get("players", [])
    bridge: list[dict] = []
    for p in players:
        names = []
        for k in ("name", "firstName", "lastName", "footballName"):
            v = p.get(k)
            if v:
                names.append(v)
        # Also include "firstName lastName" combination
        if p.get("firstName") and p.get("lastName"):
            names.append(f"{p['firstName']} {p['lastName']}")
        bridge.append({
            "scisports_player_id": p.get("scisports_player_id"),
            "names":     [n for n in names if n],
            "names_norm": [_normalize(n) for n in names if n],
            "birthDate": _parse_dob(p.get("birthDate")),
            "team_name": (p.get("team") or {}).get("name"),
        })
    return bridge


def match_one(tm_name: str, tm_dob: str | None, bridge: list[dict]) -> tuple[int | None, str]:
    """Match a single TM player to a SciSports roster entry — STRICT.

    DOB is REQUIRED. Pure name-only fallback was removed after a Phase 3
    audit revealed name collisions: e.g. 'Juanlu Sánchez' (Sevilla, manually
    rated) → matched to 'Robert Sánchez' (Chelsea GK), 'Julio Enciso' →
    'Julio Soler'. Surfacing those collisions cost a round-trip of bad data
    into player_ratings, then a manual repair.

    Returns (scisports_player_id, match_kind) or (None, '').
    match_kind: 'name+dob' (strict) or 'name+year' (dob-year + unique
    name-token match).
    """
    nm = _normalize(tm_name)
    if not nm:
        return None, ""
    tokens = set(nm.split())
    if not tokens:
        return None, ""
    if not tm_dob:
        return None, "no_dob"
    tm_year = tm_dob[:4] if len(tm_dob) >= 4 else None

    # Pass 1: exact name + exact DOB
    hits = []
    for b in bridge:
        if any(nm == n for n in b["names_norm"]) and b["birthDate"] == tm_dob:
            hits.append(b)
    if len(hits) == 1:
        return hits[0]["scisports_player_id"], "name+dob"
    if len(hits) > 1:
        return None, "ambiguous"

    # Pass 2: token-subset name + exact DOB year (still narrowed by DOB)
    if tm_year:
        candidates = []
        for b in bridge:
            if not (b["birthDate"] and b["birthDate"].startswith(tm_year)):
                continue
            for n_norm in b["names_norm"]:
                sci_tokens = set(n_norm.split())
                if tokens.issubset(sci_tokens) or sci_tokens.issubset(tokens):
                    candidates.append(b)
                    break
        if len(candidates) == 1:
            return candidates[0]["scisports_player_id"], "name+year"
        if len(candidates) > 1:
            return None, "ambiguous"

    # Intentionally NO pass 3 — name-only matches are too lossy across the
    # PL roster (see Phase 3 audit). Unmatched players are logged as a
    # follow-up cohort; better to under-seed than to corrupt.
    return None, ""


def make_sciskill_response(sci_id: int, ca: float, pa: float) -> dict:
    """Synthesise the response shape the client.get_player_sciskills call
    would return for one player."""
    return {
        "items": [{
            "playerId":          sci_id,
            "currentSciSkill":   ca,
            "potentialSciSkill": pa,
            # Marker so downstream consumers know this is seeded.
            "_seeded":           True,
        }],
        "total": 1,
        "offset": 0,
        "limit": 1,
    }


def main() -> None:
    entries = load_xlsx_entries()
    print(f"xlsx rated rows: {len(entries)}")
    if not entries:
        print("  Nothing to seed.")
        return

    con = sqlite3.connect(config.SQLITE_FILE)
    try:
        tm_index = load_player_universe_index(con)
    finally:
        con.close()
    print(f"player_universe rows: {len(tm_index)}")

    bridge = load_roster_bridge()
    print(f"SciSports PL roster: {len(bridge)} players")
    print()

    matched = 0
    by_kind: dict[str, int] = {}
    seeded_paths: list[Path] = []
    unmatched_names: list[str] = []

    for e in entries:
        tm = tm_index.get(e["tm_player_id"])
        tm_name = (tm or {}).get("name") or e["name"]
        tm_dob = (tm or {}).get("dob")
        sci_id, kind = match_one(tm_name, tm_dob, bridge)
        if sci_id is None:
            unmatched_names.append(f"{e['tm_player_id']:>9d}  {tm_name}")
            continue
        # Write cache entry — shape MUST match how script 30 will call
        # client.get_player_sciskills(PlayerIds=<sci_id>, limit=1). The client
        # uses Offset/Limit as PascalCase params, plus the extra filter.
        params = {"Offset": 0, "Limit": 1, "PlayerIds": sci_id}
        synth = make_sciskill_response(sci_id, e["ca"], e["pa"])
        path = write_cached(
            SCISKILL_ENDPOINT, params, synth,
            ttl_seconds=TTL_SCISKILL, source="manual_seed",
        )
        seeded_paths.append(path)
        matched += 1
        by_kind[kind] = by_kind.get(kind, 0) + 1

    print(f"Seeded {matched} cache entries (source=manual_seed, TTL=7d)")
    print()
    print("Match-kind breakdown:")
    for k, n in sorted(by_kind.items()):
        print(f"  {k:20s} {n}")
    if unmatched_names:
        print()
        print(f"{len(unmatched_names)} unmatched rated players (no PL roster hit "
              "— likely not in 25/26 PL squad):")
        for line in unmatched_names[:15]:
            print(f"  {line}")
        if len(unmatched_names) > 15:
            print(f"  … and {len(unmatched_names) - 15} more")


if __name__ == "__main__":
    main()
