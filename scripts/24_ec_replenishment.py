"""
Step 24 — Write Replenishment_Leads_Workbook.xlsx.

Internal supply-side companion to Sheet 1 (Brokerage Opportunities). Surfaces
the upcoming back-fill demand: clubs about to sell a sellable asset, and what
they'll need to replace. No commission / sell-on / lifecycle math is surfaced
anywhere; commercial pricing judgement is applied on review.

Sheet A — Replenishment Summary
    Aggregate counts by selling-club league, by position bucket, and by
    combined seller-pressure score band (0-30 / 30-60 / 60+). One-glance
    picture of where back-fill demand concentrates.

Sheet B — Replenishment Leads
    One row per (selling_club, player_id) pair on Sheet 1. The selling club is
    the player's parent_club (legal owner — they're the one collecting the
    sale proceeds and able to spend on a replacement).

Idempotent. Overwrites the workbook on every run.
"""

import sqlite3
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import config
import club_display as cd

OUTPUT_PATH = Path("Replenishment_Leads_Workbook.xlsx")
WORKBOOK_FOR_DISPLAY = Path("BrokerageWorkbook.xlsx")

PRESSURE_BANDS = [
    ("0-30",  0,    30),
    ("30-60", 30.001, 60),
    ("60+",   60.001, 200),
]

SHEET_B_HEADERS = [
    "selling_club", "league", "parent_club", "position_now_open",
    "max_transfer_fee_eur (replacement-budget proxy)",
    "max_wage_pw_eur", "preferred_agencies", "priority_flag",
    "candidate_from_portfolio",
]
SHEET_B_WIDTHS = [34, 8, 28, 18, 36, 16, 36, 14, 28]


# ─── Sheet A: summary aggregates ──────────────────────────────────────────────

def _band_for(score: float | None) -> str:
    if score is None:
        return "unknown"
    for label, lo, hi in PRESSURE_BANDS:
        if lo <= score <= hi:
            return label
    return "unknown"


def build_sheet_a(con: sqlite3.Connection) -> dict:
    """Returns a dict of three aggregations:
       by_league: {league_id: count}
       by_position: {position_bucket: count}
       by_pressure_band: {'0-30': count, ...}
    """
    rows = con.execute("""
        SELECT pu.player_id, pu.position_bucket, cp.league_id AS parent_league,
               cp.total_pressure_score
        FROM matches m
        JOIN player_universe pu ON pu.player_id = m.player_id
        LEFT JOIN club_pressure cp ON cp.club_id = pu.parent_club_id
        GROUP BY pu.player_id
    """).fetchall()
    by_league: dict[str, int] = {}
    by_position: dict[str, int] = {}
    by_band: dict[str, int] = {label: 0 for label, _, _ in PRESSURE_BANDS}
    by_band["unknown"] = 0
    for _, bucket, league, score in rows:
        if league:
            by_league[league] = by_league.get(league, 0) + 1
        by_position[bucket] = by_position.get(bucket, 0) + 1
        by_band[_band_for(score)] = by_band.get(_band_for(score), 0) + 1
    return {"by_league": by_league, "by_position": by_position, "by_band": by_band}


def write_sheet_a(wb: Workbook, agg: dict) -> None:
    ws = wb.active
    ws.title = "Replenishment Summary"
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F3864")
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)
    section_fill = PatternFill("solid", fgColor="D9E1F2")
    section_font = Font(bold=True)

    def section(row: int, title: str) -> int:
        c = ws.cell(row=row, column=1, value=title)
        c.font = section_font
        c.fill = section_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        return row + 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 6

    # Headers
    ws.cell(row=1, column=1, value="Replenishment leads — distribution").font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

    r = 3
    r = section(r, "By selling-club league")
    for k in sorted(agg["by_league"].keys()):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=agg["by_league"][k])
        r += 1
    r += 1
    r = section(r, "By position bucket needed")
    for k in ("GK","CB","LB","RB","DM","CM","AM","LW","RW","ST_CF"):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=agg["by_position"].get(k, 0))
        r += 1
    r += 1
    r = section(r, "By total_pressure_score band")
    for label in ["0-30", "30-60", "60+", "unknown"]:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=agg["by_band"].get(label, 0))
        r += 1

    # Header colour band
    for col in range(1, 4):
        c = ws.cell(row=2, column=col)
        c.fill = fill
        c.font = bold
        c.alignment = centre
    ws.cell(row=2, column=1, value="category")
    ws.cell(row=2, column=2, value="count")


# ─── Sheet B: per-(selling_club, player) leads ────────────────────────────────

def build_sheet_b(con: sqlite3.Connection) -> list[dict]:
    """Pull one row per (selling_club, player_id) pair where the player appears
    in `matches`. selling_club = parent_club (legal owner). priority_flag is set
    if the player's best match_score is in the top 20% globally.
    """
    # Compute the top-20% match_score threshold.
    scores = [r[0] for r in con.execute("SELECT match_score FROM matches WHERE match_score IS NOT NULL")]
    if not scores:
        return []
    scores.sort(reverse=True)
    cutoff_idx = max(0, int(len(scores) * 0.20) - 1)
    top_20_threshold = scores[cutoff_idx] if scores else 0.0

    rows = con.execute("""
        SELECT pu.player_id, pu.name, pu.position_bucket, pu.parent_club,
               pu.parent_club_id, pu.current_club, pu.on_loan,
               pu.last_fee_paid_eur, cp.league_id AS selling_league,
               MAX(m.match_score) AS best_match_score,
               mco.agent_preferences
        FROM matches m
        JOIN player_universe pu ON pu.player_id = m.player_id
        LEFT JOIN club_pressure cp ON cp.club_id = pu.parent_club_id
        LEFT JOIN map_club_overview mco ON mco.club_id = pu.parent_club_id
        GROUP BY pu.player_id
        ORDER BY best_match_score DESC, pu.name
    """).fetchall()
    cols = [d[0] for d in con.execute(
        "SELECT pu.player_id, pu.name, pu.position_bucket, pu.parent_club, "
        "pu.parent_club_id, pu.current_club, pu.on_loan, "
        "pu.last_fee_paid_eur, cp.league_id AS selling_league, "
        "MAX(m.match_score) AS best_match_score, mco.agent_preferences "
        "FROM matches m JOIN player_universe pu ON pu.player_id = m.player_id "
        "LEFT JOIN club_pressure cp ON cp.club_id = pu.parent_club_id "
        "LEFT JOIN map_club_overview mco ON mco.club_id = pu.parent_club_id LIMIT 0"
    ).description]
    leads: list[dict] = []
    for r in rows:
        d = dict(zip(cols, r))
        d["priority_flag"] = 1 if (d["best_match_score"] or 0) >= top_20_threshold else 0
        leads.append(d)
    return leads


def write_sheet_b(wb: Workbook, leads: list[dict]) -> int:
    ws = wb.create_sheet("Replenishment Leads", 1)
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F3864")
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, h in enumerate(SHEET_B_HEADERS, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = bold
        c.fill = fill
        c.alignment = centre
    for ci, w in enumerate(SHEET_B_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"

    money = '"€"#,##0;[Red]"€"-#,##0'
    priority_fill = PatternFill("solid", fgColor="FCE4D6")
    placeholder_fill = PatternFill("solid", fgColor="F2F2F2")
    display_map = cd.load_display_map(WORKBOOK_FOR_DISPLAY)
    for ri, d in enumerate(leads, start=2):
        # selling_club = parent_club (legal owner); show current_club as 'parent_club' context column
        # ONLY when the player is on loan (otherwise both are the same and the column is redundant).
        selling_display = cd.display_for(d.get("parent_club_id"), d["parent_club"], display_map)
        loan_current_display = (
            cd.display_for(d.get("current_club_id") if "current_club_id" in d else None, d["current_club"], display_map)
            if d["on_loan"] else ""
        )
        selling_league_display = config.LEAGUE_DISPLAY.get(d["selling_league"], d["selling_league"]) if d["selling_league"] else "—"
        ws.cell(row=ri, column=1, value=selling_display)
        ws.cell(row=ri, column=2, value=selling_league_display)
        ws.cell(row=ri, column=3, value=loan_current_display)
        ws.cell(row=ri, column=4, value=d["position_bucket"])
        c = ws.cell(row=ri, column=5, value=d["last_fee_paid_eur"]); c.number_format = money
        c = ws.cell(row=ri, column=6, value=None); c.number_format = money  # no wage data
        ws.cell(row=ri, column=7, value=d.get("agent_preferences") or "—")
        c = ws.cell(row=ri, column=8, value=d["priority_flag"])
        if d["priority_flag"] == 1:
            c.fill = priority_fill
        cp = ws.cell(row=ri, column=9, value=None)
        cp.fill = placeholder_fill
    last_col = get_column_letter(len(SHEET_B_HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{len(leads) + 1}"
    return len(leads)


# ─── Main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    if not Path(config.SQLITE_FILE).exists():
        sys.exit(f"Missing {config.SQLITE_FILE} — run 02→22 first.")

    wb = Workbook()
    with sqlite3.connect(config.SQLITE_FILE) as con:
        agg = build_sheet_a(con)
        write_sheet_a(wb, agg)
        leads = build_sheet_b(con)
        n_leads = write_sheet_b(wb, leads)
    wb.save(OUTPUT_PATH)

    print(f"Wrote {OUTPUT_PATH.resolve()}")
    print(f"  Sheet A 'Replenishment Summary' — {len(agg['by_league'])} leagues, "
          f"{len(agg['by_position'])} positions, {len([k for k,v in agg['by_band'].items() if v>0])} band(s) populated")
    print(f"  Sheet B 'Replenishment Leads'   — {n_leads} (selling_club, player) rows")


if __name__ == "__main__":
    main()
