"""
Step 25 — Zero-Match Diagnostic sheet.

For every sellable player who landed zero rows on Sheet 1, walk the matcher's
filter cascade and label the precise failure stage + reason. The output is a
new sheet "Zero-Match Diagnostic" in BrokerageWorkbook.xlsx so Ryan can review
why each orphaned player was suppressed.

Failure stages (in priority order):
  1. no_position_demand     — no buyer request exists at this position bucket
  2. all_buyers_below_min   — buyers exist but none clear MIN_BROKERAGE_FEE (€15m)
  3. side_preference_excl   — all budget-feasible buyers fail side preference
  4. tier_rule_exclusive    — all side-feasible buyers fail league-tier (downward moves)
  5. score_floor            — pairs survived all four filters but best score < 10

Within stage 5, the reason is decomposed: low sellability, indicative fee
exceeds buyer budgets, parent-outside-coverage, or a combination.

Idempotent. Replaces the "Zero-Match Diagnostic" sheet on every run.
"""

import sqlite3
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import config
import club_display as cd
import player_display as pd_disp

WORKBOOK_PATH = Path("BrokerageWorkbook.xlsx")
SHEET_NAME = "Zero-Match Diagnostic"

# Mirror the matcher's logic exactly so the diagnostic matches reality.
INTRINSIC_SIDE = {"LB": "Left", "RB": "Right", "LW": "Left", "RW": "Right"}
WAGE_FEASIBILITY = 0.7
MATCH_SCORE_FLOOR = 10.0
LOW_SELLABILITY_THRESHOLD = 30.0


def demand_intensity(source, validated):
    src = (source or "").strip()
    val = (validated or "").strip().upper()
    if src == "Agent" and val == "YES":
        return 1.00
    if src == "Agent":
        return 0.85
    if src == "Intel" and val == "NO":
        return 0.60
    if src == "Inferred":
        return 0.40
    return 0.50


def side_ok(player_bucket, preferred_side):
    if not preferred_side or preferred_side.strip() == "" or preferred_side == "Either":
        return True
    player_side = INTRINSIC_SIDE.get(player_bucket)
    if player_side is None:
        return True
    return player_side == preferred_side


def league_move_allowed(player_league, buyer_league):
    p_tier = config.LEAGUE_TIERS.get(player_league)
    b_tier = config.LEAGUE_TIERS.get(buyer_league)
    if p_tier is None or b_tier is None:
        return False, ""
    if p_tier == 4 and b_tier != 4:
        return False, ""
    if b_tier == 4 and p_tier != 4:
        return False, ""
    if b_tier > p_tier:
        return False, ""
    return True, ("upward" if b_tier < p_tier else "lateral")


def budget_fit_curve(indicative_fee, buyer_max_fee):
    if buyer_max_fee < config.MIN_BROKERAGE_FEE:
        return 0.0
    threshold = 2.0 * indicative_fee
    if buyer_max_fee >= threshold:
        return 0.80
    span = threshold - config.MIN_BROKERAGE_FEE
    if span <= 0:
        return 0.80
    return 0.80 * (buyer_max_fee - config.MIN_BROKERAGE_FEE) / span


def score_pair(player_tm, sellability, buyer_max_fee, source, validated):
    sell_term = (sellability or 0.0) / 100.0
    di = demand_intensity(source, validated)
    indicative_fee = player_tm * config.tm_to_fee_multiplier(player_tm)
    bf = budget_fit_curve(indicative_fee, float(buyer_max_fee))
    raw = sell_term * di * bf * WAGE_FEASIBILITY
    return raw * 100.0, bf, di, indicative_fee


# ─── Build the cascade per player ─────────────────────────────────────────────

def diagnose_player(con, player) -> dict:
    """Returns a row dict ready for the sheet."""
    pid = player["player_id"]
    bucket = player["position_bucket"]
    player_tm = player["current_tm_value_eur"]
    sell = player["sellability_score"] or 0.0
    player_lg = player["league_id"]

    candidates = list(con.execute("""
        SELECT request_id, club_id, club_name, league, position_bucket, preferred_side,
               max_transfer_fee_eur, source, validated
        FROM map_club_requests
        WHERE position_bucket = ? AND max_transfer_fee_eur IS NOT NULL
        UNION ALL
        SELECT request_id, club_id, club_name, league, position_bucket, preferred_side,
               max_transfer_fee_eur, source, validated
        FROM inferred_club_requests
        WHERE position_bucket = ? AND max_transfer_fee_eur IS NOT NULL
    """, (bucket, bucket)))

    def base():
        return {
            "name":             player["name"],
            "age":              player["age"],
            "position":         bucket,
            "current_club":     player.get("current_club_display") or player["current_club"],
            "parent_club":      player.get("parent_club_display") or player["parent_club"],
            "parent_league":    config.LEAGUE_DISPLAY.get(player.get("parent_league") or player_lg, player.get("parent_league") or player_lg),
            "tm_value_eur":     player_tm,
            "sellability":      round(sell, 1),
            "right_priced":     player["right_priced"],
            "finished_product": player["finished_product"],
            "contract_leveraged": player["contract_leveraged"],
            "on_loan":          player["on_loan"],
            "failure_stage":    "",
            "reason_summary":   "",
            "reason_detail":    "",
            "best_buyer":       "",
            "best_possible_score": 0.0,
        }

    out = base()

    # Stage 1: no demand at all for this bucket
    if not candidates:
        out["failure_stage"]  = "no_position_demand"
        out["reason_summary"] = f"No buyer request for {bucket} anywhere"
        out["reason_detail"]  = (
            f"Zero requests across explicit (map_club_requests) and inferred "
            f"(senior_roster thinness) for position {bucket}. Player can only match "
            f"once we receive demand data for this bucket."
        )
        return out

    # Stage 2: budget filter (max_fee ≥ MIN_BROKERAGE_FEE)
    after_budget = [c for c in candidates if c["max_transfer_fee_eur"] >= config.MIN_BROKERAGE_FEE]
    if not after_budget:
        max_observed = max(c["max_transfer_fee_eur"] for c in candidates)
        out["failure_stage"]  = "all_buyers_below_min"
        out["reason_summary"] = f"All {len(candidates)} {bucket} buyers below €15m floor"
        out["reason_detail"]  = (
            f"{len(candidates)} request(s) at {bucket} exist, but none demonstrate "
            f"≥€15m spending intent (top observed max_fee = €{max_observed/1e6:.1f}m). "
            "MIN_BROKERAGE_FEE filter excludes them."
        )
        return out

    # Stage 3: side preference
    after_side = [c for c in after_budget if side_ok(bucket, c["preferred_side"])]
    if not after_side:
        sides = sorted({c["preferred_side"] for c in after_budget if c["preferred_side"]})
        out["failure_stage"]  = "side_preference_excl"
        out["reason_summary"] = f"Side preference excludes all {len(after_budget)} buyers"
        out["reason_detail"]  = (
            f"All {len(after_budget)} budget-feasible {bucket} buyers specify "
            f"side(s) {sides} incompatible with this player. (Side enforced for "
            f"LB/RB/LW/RW positions.)"
        )
        return out

    # Stage 4: league tier
    after_tier = []
    for c in after_side:
        ok, _ = league_move_allowed(player_lg, c["league"])
        if ok:
            after_tier.append(c)
    if not after_tier:
        b_leagues = sorted({c["league"] for c in after_side})
        p_tier = config.LEAGUE_TIERS.get(player_lg, "?")
        out["failure_stage"]  = "tier_rule_exclusive"
        out["reason_summary"] = f"League-tier rule excludes all {len(after_side)} buyers"
        out["reason_detail"]  = (
            f"All {len(after_side)} budget+side-feasible buyers are in leagues "
            f"that would require a downward move from player's {player_lg} "
            f"(tier {p_tier}). Buyer leagues: {b_leagues}. Tier rule allows "
            f"lateral or upward only; Tier D self-contained."
        )
        return out

    # Stage 5: score floor — compute best possible score across all surviving
    # candidates. Start with the highest-budget candidate as a default so we
    # have a meaningful "closest" buyer even when every score is 0.
    after_tier_by_budget = sorted(after_tier, key=lambda c: -(c["max_transfer_fee_eur"] or 0))
    best_buyer = after_tier_by_budget[0]
    best_score, best_bf, best_di, _ = score_pair(
        player_tm, sell, best_buyer["max_transfer_fee_eur"],
        best_buyer["source"], best_buyer["validated"],
    )
    indicative_fee = player_tm * config.tm_to_fee_multiplier(player_tm)
    for c in after_tier[1:]:
        score, bf, di, _ = score_pair(player_tm, sell, c["max_transfer_fee_eur"],
                                       c["source"], c["validated"])
        if score > best_score:
            best_score = score
            best_buyer = c
            best_bf = bf
            best_di = di

    out["best_buyer"]          = best_buyer["club_name"] if best_buyer else ""
    out["best_possible_score"] = round(best_score, 1)
    out["failure_stage"]       = "score_floor"

    # Decompose the failure
    parent_oc = player["parent_club_id"] is None or player.get("parent_league") is None
    reasons = []
    if sell < LOW_SELLABILITY_THRESHOLD:
        reasons.append(f"sellability {sell:.1f} below {LOW_SELLABILITY_THRESHOLD:.0f}")
    if parent_oc:
        reasons.append("parent_club outside our 19-league coverage")
    if best_bf < 0.3:
        reasons.append(
            f"indicative fee €{indicative_fee/1e6:.0f}m exceeds buyers' budgets "
            f"(best budget_fit {best_bf:.2f} vs {best_buyer['club_name']} at "
            f"€{best_buyer['max_transfer_fee_eur']/1e6:.0f}m)"
        )
    if player["right_priced"] == 0:
        last_fee = player["last_fee_paid_eur"]
        if last_fee and last_fee > player_tm:
            reasons.append(
                f"right_priced=0 — owning club paid €{last_fee/1e6:.0f}m vs "
                f"current TM €{player_tm/1e6:.0f}m (paper loss, won't sell at TM)"
            )

    if not reasons:
        reasons.append(
            f"combination of mid-tier sellability ({sell:.1f}), demand_intensity "
            f"({best_di:.2f}), and budget_fit ({best_bf:.2f})"
        )

    out["reason_summary"] = f"Best possible {best_score:.1f} (floor {int(MATCH_SCORE_FLOOR)})"
    out["reason_detail"]  = (
        f"{len(after_tier)} buyer(s) cleared all four filters but the best "
        f"possible match score is {best_score:.1f}, below the floor of "
        f"{int(MATCH_SCORE_FLOOR)}. Drivers: " + "; ".join(reasons) + "."
    )
    return out


# ─── Write the sheet ──────────────────────────────────────────────────────────

HEADERS = [
    "name", "age", "position", "current club", "parent club", "league of parent",
    "TM value (€)", "sellability", "right_priced", "finished_product",
    "contract_leveraged", "on_loan",
    "failure_stage", "reason_summary", "reason_detail",
    "best_buyer (if any)", "best_possible_score",
]
WIDTHS = [
    28, 5, 8, 32, 28, 16, 13, 12, 12, 14, 16, 7,
    22, 44, 80, 28, 16,
]


def write_sheet(wb, rows):
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F3864")
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = bold; c.fill = fill; c.alignment = centre
    ws.row_dimensions[1].height = 32
    for ci, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"

    money = '"€"#,##0;[Red]"€"-#,##0'
    wrap = Alignment(wrap_text=True, vertical="top")

    stage_fills = {
        "no_position_demand":   PatternFill("solid", fgColor="FCE4D6"),  # orange
        "all_buyers_below_min": PatternFill("solid", fgColor="FFF2CC"),  # yellow
        "side_preference_excl": PatternFill("solid", fgColor="DDEBF7"),  # light blue
        "tier_rule_exclusive":  PatternFill("solid", fgColor="E2EFDA"),  # light green
        "score_floor":          PatternFill("solid", fgColor="F2F2F2"),  # grey
    }

    # Sort: by failure_stage (in priority order), then by sellability DESC
    stage_order = ["no_position_demand", "all_buyers_below_min", "side_preference_excl",
                   "tier_rule_exclusive", "score_floor"]
    rows = sorted(rows, key=lambda r: (stage_order.index(r["failure_stage"]) if r["failure_stage"] in stage_order else 99,
                                       -(r["sellability"] or 0)))

    for ri, r in enumerate(rows, start=2):
        ws.cell(row=ri, column=1,  value=r["name"])
        ws.cell(row=ri, column=2,  value=r["age"])
        ws.cell(row=ri, column=3,  value=r["position"])
        ws.cell(row=ri, column=4,  value=r["current_club"])
        ws.cell(row=ri, column=5,  value=r["parent_club"])
        ws.cell(row=ri, column=6,  value=r["parent_league"])
        c = ws.cell(row=ri, column=7, value=r["tm_value_eur"]); c.number_format = money
        c = ws.cell(row=ri, column=8, value=r["sellability"]); c.number_format = "0.0"
        ws.cell(row=ri, column=9,  value=r["right_priced"])
        ws.cell(row=ri, column=10, value=r["finished_product"])
        ws.cell(row=ri, column=11, value=r["contract_leveraged"])
        ws.cell(row=ri, column=12, value=r["on_loan"])
        c = ws.cell(row=ri, column=13, value=r["failure_stage"])
        c.fill = stage_fills.get(r["failure_stage"], PatternFill())
        ws.cell(row=ri, column=14, value=r["reason_summary"])
        c = ws.cell(row=ri, column=15, value=r["reason_detail"]); c.alignment = wrap
        ws.cell(row=ri, column=16, value=r["best_buyer"])
        c = ws.cell(row=ri, column=17, value=r["best_possible_score"]); c.number_format = "0.0"

    last_col = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"


def main():
    if not Path(config.SQLITE_FILE).exists():
        sys.exit(f"Missing {config.SQLITE_FILE}.")
    if not WORKBOOK_PATH.exists():
        sys.exit(f"Missing {WORKBOOK_PATH}.")

    display_map = cd.load_display_map(WORKBOOK_PATH)
    player_map = pd_disp.load_display_map(WORKBOOK_PATH)
    with sqlite3.connect(config.SQLITE_FILE) as con:
        con.row_factory = sqlite3.Row
        matched = set(r[0] for r in con.execute("SELECT DISTINCT player_id FROM matches"))
        sellables = list(con.execute("""
            SELECT pu.player_id, pu.name, pu.age, pu.position_bucket,
                   pu.current_club, pu.current_club_id,
                   pu.parent_club, pu.parent_club_id, pu.on_loan, pu.league_id,
                   pu.current_tm_value_eur, pu.sellability_score, pu.right_priced,
                   pu.finished_product, pu.contract_leveraged, pu.last_fee_paid_eur,
                   cp.league_id AS parent_league
            FROM player_universe pu
            LEFT JOIN club_pressure cp ON cp.club_id = pu.parent_club_id
            WHERE (pu.right_priced=1 OR pu.finished_product=1
                   OR pu.finished_product IS NULL OR pu.contract_leveraged=1)
        """))
        orphans = []
        for p in sellables:
            if p["player_id"] in matched:
                continue
            d = dict(p)
            d["current_club_display"] = cd.display_for(d.get("current_club_id"), d["current_club"], display_map)
            d["parent_club_display"]  = cd.display_for(d.get("parent_club_id"),  d["parent_club"],  display_map)
            d["name"] = pd_disp.display_for(d["player_id"], d["name"], player_map)
            orphans.append(d)
        rows = [diagnose_player(con, p) for p in orphans]

    wb = load_workbook(WORKBOOK_PATH)
    write_sheet(wb, rows)
    wb.save(WORKBOOK_PATH)

    # Print summary
    from collections import Counter
    stage_counts = Counter(r["failure_stage"] for r in rows)
    print(f"Wrote {WORKBOOK_PATH} → sheet '{SHEET_NAME}' with {len(rows)} orphaned players.")
    print()
    print("By failure stage:")
    for stage in ["no_position_demand", "all_buyers_below_min", "side_preference_excl",
                  "tier_rule_exclusive", "score_floor"]:
        c = stage_counts.get(stage, 0)
        print(f"  {stage:24s}  {c:>3} players")
    print(f"  {'TOTAL':24s}  {len(rows):>3}")


if __name__ == "__main__":
    main()
