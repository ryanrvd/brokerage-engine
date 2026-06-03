"""
Reconcile `data/scisports_ratings.xlsx` against the current player universe.

Day 8 scope: every player in tm_squad_scrape (the TM-scraped PL squad universe)
PLUS the original sellable cohort (sellability_status = 'sellable_now') from
all 19 leagues. Championship players added in a follow-up phase.

Why scoped to tm_squad_scrape and not the broader player_universe: dcaribou's
current_club_name is "most recent club recorded", not "currently in squad".
Mixing dcaribou-sourced PL rows into the xlsx pollutes it with retired and
historical-association players. Only trust tm_squad_scrape for the
"currently in PL squad" classification. The sellable_now overlay ensures we
keep coverage of non-PL clean targets the matcher cares about.

User workflow:
  1. Run the pipeline through scripts 09 → reconcile_scisports.py
  2. Open `data/scisports_ratings.xlsx`, filter status=pending — that's the
     worklist. Fill in current_ability + potential_ability from Sci Sports.
  3. Save and run `python scripts/load_scisports_ratings.py`.

Reconciliation rules:
  • New player in scope, not in file → append with empty CA/PA, status=pending
  • Existing rated player still in scope → preserve CA/PA, mark active
  • Existing rated player NOT in scope → mark departed, preserve CA/PA
  • Player on Kill List → mark killed (still visible in file, flagged)

File-lock safety: if the workbook is open in Excel, openpyxl raises
PermissionError and the script exits with a clear message.
"""

from __future__ import annotations

import sqlite3
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import config
import kill_list

FILE_PATH = Path(config.DATA_DIR) / "scisports_ratings.xlsx"

COLUMNS: list[str] = [
    "tm_player_id",
    "player_name",
    "parent_club",
    "current_club",
    "position_bucket",
    "age",
    "is_on_loan",
    "current_ability",
    "potential_ability",
    "status",
    "last_updated",
    "notes",
]
EDITABLE_COLS: set[str] = {"current_ability", "potential_ability", "notes"}
COLUMN_WIDTHS: dict[str, int] = {
    "tm_player_id":      14,
    "player_name":       28,
    "parent_club":       30,
    "current_club":      30,
    "position_bucket":   10,
    "age":                6,
    "is_on_loan":         12,
    "current_ability":   16,
    "potential_ability": 16,
    "status":            12,
    "last_updated":      14,
    "notes":             42,
}
STATUS_ORDER = {"pending": 0, "active": 1, "departed": 2, "killed": 3}


# ─── DB helpers ──────────────────────────────────────────────────────────────
def get_current_cohort(con: sqlite3.Connection) -> dict[int, dict]:
    """Return {tm_player_id: {...}} for every player in scope.

    Scope = tm_squad_scrape UNION sellability_status='sellable_now'.

    tm_squad_scrape = current PL squad players (TM-scraped), including loaned-out.
    sellable_now overlay = preserves the prior cohort's coverage for the
    matcher's existing 19-league sellable set.
    """
    # Include out_of_scope players too — those are Kill List hits, and we
    # want them visible in the xlsx as status='killed' (not silently dropped
    # as 'departed', which is what would happen if we narrowed cohort here).
    rows = con.execute("""
        SELECT pu.player_id, pu.name, pu.parent_club, pu.current_club,
               pu.position_bucket, pu.age, pu.on_loan, pu.data_source,
               pu.sellability_status
        FROM player_universe pu
        WHERE pu.data_source = 'tm_squad_scrape'
           OR pu.sellability_status IN ('sellable_now', 'out_of_scope')
    """).fetchall()

    out: dict[int, dict] = {}
    for (pid, name, parent_club, current_club, position_bucket,
         age, on_loan, data_source, sellability_status) in rows:
        out[int(pid)] = {
            "player_name":     name or "",
            "parent_club":     parent_club or "",
            "current_club":    current_club or "",
            "position_bucket": position_bucket or "",
            "age":             age,
            "is_on_loan":      True if on_loan else False,
        }
    return out


def get_killed_ids(con: sqlite3.Connection) -> set[int]:
    state = kill_list.compute_kill_list_state(con)
    return {int(pid) for pid in state["excluded_ids"]}


# ─── File I/O ────────────────────────────────────────────────────────────────
def read_existing(file_path: Path) -> dict[int, dict]:
    if not file_path.exists():
        return {}
    wb = load_workbook(file_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    headers = list(rows[0])
    out: dict[int, dict] = {}
    for raw in rows[1:]:
        if not raw or raw[0] is None:
            continue
        d = dict(zip(headers, raw))
        try:
            pid = int(d.get("tm_player_id"))
        except (TypeError, ValueError):
            continue
        out[pid] = d
    return out


def _is_rated(ca, pa) -> bool:
    for v in (ca, pa):
        if v is None:
            continue
        s = str(v).strip()
        if s and s.lower() not in ("none", "nan"):
            return True
    return False


def write_workbook(file_path: Path, rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "scisports_ratings"

    ws.append(COLUMNS)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3864")
    for col_idx, _ in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    editable_fill = PatternFill("solid", fgColor="FFFBEA")
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            value = row.get(col_name)
            if isinstance(value, date):
                value = value.isoformat()
            if isinstance(value, bool):
                value = "TRUE" if value else "FALSE"
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_name in EDITABLE_COLS:
                cell.fill = editable_fill

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(col_name, 16)
    ws.freeze_panes = "A2"

    try:
        wb.save(file_path)
    except PermissionError as e:
        raise PermissionError(
            f"Cannot write {file_path} — file may be open in Excel. "
            "Close it and re-run."
        ) from e


# ─── Reconciliation ──────────────────────────────────────────────────────────
def reconcile(cohort: dict[int, dict],
              killed_ids: set[int],
              existing: dict[int, dict],
              today_iso: str) -> tuple[list[dict], dict[str, int], dict[str, int]]:
    """Return (sorted_rows, status_counts, change_counts)."""
    reconciled: dict[int, dict] = {}
    cohort_ids = set(cohort.keys())
    existing_ids = set(existing.keys())

    new_count = 0
    departed_count = 0
    reactivated_count = 0

    for pid, meta in cohort.items():
        prev = existing.get(pid, {})
        prev_status = (prev.get("status") or "").strip().lower()
        ca = prev.get("current_ability")
        pa = prev.get("potential_ability")

        if pid in killed_ids:
            new_status = "killed"
        elif _is_rated(ca, pa):
            new_status = "active"
        else:
            new_status = "pending"

        last_updated = prev.get("last_updated") or today_iso
        if new_status != prev_status:
            last_updated = today_iso

        if pid not in existing_ids:
            new_count += 1
        elif prev_status == "departed":
            reactivated_count += 1

        reconciled[pid] = {
            "tm_player_id":      pid,
            "player_name":       meta["player_name"],
            "parent_club":       meta["parent_club"],
            "current_club":      meta["current_club"],
            "position_bucket":   meta["position_bucket"],
            "age":               meta["age"],
            "is_on_loan":        meta["is_on_loan"],
            "current_ability":   ca,
            "potential_ability": pa,
            "status":            new_status,
            "last_updated":      last_updated,
            "notes":             prev.get("notes") or "",
        }

    for pid in existing_ids - cohort_ids:
        prev = existing[pid]
        prev_status = (prev.get("status") or "").strip().lower()
        new_status = "departed"
        if prev_status != "departed":
            departed_count += 1
        last_updated = prev.get("last_updated") or today_iso
        if new_status != prev_status:
            last_updated = today_iso
        reconciled[pid] = {
            "tm_player_id":      pid,
            "player_name":       prev.get("player_name") or "",
            "parent_club":       prev.get("parent_club") or "",
            "current_club":      prev.get("current_club") or "",
            "position_bucket":   prev.get("position_bucket") or "",
            "age":               prev.get("age"),
            "is_on_loan":        prev.get("is_on_loan") or False,
            "current_ability":   prev.get("current_ability"),
            "potential_ability": prev.get("potential_ability"),
            "status":            new_status,
            "last_updated":      last_updated,
            "notes":             prev.get("notes") or "",
        }

    sorted_rows = sorted(
        reconciled.values(),
        key=lambda r: (
            STATUS_ORDER.get(r["status"], 9),
            str(r["player_name"] or "").lower(),
        ),
    )

    status_counts = {
        "pending":  sum(1 for r in sorted_rows if r["status"] == "pending"),
        "active":   sum(1 for r in sorted_rows if r["status"] == "active"),
        "departed": sum(1 for r in sorted_rows if r["status"] == "departed"),
        "killed":   sum(1 for r in sorted_rows if r["status"] == "killed"),
    }
    change_counts = {
        "new": new_count,
        "departed": departed_count,
        "reactivated": reactivated_count,
    }
    return sorted_rows, status_counts, change_counts


def main() -> None:
    today_iso = date.today().isoformat()
    con = sqlite3.connect(config.SQLITE_FILE)
    try:
        cohort = get_current_cohort(con)
        killed_ids = get_killed_ids(con)
    finally:
        con.close()

    existing = read_existing(FILE_PATH)
    existing_count = len(existing)
    overlap_count = len(set(existing.keys()) & set(cohort.keys()))
    sorted_rows, status_counts, change_counts = reconcile(
        cohort, killed_ids, existing, today_iso,
    )

    write_workbook(FILE_PATH, sorted_rows)

    print(f"Reconciled {FILE_PATH}:")
    print(f"  Existing rows preserved:                {existing_count}")
    print(f"  New PL squad rows added (status=pending+rated): {change_counts['new']}")
    print(f"  Rows overlapping (already in xlsx):     {overlap_count}")
    print(f"  Departures marked:                      {change_counts['departed']}")
    print(f"  Reactivations:                          {change_counts['reactivated']}")
    print(f"  Total rows:                             {len(sorted_rows)}")
    print()
    print("By status:")
    for status in ("active", "killed", "pending", "departed"):
        label = {
            "active":   "active (rated)",
            "killed":   "killed",
            "pending":  "pending (awaiting rating)",
            "departed": "departed",
        }[status]
        print(f"  {label:35s} {status_counts.get(status, 0)}")


if __name__ == "__main__":
    main()
