"""
Step 17 — Write Sheets 5, 6, 7 of BrokerageWorkbook.xlsx (Day 4).

Sheet 5 — Live Demand Signal
    Cross-league count of clubs needing each position across the 10 demand-mapped
    leagues. Columns: Position, Clubs Needing, Top 5 Clubs by Max Transfer Fee.
    Footnote: covers 10 of 19 leagues (deferred per BACKLOG.md).

Sheet 6 — Live Supply Signal
    Sister sheet to 5. Count of sellable players per position across the full 19-
    league universe. Columns: Position, Sellable Players, Top 5 by Sellability Score.

Sheet 7 — Demand Map Mirror
    Consolidated demand from the 8 manual workbooks. One row per request (with
    "Either" expansions collapsed back to a single representative row to mirror
    what's in the source workbook, not the duplicated bucket-rows we use for
    aggregation). Sorted by league then club.

All three derive from the map_* tables loaded by scripts/16_load_market_maps.py.
"""

import sqlite3
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import config

WORKBOOK_PATH = Path("BrokerageWorkbook.xlsx")

SHEET5_NAME = "Live Demand Signal"
SHEET6_NAME = "Live Supply Signal"
SHEET7_NAME = "Demand Map Mirror"

# Canonical 10-bucket vocabulary, ordered for readability.
BUCKET_ORDER = ["GK", "CB", "LB", "RB", "DM", "CM", "AM", "LW", "RW", "ST_CF"]
BUCKET_DISPLAY = {
    "GK": "GK (Goalkeeper)",
    "CB": "CB (Centre Back)",
    "LB": "LB (Left Back)",
    "RB": "RB (Right Back)",
    "DM": "DM (Defensive Mid)",
    "CM": "CM (Centre Mid)",
    "AM": "AM (Attacking Mid)",
    "LW": "LW (Left Winger)",
    "RW": "RW (Right Winger)",
    "ST_CF": "ST_CF (Centre Forward)",
}

# Source-of-truth count for the footnote.
DEMAND_LEAGUES_COVERED = 10
DEMAND_LEAGUES_TOTAL = 19


# ─── Workbook helpers ───────────────────────────────────────────────────────────

def _open_workbook(path: Path) -> Workbook:
    if path.exists():
        wb = load_workbook(path)
        for sn in (SHEET5_NAME, SHEET6_NAME, SHEET7_NAME):
            if sn in wb.sheetnames:
                del wb[sn]
        return wb
    wb = Workbook()
    wb.active.title = "Sheet 1 (placeholder)"
    return wb


def _write_header(ws, headers: list[str], row: int = 1) -> None:
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F3864")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = bold
        cell.fill = fill
        cell.alignment = center


def _write_footnote(ws, text: str, after_row: int, col_span: int) -> None:
    italic = Font(italic=True, color="666666", size=10)
    cell = ws.cell(row=after_row + 2, column=1, value=text)
    cell.font = italic
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=after_row + 2, start_column=1,
                   end_row=after_row + 2, end_column=col_span)


# ─── Sheet 5 — Live Demand Signal ───────────────────────────────────────────────

SHEET5_HEADERS = [
    "Position Bucket",
    "Clubs Needing",
    "Total Requests",
    "Top 5 Clubs (by Max Transfer Fee €)",
    "Leagues Represented",
]
SHEET5_WIDTHS = [24, 14, 14, 80, 30]


def write_demand_signal_sheet(wb: Workbook, con: sqlite3.Connection) -> int:
    ws = wb.create_sheet(SHEET5_NAME)
    _write_header(ws, SHEET5_HEADERS)

    # Per-bucket: distinct clubs, request count, leagues set, top-5 clubs by max budget.
    rows_out: list[tuple] = []
    for bucket in BUCKET_ORDER:
        clubs_set: set[tuple[str, str]] = set()   # (club_name, league)
        request_count = 0
        leagues_set: set[str] = set()
        # Aggregate by (club, league) — pick that club's MAX budget across their requests
        club_max_budget: dict[tuple[str, str], int] = {}
        rows = con.execute("""
            SELECT club_name, league, max_transfer_fee_eur
            FROM map_club_requests
            WHERE position_bucket = ?
              AND workbook_source = 'manual_workbook'
        """, (bucket,)).fetchall()
        for club, lg, budget in rows:
            clubs_set.add((club, lg))
            leagues_set.add(lg)
            request_count += 1
            b = budget or 0
            key = (club, lg)
            if b > club_max_budget.get(key, -1):
                club_max_budget[key] = b
        top5 = sorted(club_max_budget.items(), key=lambda kv: -kv[1])[:5]
        top5_str = "; ".join(
            f"{club} ({lg}, €{budget:,})" if budget else f"{club} ({lg}, n/a)"
            for (club, lg), budget in top5
        )
        rows_out.append((
            BUCKET_DISPLAY[bucket],
            len(clubs_set),
            request_count,
            top5_str,
            ", ".join(sorted(leagues_set)),
        ))

    for ri, r in enumerate(rows_out, start=2):
        for ci, val in enumerate(r, start=1):
            ws.cell(row=ri, column=ci, value=val)

    for ci, w in enumerate(SHEET5_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(SHEET5_HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows_out) + 1}"
    _write_footnote(
        ws,
        f"Covers {DEMAND_LEAGUES_COVERED} of {DEMAND_LEAGUES_TOTAL} leagues "
        "(England 1+2, France 1+2, Italy, Spain, Germany, Netherlands, Portugal, Belgium); "
        "other 9 leagues' demand mapping deferred to Stage 2 per BACKLOG.md "
        "'Maps Auto-Sync Infrastructure'. Source: data/market_maps/*.xlsx "
        "(manual Google Sheets downloads, stale on arrival).",
        after_row=len(rows_out) + 1,
        col_span=len(SHEET5_HEADERS),
    )
    return len(rows_out)


# ─── Sheet 6 — Live Supply Signal ───────────────────────────────────────────────

SHEET6_HEADERS = [
    "Position Bucket",
    "Sellable Players",
    "Top 5 (by Sellability Score)",
    "Leagues Represented",
]
SHEET6_WIDTHS = [24, 18, 80, 30]


def write_supply_signal_sheet(wb: Workbook, con: sqlite3.Connection) -> int:
    ws = wb.create_sheet(SHEET6_NAME)
    _write_header(ws, SHEET6_HEADERS)

    # "Sellable" = same filter Sheet 4 uses (player passes at least one of the three flags).
    # We restrict to the 19-league universe (which is all of player_universe).
    rows_out: list[tuple] = []
    for bucket in BUCKET_ORDER:
        rows = con.execute("""
            SELECT name, current_club, league, sellability_score
            FROM player_universe
            WHERE position_bucket = ?
              AND sellability_status = 'sellable_now'
            ORDER BY sellability_score DESC NULLS LAST
        """, (bucket,)).fetchall()
        total = len(rows)
        top5 = rows[:5]
        leagues_set = {r[2] for r in rows if r[2]}
        top5_str = "; ".join(
            f"{name} ({club}, {sell:.1f})" if sell is not None
            else f"{name} ({club}, n/a)"
            for name, club, _, sell in top5
        )
        rows_out.append((
            BUCKET_DISPLAY[bucket],
            total,
            top5_str,
            ", ".join(sorted(leagues_set)),
        ))

    for ri, r in enumerate(rows_out, start=2):
        for ci, val in enumerate(r, start=1):
            ws.cell(row=ri, column=ci, value=val)

    for ci, w in enumerate(SHEET6_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(SHEET6_HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows_out) + 1}"
    _write_footnote(
        ws,
        "Covers all 19 leagues in player_universe. 'Sellable' = passes at least one of "
        "right_priced / finished_product / contract_leveraged (same filter as Sheet 4 "
        "Sellable Assets).",
        after_row=len(rows_out) + 1,
        col_span=len(SHEET6_HEADERS),
    )
    return len(rows_out)


# ─── Sheet 7 — Demand Map Mirror ────────────────────────────────────────────────

SHEET7_HEADERS = [
    "Club", "League", "Date Last Updated",
    "Position Category", "Preferred Side", "Role Notes",
    "Max Transfer Fee (€)", "Max Wage PW (€)",
    "Source", "Validated", "Validated By",
    "Linked/Shortlisted Players",
    "Source Workbook",
]
SHEET7_WIDTHS = [28, 7, 14, 18, 12, 40, 16, 14, 8, 10, 14, 36, 36]


def write_demand_mirror_sheet(wb: Workbook, con: sqlite3.Connection) -> int:
    ws = wb.create_sheet(SHEET7_NAME)
    _write_header(ws, SHEET7_HEADERS)

    # Collapse the "Either" expansions back into single workbook-source rows:
    # group by (club, league, date_last_updated, position_category, preferred_side,
    # role_notes, …) and pick one representative — the bucket duplication is an
    # implementation detail for aggregation, not for presentation here.
    rows = con.execute("""
        SELECT MIN(request_id),
               club_name, league, date_last_updated,
               position_category, preferred_side, role_notes,
               max_transfer_fee_eur, max_wage_pw_eur,
               source, validated, validated_by,
               linked_shortlisted_player, source_file
        FROM map_club_requests
        WHERE workbook_source = 'manual_workbook'
        GROUP BY club_name, league, date_last_updated,
                 position_category, preferred_side, role_notes,
                 max_transfer_fee_eur, max_wage_pw_eur,
                 source, validated, validated_by,
                 linked_shortlisted_player, source_file
        ORDER BY league, club_name, MIN(request_id)
    """).fetchall()

    for ri, r in enumerate(rows, start=2):
        (_rid, club, lg, dlu, cat, side, notes, max_fee, max_wage, src,
         val, val_by, linked, source_file) = r
        ws.cell(row=ri, column=1, value=club)
        ws.cell(row=ri, column=2, value=lg)
        ws.cell(row=ri, column=3, value=dlu)
        ws.cell(row=ri, column=4, value=cat)
        ws.cell(row=ri, column=5, value=side)
        ws.cell(row=ri, column=6, value=notes)
        ws.cell(row=ri, column=7, value=max_fee)
        ws.cell(row=ri, column=8, value=max_wage)
        ws.cell(row=ri, column=9, value=src)
        ws.cell(row=ri, column=10, value=val)
        ws.cell(row=ri, column=11, value=val_by)
        ws.cell(row=ri, column=12, value=linked)
        ws.cell(row=ri, column=13, value=source_file)

    for ri in range(2, len(rows) + 2):
        ws.cell(row=ri, column=7).number_format = '"€"#,##0'
        ws.cell(row=ri, column=8).number_format = '"€"#,##0'

    for ci, w in enumerate(SHEET7_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(SHEET7_HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"
    return len(rows)


# ─── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    wb = _open_workbook(WORKBOOK_PATH)
    with sqlite3.connect(config.SQLITE_FILE) as con:
        n5 = write_demand_signal_sheet(wb, con)
        n6 = write_supply_signal_sheet(wb, con)
        n7 = write_demand_mirror_sheet(wb, con)
    wb.save(WORKBOOK_PATH)
    print(f"Saved {WORKBOOK_PATH.resolve()}")
    print(f"  Sheet 5 '{SHEET5_NAME}'  — {n5} position buckets ({DEMAND_LEAGUES_COVERED}/{DEMAND_LEAGUES_TOTAL} leagues)")
    print(f"  Sheet 6 '{SHEET6_NAME}'  — {n6} position buckets (19/19 leagues)")
    print(f"  Sheet 7 '{SHEET7_NAME}'  — {n7} consolidated demand rows")


if __name__ == "__main__":
    main()
