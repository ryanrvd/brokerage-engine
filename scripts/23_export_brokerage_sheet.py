"""
Step 23 — Write Sheet 1 (Brokerage Opportunities) + maintain the Kill List tab.

One row per (player, buyer_request) pair from the `matches` table, sorted by
match_score DESC. Excludes rows for any player on the Kill List (manual entries
+ agency-rule auto entries — see kill_list.py for the unified logic).

Columns are deliberately spare: commercial pricing is Yatin's call on review.

Kill List behaviour:
  • Single source of truth (kill_list.py); shared with the Streamlit app.
  • Manual entries (user-typed in BrokerageWorkbook.xlsx → Kill List tab) are
    preserved verbatim across pipeline runs.
  • Agency-rule auto entries are regenerated every run from
    data/blocked_agencies.csv — any player whose `agency` field matches a row
    there is auto-killed. Currently: CAA Base, Gestifute, Gol International,
    Unique Sports Group.
  • A "source" column distinguishes manual rows from agency:<name> rows.
  • Fuzzy match warns on zero hits OR ambiguous (multi-hit) manual entries —
    never silently drops or over-drops.
"""

import sqlite3
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
import config
import kill_list       # shared with app/db.py
import club_display    # club name → display
import player_display  # player name → display

WORKBOOK_PATH = Path("BrokerageWorkbook.xlsx")
SHEET1_NAME = "Brokerage Opportunities"
PLACEHOLDER_NAME = "Sheet 1 (placeholder)"
LEGACY_HEADLINE_NAME = "Headline Tier (€40-80m)"  # remove if present from prior runs

KILL_LIST_SHEET = "Kill List"
KILL_LIST_HEADERS = ["player", "reason why", "source"]
KILL_LIST_WIDTHS = [32, 60, 24]

HEADERS = [
    "name", "age", "position", "current club", "parent club", "league of parent",
    "TM value (€)", "contract end date", "contract years remaining", "agency",
    # Sci Sports — player rating layer
    "current_ability", "potential_ability", "talent_band",
    "buyer club", "buyer league", "max_transfer_fee_eur", "max_wage_pw_eur",
    "match_score", "sellability_score", "demand_intensity_label", "budget_fit_pct",
    # Sci Sports — match-level fit
    "level_fit", "player_ca", "club_threshold",
    "brokerage_rationale",
]
WIDTHS = [
    28, 5, 8, 32, 28, 16, 13, 16, 14, 24,
    12, 12, 18,
    28, 12, 16, 15,
    11, 14, 22, 13,
    12, 10, 14,
    90,
]
# Column indices (1-based) for Sci Sports cells — used to apply a light-grey
# fill so the SciSports-derived block reads as a visual group.
SCISPORTS_COL_INDICES = {11, 12, 13, 22, 23, 24}


def demand_intensity_label(source: str | None, validated: str | None) -> str:
    src = (source or "").strip()
    val = (validated or "").strip().upper()
    if src == "Agent" and val == "YES":
        return "Validated agent"
    if src == "Agent":
        return "Agent (partial)"
    if src == "Intel":
        return "Intel-derived"
    if src == "Inferred":
        return "Inferred (squad gap)"
    return "Unverified"


def contract_years_remaining(contract_end: str | None, snapshot: date) -> float | None:
    if not contract_end:
        return None
    try:
        ce = datetime.fromisoformat(contract_end).date()
    except (TypeError, ValueError):
        return None
    return round((ce - snapshot).days / 365.25, 2)


def build_rationale(seller: dict, buyer: dict, tier_move: str, headroom_eur: int,
                    display_map: dict[int, str] | None = None) -> str:
    """Combine seller-side rationale, buyer-side rationale, trade frame."""
    drivers: list[str] = []
    if seller.get("public_must_sell_flag") == 1:
        drivers.append("public must-sell flag (parachute/FFP)")
    if seller.get("manager_change_flag") == 1:
        drivers.append("manager change")
    if seller.get("contract_leveraged") == 1:
        drivers.append("contract leverage")
    if seller.get("on_loan"):
        drivers.append("loan with sell-readiness signal")
    if not drivers:
        comps = [
            ("contract leverage", seller.get("contract_leverage_score") or 0),
            ("squad oversupply", seller.get("squad_oversupply_score") or 0),
            ("net-spend headroom", seller.get("net_spend_score") or 0),
        ]
        comps.sort(key=lambda x: -x[1])
        if comps[0][1] > 0:
            drivers.append(f"{comps[0][0]} ({comps[0][1]:.0f}/100)")

    parent_id = seller.get("parent_club_id")
    parent_official = seller.get("parent_club") or "(unknown)"
    parent_name = (display_map or {}).get(int(parent_id), parent_official) if parent_id is not None else parent_official
    parent_league = config.LEAGUE_DISPLAY.get(seller.get("parent_league"), seller.get("parent_league") or "—")
    if parent_id is None:
        seller_text = f"parent club outside coverage — {parent_name}"
    elif drivers:
        seller_text = f"{parent_name} ({parent_league}) — " + ", ".join(drivers[:2])
    else:
        seller_text = f"{parent_name} ({parent_league}) — low structural pressure"

    src = (buyer.get("request_source") or "").strip()
    val = (buyer.get("request_validated") or "").strip().upper()
    buyer_id = buyer.get("buyer_club_id")
    buyer_display = (display_map or {}).get(int(buyer_id), buyer["buyer_club_name"]) if buyer_id is not None else buyer["buyer_club_name"]
    buyer_lg = config.LEAGUE_DISPLAY.get(buyer["buyer_league_id"], buyer["buyer_league_id"])
    if src == "Agent" and val == "YES":
        buyer_text = f"{buyer_display} ({buyer_lg}) — validated agent-confirmed need"
    elif src == "Intel":
        buyer_text = f"{buyer_display} ({buyer_lg}) — intel-derived positional gap"
    elif src == "Inferred":
        buyer_text = f"{buyer_display} ({buyer_lg}) — inferred positional gap (squad-thinness)"
    else:
        buyer_text = f"{buyer_display} ({buyer_lg}) — open need (unverified)"

    return (
        f"Seller: {seller_text}. Buyer: {buyer_text}. "
        f"Trade: {tier_move} move, €{headroom_eur/1e6:.0f}m headroom above TM."
    )


def fetch_match_rows(con: sqlite3.Connection) -> list[dict]:
    """Pull matches + all needed player / pressure fields for rationale building.

    Filter: match_score >= 10 — Sheet 1 is the Brokerage Engine view
    (sellable_now-scoped, original score floor preserved). Market View match
    rows (NULL match_score) and below-floor Brokerage rows that only
    survived via the OR-floor against market_match_score are excluded.
    The Streamlit UI is the surface for the wider Market View cohort.
    """
    sql = """
        SELECT m.match_score, m.sellability_score, m.budget_fit,
               m.request_source, m.request_validated,
               m.buyer_club_id, m.buyer_club_name, m.buyer_league_id,
               m.max_transfer_fee_eur, m.max_wage_pw_eur, m.tier_move,
               m.player_ca, m.player_pa,
               m.club_threshold_for_request, m.level_fit,
               pu.player_id, pu.name AS player_name, pu.age, pu.position_bucket,
               pu.current_tm_value_eur, pu.contract_end_date, pu.agency,
               pu.current_club, pu.current_club_id, pu.parent_club, pu.parent_club_id, pu.on_loan,
               cp.league_id AS parent_league,
               cp.manager_change_flag, cp.public_must_sell_flag,
               cp.contract_leverage_score, cp.squad_oversupply_score, cp.net_spend_score,
               pu.contract_leveraged
        FROM matches m
        JOIN player_universe pu ON pu.player_id = m.player_id
        LEFT JOIN club_pressure cp ON cp.club_id = pu.parent_club_id
        WHERE m.match_score >= 10
        ORDER BY m.match_score DESC, pu.name, m.match_score_raw DESC
    """
    rows = con.execute(sql).fetchall()
    cols = [d[0] for d in con.execute(sql + " LIMIT 0").description]
    return [dict(zip(cols, r)) for r in rows]


def _write_header(ws) -> None:
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F3864")
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = bold
        c.fill = fill
        c.alignment = centre
    ws.row_dimensions[1].height = 32
    for ci, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"


def _write_rows(ws, rows: list[dict], snapshot: date,
                display_map: dict[int, str],
                player_display_map: dict[int, str] | None = None) -> int:
    n = 0
    money = '"€"#,##0;[Red]"€"-#,##0'
    pct = '0.0"%"'
    wrap = Alignment(wrap_text=True, vertical="top")
    player_display_map = player_display_map or {}

    def disp(club_id, fallback):
        if club_id is None:
            return fallback or ""
        return display_map.get(int(club_id), fallback or "")

    def disp_player(player_id, fallback):
        if player_id is None:
            return fallback or ""
        return player_display_map.get(int(player_id), fallback or "")

    for ri, r in enumerate(rows, start=2):
        n += 1
        current_display = disp(r.get("current_club_id"), r["current_club"])
        parent_display  = disp(r.get("parent_club_id"),  r["parent_club"])
        buyer_display   = disp(r.get("buyer_club_id"),   r["buyer_club_name"])

        current_club_display = (
            f"{current_display} (on loan from {parent_display})"
            if r["on_loan"] and r["parent_club"] else current_display
        )
        years_remaining = contract_years_remaining(r["contract_end_date"], snapshot)
        headroom = (r["max_transfer_fee_eur"] or 0) - (r["current_tm_value_eur"] or 0)
        rationale = build_rationale(
            seller=r,
            buyer={
                "request_source":    r["request_source"],
                "request_validated": r["request_validated"],
                "buyer_club_name":   r["buyer_club_name"],
                "buyer_club_id":     r.get("buyer_club_id"),
                "buyer_league_id":   r["buyer_league_id"],
            },
            tier_move=r["tier_move"],
            headroom_eur=headroom,
            display_map=display_map,
        )
        budget_fit_pct = round((r["budget_fit"] or 0.0) * 100.0, 1)
        parent_league_display = config.LEAGUE_DISPLAY.get(r["parent_league"], r["parent_league"] or "—")
        buyer_league_display  = config.LEAGUE_DISPLAY.get(r["buyer_league_id"], r["buyer_league_id"])

        # Sci Sports talent-band label — derived from CA at export time.
        _ca = r.get("player_ca")
        _pa = r.get("player_pa")
        if _ca is None:
            talent_band = "Not yet rated"
        elif _ca >= 140: talent_band = "Generational"
        elif _ca >= 120: talent_band = "Elite"
        elif _ca >= 100: talent_band = "First-team level"
        elif _ca >= 80:  talent_band = "Squad level"
        elif _ca >= 60:  talent_band = "Development"
        else:            talent_band = "Below cohort"

        ws.cell(row=ri, column=1,  value=disp_player(r.get("player_id"), r["player_name"]))
        ws.cell(row=ri, column=2,  value=r["age"])
        ws.cell(row=ri, column=3,  value=r["position_bucket"])
        ws.cell(row=ri, column=4,  value=current_club_display)
        ws.cell(row=ri, column=5,  value=parent_display)
        ws.cell(row=ri, column=6,  value=parent_league_display)
        c = ws.cell(row=ri, column=7,  value=r["current_tm_value_eur"]);  c.number_format = money
        ws.cell(row=ri, column=8,  value=r["contract_end_date"])
        ws.cell(row=ri, column=9,  value=years_remaining)
        ws.cell(row=ri, column=10, value=r["agency"])
        c = ws.cell(row=ri, column=11, value=_ca);                       c.number_format = "0.0"
        c = ws.cell(row=ri, column=12, value=_pa);                       c.number_format = "0.0"
        ws.cell(row=ri, column=13, value=talent_band)
        ws.cell(row=ri, column=14, value=buyer_display)
        ws.cell(row=ri, column=15, value=buyer_league_display)
        c = ws.cell(row=ri, column=16, value=r["max_transfer_fee_eur"]); c.number_format = money
        c = ws.cell(row=ri, column=17, value=r["max_wage_pw_eur"]);      c.number_format = money
        c = ws.cell(row=ri, column=18, value=r["match_score"]);          c.number_format = "0.0"
        c = ws.cell(row=ri, column=19, value=r["sellability_score"]);    c.number_format = "0.0"
        ws.cell(row=ri, column=20, value=demand_intensity_label(r["request_source"], r["request_validated"]))
        c = ws.cell(row=ri, column=21, value=budget_fit_pct);            c.number_format = pct
        ws.cell(row=ri, column=22, value=r.get("level_fit"))
        c = ws.cell(row=ri, column=23, value=_ca);                       c.number_format = "0.0"
        c = ws.cell(row=ri, column=24, value=r.get("club_threshold_for_request")); c.number_format = "0.0"
        c = ws.cell(row=ri, column=25, value=rationale)

        # Light-grey fill on the SciSports-derived columns for visual grouping
        _scisports_fill = PatternFill("solid", fgColor="F3F4F6")
        for col_idx in SCISPORTS_COL_INDICES:
            ws.cell(row=ri, column=col_idx).fill = _scisports_fill
        c.alignment = wrap
    # Auto filter + freeze.
    last_col = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{n + 1}"
    return n


def remove_sheet_if_exists(wb, name: str) -> None:
    if name in wb.sheetnames:
        del wb[name]


# ─── Kill List tab writer (logic lives in kill_list.py) ──────────────────────

def _write_kill_list_tab(
    wb,
    manual_entries: list[tuple[str, str]],
    auto_rows: list[dict],
    player_display_map: dict[int, str] | None = None,
) -> None:
    """Write the Kill List sheet as the LAST tab. Three columns:
    player / reason why / source. Manual rows first, then agency-rule rows.
    Player names rendered via display_map when known."""
    player_display_map = player_display_map or {}
    remove_sheet_if_exists(wb, KILL_LIST_SHEET)
    ws = wb.create_sheet(KILL_LIST_SHEET)
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="9C0006")
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, h in enumerate(KILL_LIST_HEADERS, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = bold
        c.fill = fill
        c.alignment = centre
    ws.row_dimensions[1].height = 28

    manual_fill = PatternFill("solid", fgColor="FFF2CC")  # yellow — user-editable
    auto_fill = PatternFill("solid", fgColor="F2F2F2")    # grey — auto-derived
    wrap = Alignment(wrap_text=True, vertical="top")

    row_i = 2
    for name, reason in manual_entries:
        for col, value, _ in ((1, name, None), (2, reason, None), (3, "manual", None)):
            c = ws.cell(row=row_i, column=col, value=value)
            c.fill = manual_fill
            c.alignment = wrap
        row_i += 1
    for r in sorted(auto_rows, key=lambda x: x["player_name"]):
        # Use display name if available, falling back to official name.
        name_for_display = player_display_map.get(int(r["player_id"]), r["player_name"]) if r.get("player_id") is not None else r["player_name"]
        for col, value in ((1, name_for_display), (2, r["reason"]), (3, r["source"])):
            c = ws.cell(row=row_i, column=col, value=value)
            c.fill = auto_fill
            c.alignment = wrap
        row_i += 1

    for ci, w in enumerate(KILL_LIST_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    if row_i > 2:
        ws.auto_filter.ref = f"A1:C{row_i - 1}"


def main() -> None:
    if not Path(config.SQLITE_FILE).exists():
        sys.exit(f"Missing {config.SQLITE_FILE} — run 02→22 first.")
    if not WORKBOOK_PATH.exists():
        sys.exit(f"Missing {WORKBOOK_PATH} — run 05/10/17 first.")

    snapshot = config.SNAPSHOT_DATE if isinstance(config.SNAPSHOT_DATE, date) \
               else datetime.fromisoformat(str(config.SNAPSHOT_DATE)).date()

    # Seed blocked_agencies.csv on first run
    kill_list.seed_blocked_agencies_if_missing()

    with sqlite3.connect(config.SQLITE_FILE) as con:
        all_rows = fetch_match_rows(con)
        # Compute Kill List state (manual + agency-rule)
        state = kill_list.compute_kill_list_state(con, workbook_path=WORKBOOK_PATH)
        # Compute Club Display Names state (read-preserve-write)
        display_state = club_display.compute_display_state(con, workbook_path=WORKBOOK_PATH)
        # Compute Player Display Names state (uses club display for context column)
        player_display_state = player_display.compute_display_state(
            con, workbook_path=WORKBOOK_PATH, club_display_map=display_state["display_map"]
        )

    display_map = display_state["display_map"]
    display_rows = display_state["rows"]
    n_clubs_total    = len(display_rows)
    n_clubs_manual   = sum(1 for r in display_rows if r["auto_or_manual"] == "manual")
    n_clubs_auto     = n_clubs_total - n_clubs_manual
    n_clubs_low_conf = sum(1 for r in display_rows if r["notes"] == club_display.LOW_CONFIDENCE_NOTE)

    player_display_map = player_display_state["display_map"]
    player_display_rows = player_display_state["rows"]
    n_players_total    = len(player_display_rows)
    n_players_manual   = sum(1 for r in player_display_rows if r["auto_or_manual"] == "manual")
    n_players_auto     = n_players_total - n_players_manual
    n_players_low_conf = sum(1 for r in player_display_rows if r["notes"] == player_display.LOW_CONFIDENCE_NOTE)

    manual_entries = state["manual_entries"]
    auto_rows      = state["auto_rows"]
    excluded_ids   = state["excluded_ids"]
    warnings       = state["warnings"]
    blocked_agencies = state["blocked_agencies"]

    n_before = len(all_rows)
    filtered_rows = [r for r in all_rows if r["player_id"] not in excluded_ids]
    n_after = len(filtered_rows)
    n_dropped = n_before - n_after

    wb = load_workbook(WORKBOOK_PATH)
    # Remove any prior versions of our sheets and the legacy Headline Tier.
    for nm in (PLACEHOLDER_NAME, SHEET1_NAME, LEGACY_HEADLINE_NAME):
        remove_sheet_if_exists(wb, nm)

    # Sheet 1 at position 0
    ws1 = wb.create_sheet(SHEET1_NAME, 0)
    _write_header(ws1)
    n1 = _write_rows(ws1, filtered_rows, snapshot, display_map, player_display_map)

    # ─── Display Names tabs and Kill List ────────────────────────────────
    club_display.write_display_tab(wb, display_rows, league_names=config.LEAGUE_DISPLAY)
    player_display.write_display_tab(wb, player_display_rows, league_names=config.LEAGUE_DISPLAY)
    _write_kill_list_tab(wb, manual_entries, auto_rows, player_display_map)

    # Desired sheet order
    desired_order = [
        "Brokerage Opportunities",
        "Player Universe",
        "Seller Pressure",
        "Sellable Assets",
        "Live Demand Signal",
        "Live Supply Signal",
        "Demand Map Mirror",
        "Club Display Names",
        "Player Display Names",
        "Zero-Match Diagnostic",
        "Kill List",
    ]
    present = [n for n in desired_order if n in wb.sheetnames]
    other   = [n for n in wb.sheetnames if n not in desired_order]
    final_order = present + other  # everything else (if any) trails
    # openpyxl ordering: move sheets one at a time
    for idx, name in enumerate(final_order):
        ws_obj = wb[name]
        # Use _sheets list manipulation as openpyxl >=3 lacks public reorder
        wb._sheets.remove(ws_obj)
        wb._sheets.insert(idx, ws_obj)

    wb.save(WORKBOOK_PATH)

    # ─── Report ─────────────────────────────────────────────────────────
    print(f"Wrote {WORKBOOK_PATH.resolve()}")
    print(f"  Sheet 1 '{SHEET1_NAME}' — {n1} match rows "
          f"(was {n_before}, dropped {n_dropped} via Kill List)")
    print(f"  Kill List tab — {len(manual_entries)} manual + {len(auto_rows)} auto = "
          f"{len(manual_entries) + len(auto_rows)} rows; {len(excluded_ids)} unique players excluded")
    print(f"  Blocked agencies in force ({len(blocked_agencies)}): {', '.join(blocked_agencies)}")
    print(f"  Club Display Names tab — {n_clubs_total} clubs ({n_clubs_auto} auto, "
          f"{n_clubs_manual} manual, {n_clubs_low_conf} low-confidence)")
    print(f"  Player Display Names tab — {n_players_total} players ({n_players_auto} auto, "
          f"{n_players_manual} manual, {n_players_low_conf} low-confidence)")

    # Top 30 longest reductions
    by_reduction = sorted(
        display_rows,
        key=lambda r: -(len(r["official_name"]) - len(r["display_name"])),
    )
    print()
    print("Top 30 longest reductions (official → display):")
    print(f"  {'reduction':>10s}  {'display':28s}  {'official':40s}")
    for r in by_reduction[:30]:
        reduction = len(r["official_name"]) - len(r["display_name"])
        if reduction <= 0:
            break
        print(f"  {reduction:>10d}  {r['display_name'][:28]:28s}  {r['official_name'][:40]}")

    # Low-confidence rows sorted by current_base_pressure_score
    low_conf = [r for r in display_rows if r["notes"] == club_display.LOW_CONFIDENCE_NOTE]
    if low_conf:
        with sqlite3.connect(config.SQLITE_FILE) as con2:
            pressure_lookup = {
                cid: (cl or 0) * 0.25 + (so or 0) * 0.20 + (ns or 0) * 0.20
                for cid, cl, so, ns in con2.execute(
                    "SELECT club_id, contract_leverage_score, squad_oversupply_score, net_spend_score FROM club_pressure"
                )
            }
        low_conf.sort(key=lambda r: -pressure_lookup.get(r["club_id"], 0))
        print()
        print(f"Low-confidence display names ({len(low_conf)}, sorted by base pressure DESC — review these first):")
        print(f"  {'base_p':>6s}  {'display':28s}  {'official':50s}")
        for r in low_conf[:30]:
            base_p = pressure_lookup.get(r["club_id"], 0)
            print(f"  {base_p:>6.1f}  {r['display_name'][:28]:28s}  {r['official_name'][:50]}")
        if len(low_conf) > 30:
            print(f"  … and {len(low_conf) - 30} more")

    # Rule-test examples (5 cases where football knowledge mattered)
    print()
    print("Club display rule check (5 specific cases):")
    rule_tests = [
        "Real Madrid Club de Fútbol",
        "Real Sociedad de Fútbol S.A.D.",
        "1. Fußballclub Union Berlin",
        "Associazione Calcio Milan",
        "AFC Ajax Amsterdam",
    ]
    by_official = {r["official_name"]: r for r in display_rows}
    for official in rule_tests:
        r = by_official.get(official)
        if r:
            print(f"  {official:50s} → {r['display_name']}")
        else:
            print(f"  {official:50s} → (not present in current data)")

    # ─── Player Display reporting ────────────────────────────────────────
    p_by_reduction = sorted(
        player_display_rows,
        key=lambda r: -(len(r["official_name"]) - len(r["display_name"])),
    )
    print()
    print("Top 30 player-name reductions (official → display):")
    print(f"  {'reduction':>10s}  {'display':28s}  {'official':40s}")
    for r in p_by_reduction[:30]:
        reduction = len(r["official_name"]) - len(r["display_name"])
        if reduction <= 0:
            break
        print(f"  {reduction:>10d}  {r['display_name'][:28]:28s}  {r['official_name'][:40]}")

    p_low_conf = [r for r in player_display_rows if r["notes"] == player_display.LOW_CONFIDENCE_NOTE]
    if p_low_conf:
        p_low_conf.sort(key=lambda r: -(r.get("sellability_score") or 0))
        print()
        print(f"Low-confidence player names ({len(p_low_conf)}, sorted by sellability DESC):")
        print(f"  {'sell':>5s}  {'display':28s}  {'official':40s}")
        for r in p_low_conf[:30]:
            print(f"  {(r.get('sellability_score') or 0):>5.1f}  {r['display_name'][:28]:28s}  {r['official_name'][:40]}")
        if len(p_low_conf) > 30:
            print(f"  … and {len(p_low_conf) - 30} more")

    print()
    print("Player display rule check (8 worked examples):")
    player_rule_tests = [
        ("Harrison James Burrows",        "Anglo middle-name strip"),
        ("José Ángel Carmona",            "Compound first name stays intact"),
        ("Joseph Paul Gelhardt",          "Nickname (Joe)"),
        ("Jan-Carlo Simić",               "Hyphenated first stays"),
        ("Zeno Van Den Bosch",            "Dutch particles preserved"),
        ("Takefusa Kubo",                 "Asian name (no change)"),
        ("Issahaku Abdul Fatawu",         "Reordering"),
        ("Sergio Arribas Calvo",          "Spanish dual-surname drop"),
    ]
    p_by_official = {r["official_name"]: r for r in player_display_rows}
    for official, rule in player_rule_tests:
        r = p_by_official.get(official)
        if r:
            print(f"  [{rule:32s}] {official:34s} → {r['display_name']}")
        else:
            print(f"  [{rule:32s}] {official:34s} → (not present)")
    print()
    print("Manual Kill List entries:")
    for name, reason in manual_entries:
        status = "✓"
        for w in warnings:
            if w["entry"] == name:
                status = "⚠ ZERO" if w["kind"] == "zero" else f"⚠ AMBIG ({len(w['matches'])})"
                break
        reason_short = (reason[:50] + "…") if reason and len(reason) > 50 else (reason or "")
        print(f"  {status:14s} {name:32s} {reason_short}")

    if auto_rows:
        print()
        print(f"Agency-rule auto-kills ({len(auto_rows)}):")
        for r in sorted(auto_rows, key=lambda x: (x["source"], x["player_name"]))[:20]:
            print(f"  {r['player_name']:30s}  {r['source']}")
        if len(auto_rows) > 20:
            print(f"  … and {len(auto_rows) - 20} more")

    if warnings:
        bar = "─" * 78
        print()
        print(bar)
        print("⚠  Kill List warnings — manual entries that need attention:")
        print(bar)
        for w in warnings:
            if w["kind"] == "zero":
                print(f"  ZERO MATCH: '{w['entry']}' did not match any player.")
                print(f"             Check spelling / accents. Entry was NOT applied to Sheet 1.")
            else:
                print(f"  AMBIGUOUS:  '{w['entry']}' matched {len(w['matches'])} players:")
                for pid, pn in w["matches"]:
                    print(f"                player_id={pid} — {pn}")
                print(f"             Entry was NOT applied (safer to under-drop than over-drop).")
        print(bar)

    print()
    print(f"Workbook now has {len(wb.sheetnames)} sheets:")
    for i, name in enumerate(wb.sheetnames):
        print(f"  {i}. {name}")


if __name__ == "__main__":
    main()
