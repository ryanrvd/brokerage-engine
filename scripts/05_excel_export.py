"""
Step 05 — Export the Player Universe to Sheet 2 of the Brokerage Workbook.

Creates BrokerageWorkbook.xlsx in the project root (or replaces Sheet 2 if it
already exists, preserving Sheets 1/3/4/5/6/7 added by future days' scripts).

Columns per Day 2 spec (15):
  player name, age, current club, league, primary position, TM value (€),
  contract end date, last fee paid (€), minutes_last_18m, appearances_last_18m,
  agency, right_priced, finished_product, contract_leveraged,
  owning_club_must_sell_score (NULL until Day 3)
"""

from datetime import datetime
import sqlite3
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import config
import club_display
import player_display

WORKBOOK_PATH = Path("BrokerageWorkbook.xlsx")
SHEET_NAME = "Player Universe"
SHEET_POSITION = 1  # 0-indexed; Day 1 spec wants this as Sheet 2 of the final book

HEADERS = [
    "Player",
    "Age",
    "Current Club",
    "League",
    "Primary Position",
    "TM Value (€)",
    "Contract End",
    "Last Fee Paid (€)",
    "Minutes (last 18m)",
    "Appearances (last 18m)",
    "Agency",
    "Right Priced",
    "Finished Product",
    "Contract Leveraged",
    "Owning Club Must Sell Score",
]

# Per-column widths, indexed in HEADERS order. Tuned for readability.
COLUMN_WIDTHS = [26, 5, 44, 22, 22, 14, 12, 14, 14, 14, 28, 12, 14, 16, 22]


def fetch_rows():
    with sqlite3.connect(config.SQLITE_FILE) as con:
        con.row_factory = sqlite3.Row
        return con.execute("""
            SELECT
                pu.player_id,
                pu.name,
                pu.age,
                pu.current_club,
                pu.current_club_id,
                pu.parent_club,
                pu.parent_club_id,
                pu.on_loan,
                pu.league,
                pu.sub_position,
                pu.current_tm_value_eur,
                pu.contract_end_date,
                pu.last_fee_paid_eur,
                pu.minutes_last_18m,
                pu.appearances_last_18m,
                pu.agency,
                pu.right_priced,
                pu.finished_product,
                pu.contract_leveraged,
                cp.total_pressure_score AS owning_club_must_sell_score
            FROM player_universe pu
            LEFT JOIN club_pressure cp ON cp.club_id = pu.parent_club_id
            WHERE pu.brokerage_eligible = 1
            ORDER BY pu.league, pu.current_tm_value_eur DESC, pu.name
        """).fetchall()


def _open_or_create_workbook(path: Path) -> Workbook:
    if path.exists():
        wb = load_workbook(path)
        # Drop the old Player Universe sheet if present; we replace it on every run.
        if SHEET_NAME in wb.sheetnames:
            del wb[SHEET_NAME]
        return wb
    wb = Workbook()
    # Workbook() creates a default sheet — rename it to a placeholder so it survives
    # to become a future sheet later; we don't delete it here.
    wb.active.title = "Sheet 1 (placeholder)"
    return wb


def main() -> None:
    rows = fetch_rows()
    print(f"Fetched {len(rows):,} rows from player_universe")

    # Load display-name maps (Kill List + manual overrides preserved across runs)
    display_map = club_display.load_display_map(WORKBOOK_PATH)
    player_map = player_display.load_display_map(WORKBOOK_PATH)

    wb = _open_or_create_workbook(WORKBOOK_PATH)
    ws = wb.create_sheet(SHEET_NAME, SHEET_POSITION)

    # Header row.
    bold = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3864")
    center = Alignment(horizontal="center", vertical="center")
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center

    # Data rows.
    for row_idx, r in enumerate(rows, start=2):
        current_display = club_display.display_for(r["current_club_id"], r["current_club"], display_map)
        parent_display  = club_display.display_for(r["parent_club_id"],  r["parent_club"],  display_map)
        current_club_cell = (
            f"{current_display} (on loan from {parent_display})"
            if r["on_loan"] and r["parent_club"] else current_display
        )
        league_display = config.LEAGUE_DISPLAY.get(r["league"], r["league"])

        player_name_display = player_display.display_for(r["player_id"], r["name"], player_map)
        ws.cell(row=row_idx, column=1, value=player_name_display)
        ws.cell(row=row_idx, column=2, value=r["age"])
        ws.cell(row=row_idx, column=3, value=current_club_cell)
        ws.cell(row=row_idx, column=4, value=league_display)
        ws.cell(row=row_idx, column=5, value=r["sub_position"])
        ws.cell(row=row_idx, column=6, value=r["current_tm_value_eur"])
        # Convert ISO date string -> datetime so Excel sees it as a real date.
        ce = r["contract_end_date"]
        if ce:
            ws.cell(row=row_idx, column=7, value=datetime.fromisoformat(ce).date())
        ws.cell(row=row_idx, column=8, value=r["last_fee_paid_eur"])
        ws.cell(row=row_idx, column=9, value=r["minutes_last_18m"])
        ws.cell(row=row_idx, column=10, value=r["appearances_last_18m"])
        ws.cell(row=row_idx, column=11, value=r["agency"])
        # Flag columns: write True/False so Excel filters work naturally.
        ws.cell(row=row_idx, column=12, value=bool(r["right_priced"]))
        ws.cell(row=row_idx, column=13, value=bool(r["finished_product"]))
        ws.cell(row=row_idx, column=14, value=bool(r["contract_leveraged"]))
        ws.cell(row=row_idx, column=15, value=r["owning_club_must_sell_score"])

    # Number formats.
    for row_idx in range(2, len(rows) + 2):
        ws.cell(row=row_idx, column=6).number_format = '"€"#,##0'   # TM Value
        ws.cell(row=row_idx, column=7).number_format = "yyyy-mm-dd" # Contract End
        ws.cell(row=row_idx, column=8).number_format = '"€"#,##0'   # Last Fee
        ws.cell(row=row_idx, column=9).number_format = "#,##0"      # Minutes
        ws.cell(row=row_idx, column=10).number_format = "#,##0"     # Appearances
        ws.cell(row=row_idx, column=15).number_format = "0.0"       # Must Sell Score

    # Column widths.
    for col_idx, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header, autofilter the data range.
    ws.freeze_panes = "A2"
    last_col_letter = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f"A1:{last_col_letter}{len(rows) + 1}"

    wb.save(WORKBOOK_PATH)
    print(f"Saved {WORKBOOK_PATH.resolve()}")
    print(f"  Sheet: '{SHEET_NAME}' — {len(rows)} rows")


if __name__ == "__main__":
    main()
