"""
Step 08 — Compute the five Seller-Pressure components per club.

Reads:  senior_roster (built by 07), clubs.net_transfer_record (from dcaribou),
        data/manual_flags.csv (seeded on first run; edited by user thereafter).
Writes: club_pressure rows updated with all five components, total_pressure_score,
        scoring_basis. Also assigns position_bucket on senior_roster.

Components (weighted to 0-100 total):
  1. Contract leverage (20%) — % of senior squad in final 2 years.
       Top tier:  minutes-weighted (minutes share of leveraged players).
       2nd tier:  headcount-weighted (no minutes data); scoring_basis flags this.
  2. Squad oversupply  (20%) — count of buckets over threshold ÷ 10 × 100.
       Bucket thresholds:
         GK ≥4 | CB ≥5 | LB ≥3 | RB ≥3 | AM ≥3 | LW ≥3 | RW ≥3 | ST_CF ≥4
       DM/CM combined rule: both ≥3 → both fire; otherwise neither fires.
  3. Net spend (20%) — parse clubs.net_transfer_record (TM's pre-rolled string).
       NEGATIVE record = net buyer → score on this component.
       POSITIVE/zero  = net seller → 0.
       Normalised within each league: |net| ÷ max(|net|) × 100.
       2nd tier: 0 (no data), scoring_basis flags this.
  4. Manager change flag (15%) — manual boolean × 100 from CSV.
  5. Public must-sell flag (25%) — manual boolean × 100 from CSV.

Total = 0.20·C1 + 0.20·C2 + 0.20·C3 + 0.15·C4 + 0.25·C5.

Manual flags storage: data/manual_flags.xlsx (was CSV pre-Day 4; see 18_manual_flags_excel.py).
  - Seeded on first run with every club at 0/0.
  - User edits flag columns in Excel. Re-running this script reads back the values.
  - On rerun, new clubs (e.g. promoted/relegated) are appended at 0/0;
    existing rows are preserved verbatim.
  - Falls back to manual_flags.csv if xlsx doesn't exist yet (transition support).
"""

import csv
import sqlite3
import sys
from collections import defaultdict
from pathlib import Path

import duckdb
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent))
import config
from _position_buckets import POSITION_BUCKETS, bucket_for  # noqa: E402

MANUAL_FLAGS_XLSX = Path("data/manual_flags.xlsx")
MANUAL_FLAGS_CSV = Path("data/manual_flags.csv")  # legacy fallback during migration

OVERSUPPLY_THRESHOLDS: dict[str, int] = {
    "GK": 4, "CB": 5, "LB": 3, "RB": 3,
    "DM": 3, "CM": 3, "AM": 3, "LW": 3, "RW": 3, "ST_CF": 4,
}
COMBINED_RULE = ("DM", "CM")

WEIGHTS = {
    "contract_leverage": 0.20,
    "squad_oversupply":  0.20,
    "net_spend":         0.20,
    "manager_change":    0.15,
    "public_must_sell":  0.25,
}


# ─── Helpers ────────────────────────────────────────────────────────────────────

def parse_net_transfer_record(s: str | None) -> float | None:
    """Parse TM's net transfer string into a signed euro value.

    Examples:
        '+€49.00m'  →  +49_000_000  (net SOLD)
        '€-300k'    →  -300_000     (net BOUGHT)
        '+€300k'    →  +300_000
        '+-0' / '-' / ''  →  0
    Returns None if unparseable.
    """
    if s is None:
        return None
    s = s.strip()
    if not s or s in {"+-0", "-", "0", "€0"}:
        return 0.0
    sign = -1.0 if "-" in s else 1.0
    # Strip non-numeric/unit characters
    import re
    m = re.search(r"([\d.,]+)\s*([mk]?)", s.lower())
    if not m:
        return None
    try:
        n = float(m.group(1).replace(",", "."))
    except ValueError:
        return None
    unit = m.group(2)
    multiplier = 1_000_000 if unit == "m" else (1_000 if unit == "k" else 1)
    return sign * n * multiplier


def squad_oversupply_count(bucket_counts: dict[str, int]) -> tuple[int, list[str]]:
    """Returns (count_of_positions_over_threshold, list_of_buckets_fired).

    DM/CM are combined: both fire only when both ≥ their threshold.
    """
    fired: list[str] = []
    dm_ok = bucket_counts.get("DM", 0) >= OVERSUPPLY_THRESHOLDS["DM"]
    cm_ok = bucket_counts.get("CM", 0) >= OVERSUPPLY_THRESHOLDS["CM"]
    combined_fires = dm_ok and cm_ok
    for bucket, threshold in OVERSUPPLY_THRESHOLDS.items():
        if bucket in COMBINED_RULE:
            if combined_fires:
                fired.append(bucket)
        else:
            if bucket_counts.get(bucket, 0) >= threshold:
                fired.append(bucket)
    return len(fired), fired


def _read_flags_from_xlsx() -> dict[str, dict]:
    """Read flag values from manual_flags.xlsx. Returns {} if file missing."""
    if not MANUAL_FLAGS_XLSX.exists():
        return {}
    wb = load_workbook(MANUAL_FLAGS_XLSX, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {}
    headers = [str(h) if h is not None else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}
    if "club_id" not in idx:
        return {}
    out: dict[str, dict] = {}
    for row in rows[1:]:
        cid = row[idx["club_id"]]
        if cid is None:
            continue
        out[str(cid)] = {
            "manager_change":   int(row[idx["manager_change_flag"]] or 0),
            "public_must_sell": int(row[idx["public_must_sell_flag"]] or 0),
            "notes":            row[idx.get("notes_source", -1)] if idx.get("notes_source", -1) >= 0 else "",
        }
    return out


def _read_flags_from_csv() -> dict[str, dict]:
    """Legacy CSV reader, only used if xlsx doesn't exist yet."""
    if not MANUAL_FLAGS_CSV.exists():
        return {}
    out: dict[str, dict] = {}
    with MANUAL_FLAGS_CSV.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            out[row["club_id"]] = {
                "manager_change":   int(row.get("manager_change", "0") or "0"),
                "public_must_sell": int(row.get("public_must_sell", "0") or "0"),
                "notes":            row.get("notes", "") or "",
            }
    return out


def load_or_seed_manual_flags(all_clubs: list[dict]) -> dict[str, dict]:
    """Returns club_id → {manager_change, public_must_sell, notes}.

    Reads from manual_flags.xlsx (preferred) or falls back to manual_flags.csv
    during the Day 4 migration window. New clubs default to 0/0/''.

    NOTE: writing back to the xlsx is handled by scripts/18_manual_flags_excel.py,
    not here — this script is read-only for flags. Run 18 after 08 to refresh
    the xlsx with the latest club_pressure component scores.
    """
    existing = _read_flags_from_xlsx() or _read_flags_from_csv()

    out_flags: dict[str, dict] = {}
    new_clubs = 0
    for c in all_clubs:
        cid = c["club_id"]
        prev = existing.get(cid)
        if prev is None:
            prev = {"manager_change": 0, "public_must_sell": 0, "notes": ""}
            new_clubs += 1
        out_flags[cid] = prev

    src = "xlsx" if MANUAL_FLAGS_XLSX.exists() else ("csv" if MANUAL_FLAGS_CSV.exists() else "(none — all 0/0)")
    if new_clubs:
        print(f"  Manual flags: read from {src}; {new_clubs} new club(s) defaulted to 0/0 "
              "(run 18_manual_flags_excel.py to persist them)")
    else:
        print(f"  Manual flags: read from {src} ({len(existing)} prior rows)")
    return out_flags


# ─── Mapping audit print ────────────────────────────────────────────────────────

def print_mapping_audit(con: sqlite3.Connection) -> None:
    print("Position bucket mapping (TM sub_position → bucket):")
    rows = con.execute("""
        SELECT sub_position, COUNT(*) AS n
        FROM senior_roster
        GROUP BY sub_position
        ORDER BY n DESC
    """).fetchall()
    total = sum(n for _, n in rows)
    print(f"  {'sub_position':30s} {'bucket':8s} {'count':>7s}  {'%':>5s}")
    print(f"  {'-'*30} {'-'*8} {'-'*7}  {'-'*5}")
    judgment = {"Second Striker", "Left Midfield", "Right Midfield"}
    for sp, n in rows:
        bucket = bucket_for(sp)
        flag = " *" if sp in judgment else ""
        pct = 100 * n / total if total else 0
        print(f"  {sp or '(none)':30s} {bucket or '— drop':8s} {n:>7d}  {pct:>4.1f}%{flag}")
    print("  * judgment-call mapping (see CLAUDE/Day 3 plan).")


# ─── Component computations ────────────────────────────────────────────────────

def compute_contract_leverage(con: sqlite3.Connection) -> dict[str, tuple[float, str]]:
    """club_id → (component_score_0_100, basis_note)."""
    out: dict[str, tuple[float, str]] = {}

    # Top tier (data_source='dcaribou'): minutes-weighted, fallback to headcount if no minutes.
    rows = con.execute("""
        SELECT club_id,
               COALESCE(SUM(minutes_last_18m), 0) AS total_min,
               COALESCE(SUM(CASE WHEN contract_leveraged=1
                                 THEN minutes_last_18m ELSE 0 END), 0) AS leveraged_min,
               COUNT(*) AS headcount,
               SUM(CASE WHEN contract_leveraged=1 THEN 1 ELSE 0 END) AS leveraged_hc
        FROM senior_roster
        WHERE data_source = 'dcaribou'
        GROUP BY club_id
    """).fetchall()
    for cid, total_min, lev_min, hc, lev_hc in rows:
        if total_min > 0:
            score = 100.0 * lev_min / total_min
            basis = ""
        elif hc > 0:
            score = 100.0 * lev_hc / hc
            basis = "headcount-weighted (no minutes)"
        else:
            score = 0.0
            basis = "no squad data"
        out[cid] = (score, basis)

    # Second tier (data_source='tm_squad_scrape'): headcount-weighted.
    rows = con.execute("""
        SELECT club_id, COUNT(*) AS hc,
               SUM(CASE WHEN contract_leveraged=1 THEN 1 ELSE 0 END) AS lev_hc
        FROM senior_roster
        WHERE data_source = 'tm_squad_scrape'
        GROUP BY club_id
    """).fetchall()
    for cid, hc, lev_hc in rows:
        score = 100.0 * lev_hc / hc if hc else 0.0
        out[cid] = (score, "headcount-weighted (second tier, no minutes)")

    return out


def compute_squad_oversupply(con: sqlite3.Connection) -> dict[str, tuple[float, list[str]]]:
    """club_id → (score_0_100, list_of_buckets_oversupplied).

    Filters to the active first-team pool:
      - Top tier (dcaribou): players with minutes_last_18m > 0. Excludes loanees
        registered to the club but not playing, plus academy/youth who clutter the
        roster (Crystal Palace had 13 GKs / 20 CBs before this filter).
      - Second tier (tm_squad_scrape): all players. We have no minutes data so
        the kader page is treated as authoritative on first-team membership.
    """
    rows = con.execute("""
        SELECT club_id, position_bucket, COUNT(*) AS n
        FROM senior_roster
        WHERE position_bucket IS NOT NULL
          AND (
                (data_source = 'dcaribou' AND minutes_last_18m IS NOT NULL AND minutes_last_18m > 0)
             OR (data_source = 'tm_squad_scrape')
          )
        GROUP BY club_id, position_bucket
    """).fetchall()
    per_club: dict[str, dict[str, int]] = defaultdict(dict)
    for cid, bucket, n in rows:
        per_club[cid][bucket] = n
    out: dict[str, tuple[float, list[str]]] = {}
    for cid, bucket_counts in per_club.items():
        count, fired = squad_oversupply_count(bucket_counts)
        out[cid] = (count * 10.0, fired)
    return out


def compute_net_spend(con: sqlite3.Connection, dcaribou_ids: list[str]) -> dict[str, tuple[float | None, str]]:
    """club_id → (score_0_100 or None, raw_record_string).

    Pulls clubs.net_transfer_record directly from dcaribou DuckDB.
    Normalises within each league: clubs with negative records (net buyers)
    are scaled to the league's max |negative|.
    """
    src = duckdb.connect(config.DUCKDB_FILE, read_only=True)
    ids_sql = "(" + ",".join(f"'{i}'" for i in dcaribou_ids) + ")"
    rows = src.execute(f"""
        SELECT club_id, domestic_competition_id, net_transfer_record
        FROM clubs
        WHERE domestic_competition_id IN {ids_sql}
          AND last_season >= '2025'
    """).fetchall()

    parsed: list[tuple[str, str, float | None, str]] = []
    for cid, lid, raw in rows:
        val = parse_net_transfer_record(raw)
        parsed.append((str(cid), lid, val, raw or ""))

    # Per-league max-|negative| for normalisation.
    league_max_neg: dict[str, float] = {}
    for _, lid, val, _ in parsed:
        if val is not None and val < 0:
            league_max_neg[lid] = max(league_max_neg.get(lid, 0.0), -val)

    out: dict[str, tuple[float | None, str]] = {}
    for cid, lid, val, raw in parsed:
        if val is None or val >= 0:
            score: float | None = 0.0
        else:
            mx = league_max_neg.get(lid, 0.0)
            score = (100.0 * (-val) / mx) if mx > 0 else 0.0
        out[cid] = (score, raw)
    return out


# ─── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    snapshot = config.SNAPSHOT_DATE
    dcaribou_ids = [l["id"] for l in config.LEAGUES if not l.get("external")]
    external_ids = [l["id"] for l in config.LEAGUES if l.get("external")]

    with sqlite3.connect(config.SQLITE_FILE) as con:
        # 1. Mapping audit.
        print_mapping_audit(con)
        print()

        # 2. Assign position_bucket on senior_roster.
        rows = con.execute("SELECT player_id, sub_position FROM senior_roster").fetchall()
        for pid, sp in rows:
            con.execute(
                "UPDATE senior_roster SET position_bucket = ? WHERE player_id = ?",
                (bucket_for(sp), pid),
            )
        con.commit()
        n_bucketed = con.execute(
            "SELECT COUNT(*) FROM senior_roster WHERE position_bucket IS NOT NULL"
        ).fetchone()[0]
        n_total = con.execute("SELECT COUNT(*) FROM senior_roster").fetchone()[0]
        print(f"position_bucket assigned: {n_bucketed:,}/{n_total:,} senior_roster rows")
        print()

        # 3. Make sure club_pressure has rows for every club in senior_roster
        #    (second-tier clubs aren't seeded by 03).
        all_clubs_rows = con.execute("""
            SELECT DISTINCT club_id, club_name, league_id, league
            FROM senior_roster
            WHERE club_id IS NOT NULL
        """).fetchall()
        all_clubs = [
            {"club_id": r[0], "club_name": r[1], "league_id": r[2], "league": r[3]}
            for r in all_clubs_rows
        ]
        # Seed/refresh club_pressure from senior_roster. Use INSERT OR IGNORE to
        # create new rows (preserving manual flags at 0/0), then UPDATE name/league/
        # league_id on every row — this catches clubs that have changed league
        # between dcaribou snapshots (promotion/relegation), which otherwise leaves
        # stale league labels on Sheet 3 and corrupts the per-league net_spend
        # normalisation in component 3.
        existing_meta = {
            cid: (lg, lid) for cid, lg, lid in
            con.execute("SELECT club_id, league, league_id FROM club_pressure").fetchall()
        }
        n_inserted = 0
        relabelled: list[tuple[str, str, str, str]] = []  # (club_name, old_lid, new_lid, name)
        for c in all_clubs:
            cid = c["club_id"]
            if cid not in existing_meta:
                con.execute("""
                    INSERT INTO club_pressure
                        (club_id, name, league, league_id,
                         manager_change_flag, public_must_sell_flag, snapshot_date)
                    VALUES (?, ?, ?, ?, 0, 0, ?)
                """, (cid, c["club_name"], c["league"], c["league_id"], str(snapshot)))
                n_inserted += 1
            else:
                old_lg, old_lid = existing_meta[cid]
                if old_lid != c["league_id"]:
                    relabelled.append((c["club_name"], old_lid, c["league_id"], c["club_name"]))
                con.execute("""
                    UPDATE club_pressure
                    SET name = ?, league = ?, league_id = ?
                    WHERE club_id = ?
                """, (c["club_name"], c["league"], c["league_id"], cid))
        if n_inserted:
            print(f"  Inserted {n_inserted} new club_pressure row(s)")
        if relabelled:
            print(f"  Relabelled {len(relabelled)} club_pressure row(s) where league changed since prior run:")
            for name, old_lid, new_lid, _ in sorted(relabelled, key=lambda x: (x[1], x[2], x[0])):
                print(f"    {name:36s} {old_lid:5s} → {new_lid}")
        # Drop any pre-existing club_pressure rows for clubs that aren't currently in
        # senior_roster. These are stale dcaribou entries (e.g. clubs relegated out of
        # our 19-league set since the dataset was last refreshed) that pollute Sheet 3.
        active_ids = {c["club_id"] for c in all_clubs}
        existing_ids = {r[0] for r in con.execute("SELECT club_id FROM club_pressure")}
        stale_ids = existing_ids - active_ids
        if stale_ids:
            con.executemany(
                "DELETE FROM club_pressure WHERE club_id = ?",
                [(cid,) for cid in stale_ids],
            )
            print(f"  Dropped {len(stale_ids)} stale club_pressure rows (no current senior_roster)")
        con.commit()

        # 4. Manual flags CSV.
        flags = load_or_seed_manual_flags(all_clubs)

        # 5. Compute components.
        cl_scores = compute_contract_leverage(con)
        so_scores = compute_squad_oversupply(con)
        ns_scores = compute_net_spend(con, dcaribou_ids)

        # 6. Compose total + scoring_basis per club; write back.
        clubs_in_cp = con.execute("""
            SELECT club_id, league_id FROM club_pressure
        """).fetchall()
        updates = []
        for cid, lid in clubs_in_cp:
            cl_score, cl_basis = cl_scores.get(cid, (0.0, "no squad data"))
            so_score, so_buckets = so_scores.get(cid, (0.0, []))
            if lid in external_ids:
                ns_score, ns_raw = 0.0, ""
                ns_basis = "net_spend: unavailable (dcaribou top-tier only)"
            else:
                ns_score, ns_raw = ns_scores.get(cid, (0.0, ""))
                ns_basis = ""
            mc = flags.get(cid, {}).get("manager_change", 0)
            ps = flags.get(cid, {}).get("public_must_sell", 0)
            total = (
                WEIGHTS["contract_leverage"] * cl_score
                + WEIGHTS["squad_oversupply"] * so_score
                + WEIGHTS["net_spend"] * ns_score
                + WEIGHTS["manager_change"] * (mc * 100.0)
                + WEIGHTS["public_must_sell"] * (ps * 100.0)
            )
            basis_parts = [p for p in (cl_basis, ns_basis) if p]
            if so_buckets:
                basis_parts.append(f"oversupplied: {', '.join(so_buckets)}")
            scoring_basis = "; ".join(basis_parts) if basis_parts else ""
            updates.append((
                round(cl_score, 1), round(so_score, 1), round(ns_score, 1),
                int(mc), int(ps), round(total, 1), scoring_basis, cid,
            ))
        con.executemany("""
            UPDATE club_pressure SET
                contract_leverage_score = ?,
                squad_oversupply_score  = ?,
                net_spend_score         = ?,
                manager_change_flag     = ?,
                public_must_sell_flag   = ?,
                total_pressure_score    = ?,
                scoring_basis           = ?
            WHERE club_id = ?
        """, updates)
        con.commit()

        # 7. Print top 20 by total_pressure_score.
        print()
        print("Top 20 clubs by total_pressure_score:")
        print(f"  {'club':36s} {'lg':5s} {'CL':>5s} {'SO':>5s} {'NS':>5s} {'MC':>3s} {'PS':>3s} {'TOT':>6s}")
        rows = con.execute("""
            SELECT name, league_id,
                   contract_leverage_score, squad_oversupply_score, net_spend_score,
                   manager_change_flag, public_must_sell_flag, total_pressure_score
            FROM club_pressure
            WHERE total_pressure_score IS NOT NULL
            ORDER BY total_pressure_score DESC
            LIMIT 20
        """).fetchall()
        for r in rows:
            print(f"  {(r[0] or '')[:36]:36s} {r[1]:5s} "
                  f"{r[2] or 0:>5.1f} {r[3] or 0:>5.1f} {r[4] or 0:>5.1f} "
                  f"{r[5] or 0:>3d} {r[6] or 0:>3d} {r[7] or 0:>6.1f}")

        n_cp = con.execute("SELECT COUNT(*) FROM club_pressure").fetchone()[0]
        print(f"\n{n_cp} clubs scored.")
        flags_path = MANUAL_FLAGS_XLSX if MANUAL_FLAGS_XLSX.exists() else MANUAL_FLAGS_CSV
        print(f"Manual flags source: {flags_path.resolve() if flags_path.exists() else '(none yet)'}")


if __name__ == "__main__":
    main()
