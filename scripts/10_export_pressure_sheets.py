"""
Step 10 — Write Sheet 3 (Seller Pressure) and Sheet 4 (Sellable Asset Ledger)
into BrokerageWorkbook.xlsx.

Sheet 3 — Seller Pressure
    One row per club across all 19 leagues. Columns:
      club, league, contract_leverage_score, squad_oversupply_score,
      net_spend_score, manager_change_flag, public_must_sell_flag,
      total_pressure_score, top_3_likely_to_move, scoring_basis

Sheet 4 — Sellable Asset Ledger
    Subset of player_universe where at least one of right_priced/finished_product/
    contract_leveraged is true (or finished_product is NULL = unknown).
    Joined with owning-club pressure. Columns:
      player, age, primary_position, position_bucket, current_tm_value_eur,
      contract_years_remaining, current_club, owning_club_pressure_score,
      sellability_score, agency, wage_estimate (NULL until Day 6)
"""

from datetime import datetime
import sqlite3
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import config
import club_display as cd  # alias to avoid clashing with a local var named club_display
import player_display as pd_disp

WORKBOOK_PATH = Path("BrokerageWorkbook.xlsx")

SHEET3_NAME = "Seller Pressure"
SHEET4_NAME = "Sellable Assets"

SHEET3_HEADERS = [
    "Club", "League",
    "Contract Leverage (0-100)", "Squad Oversupply (0-100)", "Net Spend (0-100)",
    "Manager Change (0/1)", "Public Must-Sell (0/1)",
    "Total Pressure (0-100)", "Top 3 Likely to Move", "Scoring Basis",
]
SHEET3_WIDTHS = [34, 22, 18, 18, 14, 14, 14, 16, 36, 40]

SHEET4_HEADERS = [
    "Player", "Age", "Primary Position", "Position Bucket",
    "TM Value (€)", "Contract Years Remaining", "Current Club",
    "Owning Club Pressure", "Sellability Score", "Agency", "Wage Estimate (€/wk)",
    "Scoring Notes",
]
SHEET4_WIDTHS = [26, 5, 18, 14, 14, 12, 44, 16, 14, 28, 16, 28]


def _open_workbook(path: Path) -> Workbook:
    if path.exists():
        wb = load_workbook(path)
        for sn in (SHEET3_NAME, SHEET4_NAME):
            if sn in wb.sheetnames:
                del wb[sn]
        return wb
    wb = Workbook()
    wb.active.title = "Sheet 1 (placeholder)"
    return wb


def _write_header(ws, headers: list[str]) -> None:
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F3864")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = bold
        cell.fill = fill
        cell.alignment = center


def write_seller_pressure_sheet(wb: Workbook, con: sqlite3.Connection) -> int:
    ws = wb.create_sheet(SHEET3_NAME, 2)  # position after Player Universe (Sheet 2)
    _write_header(ws, SHEET3_HEADERS)
    display_map = cd.load_display_map(WORKBOOK_PATH)

    # Build official_name → display_name map for swapping the top_3_likely_to_move text
    player_display_by_id = pd_disp.load_display_map(WORKBOOK_PATH)
    official_to_display: dict[str, str] = {}
    for pid, official in con.execute("SELECT player_id, name FROM player_universe"):
        dn = player_display_by_id.get(int(pid))
        if dn:
            official_to_display[official] = dn

    rows = con.execute("""
        SELECT club_id, name, league,
               contract_leverage_score, squad_oversupply_score, net_spend_score,
               manager_change_flag, public_must_sell_flag,
               total_pressure_score, top_3_likely_to_move, scoring_basis
        FROM club_pressure
        ORDER BY total_pressure_score DESC NULLS LAST, league, name
    """).fetchall()
    # SHEET3 headers (positional): name, league, ..., total_pressure_score, top_3_likely_to_move, scoring_basis
    # top_3_likely_to_move is at column index 9 (0-based) from rest after splitting club_id+name
    # In `rest`: 0=league 1=cl 2=so 3=ns 4=mgr 5=ps 6=total 7=top_3 8=scoring_basis
    TOP_3_REST_IDX = 7

    for ri, r in enumerate(rows, start=2):
        club_id, official_name, *rest = r
        display_name = cd.display_for(club_id, official_name, display_map)
        ws.cell(row=ri, column=1, value=display_name)
        for ci, val in enumerate(rest, start=2):
            # Swap player names in the top_3 text field
            if ci == 2 + TOP_3_REST_IDX:
                val = pd_disp.apply_to_text_list(val, official_to_display)
            ws.cell(row=ri, column=ci, value=val)
    # Number formats.
    for ri in range(2, len(rows) + 2):
        for ci in (3, 4, 5, 8):
            ws.cell(row=ri, column=ci).number_format = "0.0"
    for ci, w in enumerate(SHEET3_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(SHEET3_HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"
    return len(rows)


def write_sellable_assets_sheet(wb: Workbook, con: sqlite3.Connection) -> int:
    ws = wb.create_sheet(SHEET4_NAME, 3)
    _write_header(ws, SHEET4_HEADERS)
    snapshot = datetime.fromisoformat(str(config.SNAPSHOT_DATE)).date()
    # Join on PARENT_club_id (Day 3.5 fix). Detect whether the parent is inside our
    # 19-league coverage via LEFT JOIN — NULL pressure means "parent outside coverage".
    # Load manual wages (data/manual_wages.xlsx) if present — populated by Ryan
    # via scripts/26_manual_wages_template.py. Missing entries → blank wage cell.
    player_wages: dict[int, float] = {}
    wages_path = Path("data/manual_wages.xlsx")
    if wages_path.exists():
        from openpyxl import load_workbook as _lw
        wb_w = _lw(wages_path, data_only=True, read_only=True)
        ws_w = wb_w.active
        all_w = list(ws_w.iter_rows(values_only=True))
        wb_w.close()
        if all_w:
            hs = [str(h) if h is not None else "" for h in all_w[0]]
            idx_w = {h: i for i, h in enumerate(hs)}
            if "player_id" in idx_w and "wage_pw_eur" in idx_w:
                for r in all_w[1:]:
                    if not r or r[idx_w["player_id"]] is None:
                        continue
                    try:
                        pid = int(r[idx_w["player_id"]])
                        wage = r[idx_w["wage_pw_eur"]]
                        if wage not in (None, ""):
                            player_wages[pid] = float(wage)
                    except (TypeError, ValueError):
                        continue

    display_map = cd.load_display_map(WORKBOOK_PATH)
    player_map = pd_disp.load_display_map(WORKBOOK_PATH)

    rows = con.execute("""
        SELECT pu.player_id, pu.name, pu.age, pu.sub_position, pu.position_bucket,
               pu.current_tm_value_eur, pu.contract_end_date,
               pu.current_club, pu.current_club_id, pu.parent_club, pu.on_loan, pu.parent_club_id,
               cp.total_pressure_score, pu.sellability_score, pu.agency
        FROM player_universe pu
        LEFT JOIN club_pressure cp ON cp.club_id = pu.parent_club_id
        WHERE pu.sellability_status = 'sellable_now'
        ORDER BY pu.sellability_score DESC NULLS LAST, pu.name
    """).fetchall()
    for ri, r in enumerate(rows, start=2):
        (pid, name, age, sub_pos, bucket, mv, ce, current_club, current_club_id, parent_club, on_loan,
         parent_id, parent_pressure, sell, agency) = r
        years_remaining = None
        if ce:
            try:
                ce_d = datetime.fromisoformat(ce).date()
                years_remaining = round((ce_d - snapshot).days / 365.25, 2)
            except (TypeError, ValueError):
                pass
        cc_display = cd.display_for(current_club_id, current_club, display_map)
        pc_display = cd.display_for(parent_id, parent_club, display_map)
        club_cell = (
            f"{cc_display} (on loan from {pc_display})"
            if on_loan and parent_club else cc_display
        )
        # Pressure cell: numeric if parent in coverage; "—" if outside; same as
        # multiplicative=0 branch in the sellability formula.
        parent_in_coverage = parent_pressure is not None
        pressure_cell: object = parent_pressure if parent_in_coverage else "—"
        notes_parts: list[str] = []
        if on_loan and not parent_in_coverage:
            notes_parts.append("parent-outside-coverage")
        elif on_loan:
            notes_parts.append("loan: +15 if finished_product=true")
        ws.cell(row=ri, column=1, value=pd_disp.display_for(pid, name, player_map))
        ws.cell(row=ri, column=2, value=age)
        ws.cell(row=ri, column=3, value=sub_pos)
        ws.cell(row=ri, column=4, value=bucket)
        ws.cell(row=ri, column=5, value=mv)
        ws.cell(row=ri, column=6, value=years_remaining)
        ws.cell(row=ri, column=7, value=club_cell)
        ws.cell(row=ri, column=8, value=pressure_cell)
        ws.cell(row=ri, column=9, value=sell)
        ws.cell(row=ri, column=10, value=agency)
        ws.cell(row=ri, column=11, value=player_wages.get(pid))  # wage_pw from manual_wages.xlsx
        ws.cell(row=ri, column=12, value="; ".join(notes_parts) or None)
        if not parent_in_coverage:
            # Centre-align the em-dash so it reads as "data gap" not "zero".
            ws.cell(row=ri, column=8).alignment = Alignment(horizontal="center")
    for ri in range(2, len(rows) + 2):
        ws.cell(row=ri, column=5).number_format = '"€"#,##0'
        ws.cell(row=ri, column=6).number_format = "0.00"
        # Column 8 (pressure) is only numeric-formatted when the cell holds a number;
        # leave it default for "—" string cells.
        cell8 = ws.cell(row=ri, column=8)
        if isinstance(cell8.value, (int, float)):
            cell8.number_format = "0.0"
        ws.cell(row=ri, column=9).number_format = "0.00"
        ws.cell(row=ri, column=11).number_format = '"€"#,##0'
    for ci, w in enumerate(SHEET4_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(SHEET4_HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"
    return len(rows)


def main() -> None:
    wb = _open_workbook(WORKBOOK_PATH)
    with sqlite3.connect(config.SQLITE_FILE) as con:
        n3 = write_seller_pressure_sheet(wb, con)
        n4 = write_sellable_assets_sheet(wb, con)
    wb.save(WORKBOOK_PATH)
    print(f"Saved {WORKBOOK_PATH.resolve()}")
    print(f"  Sheet 3 '{SHEET3_NAME}' — {n3} clubs")
    print(f"  Sheet 4 '{SHEET4_NAME}' — {n4} sellable assets")


if __name__ == "__main__":
    main()
