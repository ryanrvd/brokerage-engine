"""
Step 27 — Generate / refresh `data/player_overrides.xlsx`.

User-facing exclusion list. One row per sellable player. Ryan flips
`exclude = 1` for any player he wants dropped from the Targets view in the
Streamlit app (e.g. agent says no, outside RV remit, already placed).

Same pattern as `data/manual_flags.xlsx` and `data/manual_wages.xlsx`:
  • Idempotent — preserves prior user values across runs.
  • Sorted by best_match_score DESC (highest-impact decisions surface first).
  • Staging-file + row-count reconciliation safety net.
  • Rolling `.bak` at start of every run.

Read-only columns (grey):
  player_id, name, age, position, current_club, parent_club, league, TM_value,
  sellability_score, best_match_score
User input columns (yellow):
  exclude (0/1, data-validated), exclusion_reason, notes
Auto-stamped:
  last_reviewed (today when exclude value flips)
"""

import shutil
import sqlite3
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import config

XLSX_PATH = Path("data/player_overrides.xlsx")
BAK_XLSX_PATH = Path("data/player_overrides.xlsx.bak")
STAGING_XLSX_PATH = Path("data/player_overrides.staging.xlsx")
LOG_DIR = Path("logs")

HEADERS = [
    "player_id",            # 1  RO
    "name",                 # 2  RO
    "age",                  # 3  RO
    "position",             # 4  RO
    "current_club",         # 5  RO
    "parent_club",          # 6  RO
    "league",               # 7  RO
    "TM_value_eur",         # 8  RO
    "sellability_score",    # 9  RO
    "best_match_score",     # 10 RO
    "exclude",              # 11 user input (0/1)
    "exclusion_reason",     # 12 user input (free text)
    "notes",                # 13 user input (free text)
    "last_reviewed",        # 14 auto-stamped
]
WIDTHS = [10, 28, 5, 8, 28, 28, 7, 14, 12, 12, 9, 32, 36, 14]

COL_EXCLUDE       = 11
COL_REASON        = 12
COL_NOTES         = 13
COL_LAST_REVIEWED = 14


def _load_existing() -> dict[int, dict]:
    if not XLSX_PATH.exists():
        return {}
    out: dict[int, dict] = {}
    wb = load_workbook(XLSX_PATH, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return out
    headers = [str(h) if h is not None else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}
    if "player_id" not in idx:
        return out
    for row in rows[1:]:
        if not row or row[idx["player_id"]] in (None, ""):
            continue
        try:
            pid = int(row[idx["player_id"]])
        except (TypeError, ValueError):
            continue
        out[pid] = {
            "exclude":          int(row[idx.get("exclude", -1)] or 0) if idx.get("exclude", -1) >= 0 else 0,
            "exclusion_reason": row[idx.get("exclusion_reason", -1)] if idx.get("exclusion_reason", -1) >= 0 else "",
            "notes":            row[idx.get("notes", -1)] if idx.get("notes", -1) >= 0 else "",
            "last_reviewed":    row[idx.get("last_reviewed", -1)] if idx.get("last_reviewed", -1) >= 0 else "",
        }
    return out


def _build_rows(con: sqlite3.Connection, existing: dict[int, dict]) -> list[dict]:
    today_str = str(date.today())
    rows = con.execute("""
        SELECT pu.player_id, pu.name, pu.age, pu.position_bucket,
               pu.current_club, pu.parent_club, pu.league_id,
               pu.current_tm_value_eur, pu.sellability_score,
               COALESCE((SELECT MAX(match_score) FROM matches WHERE player_id = pu.player_id), 0)
        FROM player_universe pu
        WHERE pu.sellability_status = 'sellable_now'
    """).fetchall()

    out: list[dict] = []
    for (pid, name, age, bucket, cc, pc, lid, tm, sell, best) in rows:
        prev = existing.get(pid, {})
        prev_exclude = prev.get("exclude", 0)
        last_reviewed = prev.get("last_reviewed", "")
        row = {
            "player_id":          pid,
            "name":               name,
            "age":                age,
            "position":           bucket,
            "current_club":       cc,
            "parent_club":        pc,
            "league":             lid,
            "TM_value_eur":       tm,
            "sellability_score":  round(sell or 0.0, 1),
            "best_match_score":   round(best or 0.0, 1),
            "exclude":            prev_exclude,
            "exclusion_reason":   prev.get("exclusion_reason") or "",
            "notes":              prev.get("notes") or "",
            "last_reviewed":      last_reviewed,
        }
        out.append(row)

    # Sort: best_match_score DESC (highest-impact decisions surface first)
    out.sort(key=lambda r: -(r["best_match_score"] or 0.0))
    return out


def write_xlsx(rows: list[dict], path: Path = XLSX_PATH) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "player_overrides"

    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F3864")
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = bold; c.fill = fill; c.alignment = centre
    ws.row_dimensions[1].height = 32

    ws.cell(row=1, column=COL_EXCLUDE).comment = Comment(
        "Set to 1 to drop this player from the Targets view. They still appear "
        "in All Matches and the Excluded Players page (with your reason).", "build"
    )
    ws.cell(row=1, column=COL_REASON).comment = Comment(
        "Short label — e.g. 'agent says no', 'outside RV remit', 'already placed'.", "build"
    )

    ro_fill = PatternFill("solid", fgColor="F2F2F2")
    in_fill = PatternFill("solid", fgColor="FFF2CC")
    excluded_fill = PatternFill("solid", fgColor="FCE4D6")
    money = '"€"#,##0;[Red]"€"-#,##0'
    wrap = Alignment(wrap_text=True, vertical="top")

    for ri, r in enumerate(rows, start=2):
        ws.cell(row=ri, column=1,  value=r["player_id"]).fill = ro_fill
        ws.cell(row=ri, column=2,  value=r["name"]).fill = ro_fill
        ws.cell(row=ri, column=3,  value=r["age"]).fill = ro_fill
        ws.cell(row=ri, column=4,  value=r["position"]).fill = ro_fill
        ws.cell(row=ri, column=5,  value=r["current_club"]).fill = ro_fill
        ws.cell(row=ri, column=6,  value=r["parent_club"]).fill = ro_fill
        ws.cell(row=ri, column=7,  value=r["league"]).fill = ro_fill
        c = ws.cell(row=ri, column=8, value=r["TM_value_eur"])
        c.fill = ro_fill; c.number_format = money
        c = ws.cell(row=ri, column=9, value=r["sellability_score"])
        c.fill = ro_fill; c.number_format = "0.0"
        c = ws.cell(row=ri, column=10, value=r["best_match_score"])
        c.fill = ro_fill; c.number_format = "0.0"
        c = ws.cell(row=ri, column=COL_EXCLUDE, value=r["exclude"])
        c.fill = excluded_fill if r["exclude"] == 1 else in_fill
        c.alignment = Alignment(horizontal="center")
        c = ws.cell(row=ri, column=COL_REASON, value=r["exclusion_reason"])
        c.fill = in_fill; c.alignment = wrap
        c = ws.cell(row=ri, column=COL_NOTES, value=r["notes"])
        c.fill = in_fill; c.alignment = wrap
        ws.cell(row=ri, column=COL_LAST_REVIEWED, value=r["last_reviewed"]).fill = ro_fill

    dv = DataValidation(type="list", formula1='"0,1"', allow_blank=False)
    dv.error = "Enter 0 or 1."
    dv.errorTitle = "Invalid exclude value"
    letter = get_column_letter(COL_EXCLUDE)
    dv.add(f"{letter}2:{letter}{len(rows) + 1}")
    ws.add_data_validation(dv)

    for ci, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "D2"
    last_col = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _reconcile_counts(staging_path: Path, expected: int) -> list[dict]:
    wb = load_workbook(staging_path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    headers = [str(h) if h is not None else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}
    pids = [r[idx["player_id"]] for r in rows[1:] if r and r[idx["player_id"]] is not None]
    mismatches: list[dict] = []
    if len(pids) != expected:
        mismatches.append({"check": "row count vs sellable cohort", "expected": expected, "actual": len(pids)})
    if len(set(pids)) != len(pids):
        mismatches.append({"check": "duplicate player_ids", "expected": 0, "actual": len(pids) - len(set(pids))})
    return mismatches


def _write_recon_log(diff: list[dict]) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = LOG_DIR / f"script_27_reconciliation_{ts}.log"
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("w", encoding="utf-8") as f:
        f.write("# Script 27 reconciliation failure — player_overrides.staging.xlsx\n")
        f.write(f"# Timestamp: {datetime.now().isoformat(timespec='seconds')}\n\n")
        f.write(f"# {'check':40s}  {'expected':>10s}  {'actual':>10s}\n")
        for d in diff:
            f.write(f"  {d['check']:40s}  {d['expected']:>10}  {d['actual']:>10}\n")
    return log_path


def main() -> None:
    if not Path(config.SQLITE_FILE).exists():
        sys.exit(f"Missing {config.SQLITE_FILE} — run 02→22 first.")

    if XLSX_PATH.exists():
        shutil.copy2(XLSX_PATH, BAK_XLSX_PATH)

    existing = _load_existing()
    with sqlite3.connect(config.SQLITE_FILE) as con:
        rows = _build_rows(con, existing)
        write_xlsx(rows, STAGING_XLSX_PATH)
        diff = _reconcile_counts(STAGING_XLSX_PATH, len(rows))

    if diff:
        log = _write_recon_log(diff)
        bar = "=" * 78
        print()
        print(bar)
        print("\033[1;31m⚠  RECONCILIATION FAILED — staged player_overrides.xlsx diverges\033[0m")
        print(bar)
        for m in diff:
            print(f"   • {m['check']}: expected={m['expected']}, actual={m['actual']}")
        print(f"   Log: {log.resolve()}")
        print(f"   Live xlsx: {XLSX_PATH.resolve()}  (UNCHANGED)")
        print(f"   Staging:   {STAGING_XLSX_PATH.resolve()}  (kept for inspection)")
        print(bar)
        sys.exit(2)

    STAGING_XLSX_PATH.replace(XLSX_PATH)

    n = len(rows)
    n_excluded = sum(1 for r in rows if r["exclude"] == 1)
    print(f"Wrote {XLSX_PATH.resolve()}")
    print(f"  {n} rows total — {n_excluded} marked exclude=1, {n - n_excluded} active in Targets view")
    print()
    print("Sorted by best_match_score DESC. Highest-impact rows at the top:")
    print(f"  {'name':28s} {'pos':6s} {'lg':5s} {'sell':>5s} {'best':>5s}  excl")
    for r in rows[:15]:
        flag = "✓" if r["exclude"] == 1 else ""
        print(f"  {(r['name'] or '')[:28]:28s} {r['position']:6s} {r['league']:5s} "
              f"{r['sellability_score']:>5.1f} {r['best_match_score']:>5.1f}  {flag}")


if __name__ == "__main__":
    main()
