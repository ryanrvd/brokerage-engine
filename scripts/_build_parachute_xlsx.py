"""One-shot helper: build data/parachute_payments.xlsx with pre-populated
yr1/yr2/yr3 cohorts.

yr1: from data/manual_league_overrides.csv (PL → Championship transitions for 26/27)
yr2: derived from dcaribou games — clubs in GB1 24/25 but not 25/26
yr3: derived from dcaribou games — clubs in GB1 23/24 but not 24/25

Run once: `.venv/bin/python scripts/_build_parachute_xlsx.py`
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

OUT = Path("data/parachute_payments.xlsx")

HEADERS = [
    "club_id",
    "club_name",
    "parachute_year",
    "season_relegated",
    "needs_review",
    "notes",
]
WIDTHS = [10, 42, 16, 18, 14, 100]

# Pre-populated cohort. needs_review=1 highlights rows where I'm not 100% confident.
ROWS = [
    # yr1 — from manual_league_overrides.csv (relegated end of 25/26 → parachute yr1 in 26/27)
    {"club_id": "543",  "club_name": "Wolverhampton Wanderers Football Club",
     "parachute_year": 1, "season_relegated": "25/26", "needs_review": 0,
     "notes": "seeded from manual_league_overrides.csv (newly relegated for 26/27)"},
    {"club_id": "1132", "club_name": "Burnley Football Club",
     "parachute_year": 1, "season_relegated": "25/26", "needs_review": 0,
     "notes": "seeded from manual_league_overrides.csv (newly relegated for 26/27); previous 23/24 relegation window expired after 24/25 Championship-winning promotion"},

    # yr2 — derived from dcaribou: GB1 in 24/25 but not in 25/26
    {"club_id": "1003", "club_name": "Leicester City",
     "parachute_year": 2, "season_relegated": "24/25", "needs_review": 0,
     "notes": "derived from dcaribou games (in GB1 24/25, not GB1 25/26)"},
    {"club_id": "180",  "club_name": "Southampton FC",
     "parachute_year": 2, "season_relegated": "24/25", "needs_review": 0,
     "notes": "derived from dcaribou games (in GB1 24/25, not GB1 25/26)"},

    # yr3 — derived from dcaribou: GB1 in 23/24 but not in 24/25
    {"club_id": "350",  "club_name": "Sheffield United",
     "parachute_year": 3, "season_relegated": "23/24", "needs_review": 0,
     "notes": "derived from dcaribou games (in GB1 23/24, not GB1 24/25)"},
    {"club_id": "1031", "club_name": "Luton Town",
     "parachute_year": 3, "season_relegated": "23/24", "needs_review": 1,
     "notes": "derived from dcaribou games (in GB1 23/24, not GB1 24/25); NOT IN club_pressure — likely relegated to League One (GB3, outside our 19-league coverage). Flag will have no downstream effect. Keep row for transparency, or delete if you want a clean registry."},
]
# Excluded from list (worth noting for the user):
#   Ipswich Town (677) — would be yr2 per dcaribou (relegated 24/25) but
#   manual_league_overrides.csv shows them promoted back to GB1 for 26/27,
#   so they're out of the parachute window while in PL.

PARACHUTE_NOTE = (
    "Auto-pre-populated. yr1 seeded from manual_league_overrides.csv; "
    "yr2/yr3 derived from dcaribou games (Premier League → Championship transitions). "
    "Excluded: Ipswich Town (677) — relegated end of 24/25 but promoted back to PL "
    "for 26/27 → out of parachute window. Edit needs_review=1 rows; delete any row "
    "you disagree with; pipeline re-reads on next run."
)


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "parachute_payments"

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3864")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    ws.cell(row=1, column=1).comment = Comment(PARACHUTE_NOTE, "build")

    review_fill = PatternFill("solid", fgColor="FFF2CC")  # pale yellow — needs_review=1
    edit_fill   = PatternFill("solid", fgColor="FFFFFF")  # white — editable

    wrap = Alignment(wrap_text=True, vertical="top")
    for ri, r in enumerate(ROWS, start=2):
        needs = r["needs_review"] == 1
        for ci, key in enumerate(["club_id", "club_name", "parachute_year",
                                  "season_relegated", "needs_review", "notes"], start=1):
            cell = ws.cell(row=ri, column=ci, value=r[key])
            cell.fill = review_fill if needs else edit_fill
            if key == "notes":
                cell.alignment = wrap

    # Data validation on parachute_year (1/2/3) and needs_review (0/1)
    dv_year = DataValidation(type="list", formula1='"1,2,3"', allow_blank=False)
    dv_year.error = "Enter 1, 2, or 3."
    dv_year.errorTitle = "Invalid parachute_year"
    dv_year.add(f"C2:C{len(ROWS) + 1}")
    ws.add_data_validation(dv_year)

    dv_review = DataValidation(type="list", formula1='"0,1"', allow_blank=False)
    dv_review.error = "Enter 0 or 1."
    dv_review.errorTitle = "Invalid needs_review"
    dv_review.add(f"E2:E{len(ROWS) + 1}")
    ws.add_data_validation(dv_review)

    for ci, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(ROWS) + 1}"
    ws.row_dimensions[1].height = 36

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT.resolve()} — {len(ROWS)} row(s)")
    n_review = sum(1 for r in ROWS if r["needs_review"] == 1)
    print(f"  {n_review} row(s) flagged needs_review=1 (highlighted pale yellow)")


if __name__ == "__main__":
    main()
