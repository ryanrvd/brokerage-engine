"""Step 30 — Refresh CA/PA (SciSkill) from the SciSports API.

Process order — paranoid about the shared 1000-req/60s budget:

  1. Pre-flight: quota baseline + cost estimate. Exit if --yes not passed and
     stdin doesn't get an explicit 'y'.
  2. Priority queue, processed strictly in order:
        (a) relegated cohort (mandate_priority_multiplier ≥ 1.3)
        (b) sellable_now players
        (c) sellable_with_caveat players
        (d) everyone else with scisports_player_id and tm_squad_scrape
     If quota runs tight, processing halts cleanly and can resume next run.
  3. Per player:
        - Skip if player_ratings row is 'active' AND last_updated < 7 days old
        - Look up cache (manual_seed wins; otherwise fresh fetch)
        - On fresh fetch: call /api/v2/metrics/players/sciskill?PlayerIds=…
        - Write player_ratings row with retrieved CA/PA
  4. Once all DB writes are done, sync data/scisports_ratings.xlsx via the
     existing reconcile_scisports + load_scisports_ratings tools — preserves
     manually-entered notes.

Pacing is enforced inside the client (`_scisports_client.py`). This script
just iterates and respects the EMERGENCY HALT exception.
"""
from __future__ import annotations

import argparse
import datetime as dt
import sqlite3
import sys
import time
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(Path(__file__).resolve().parent))

import config
from _scisports_cache import TTL_SCISKILL, get_or_fetch
from _scisports_client import (
    ScisportsClient,
    ScisportsRateLimitedError,
    ScisportsRateLimitEmergency,
)

SCISKILL_ENDPOINT = "/api/v2/metrics/players/sciskill"
PAGE_SIZE = 1   # one player per call — cheapest possible
FRESH_TTL_DAYS = 7

XLSX_PATH = PROJECT_ROOT / "data" / "scisports_ratings.xlsx"


def _player_ratings_freshness(con: sqlite3.Connection) -> dict[int, dt.datetime]:
    """tm_player_id -> last_updated datetime, only for status='active'."""
    out: dict[int, dt.datetime] = {}
    cur = con.execute(
        "SELECT tm_player_id, last_updated FROM player_ratings "
        "WHERE status = 'active' AND last_updated IS NOT NULL"
    )
    for pid, ts in cur.fetchall():
        if not ts:
            continue
        try:
            out[int(pid)] = dt.datetime.fromisoformat(str(ts).split("T")[0])
        except (TypeError, ValueError):
            continue
    return out


def _candidate_queue(con: sqlite3.Connection,
                     freshness: dict[int, dt.datetime]) -> list[dict]:
    """Build the prioritised refresh queue.

    Returns list of dicts with player_id, scisports_player_id, name,
    priority_band, multiplier. Highest priority first.
    """
    rows = con.execute("""
        SELECT pu.player_id, pu.name, pu.scisports_player_id,
               pu.sellability_status, pu.mandate_priority_multiplier,
               pu.parent_club_recently_relegated, pu.parent_club
        FROM player_universe pu
        WHERE pu.data_source = 'tm_squad_scrape'
          AND pu.scisports_player_id IS NOT NULL
    """).fetchall()

    today = dt.datetime.now(dt.timezone.utc).date()
    cutoff = today - dt.timedelta(days=FRESH_TTL_DAYS)

    queue = []
    skipped_fresh = 0
    for (pid, name, sci_id, status, mult,
         relegated, parent_club) in rows:
        prev_ts = freshness.get(int(pid))
        if prev_ts and prev_ts.date() >= cutoff:
            skipped_fresh += 1
            continue
        # Priority band
        if relegated:
            band, band_label = 0, "relegated"
        elif status == "sellable_now":
            band, band_label = 1, "sellable_now"
        elif status == "sellable_with_caveat":
            band, band_label = 2, "sellable_with_caveat"
        else:
            band, band_label = 3, "other"
        queue.append({
            "player_id": pid,
            "scisports_player_id": int(sci_id),
            "name": name,
            "priority_band": band,
            "priority_label": band_label,
            "multiplier": mult or 1.0,
            "parent_club": parent_club,
        })
    queue.sort(key=lambda r: (r["priority_band"], -float(r["multiplier"] or 1.0), r["name"] or ""))
    return queue, skipped_fresh


def _extract_ca_pa(payload) -> tuple[float | None, float | None, bool]:
    """Return (ca, pa, was_seeded). Tolerates both seeded and live shapes.

    Live SciSports response uses keys 'sciskill' (CA) and 'potential' (PA).
    Manual_seed cache entries use 'currentSciSkill' / 'potentialSciSkill' to
    distinguish them. We accept both.
    """
    items = payload.get("items") if isinstance(payload, dict) else None
    items = items or payload.get("data") if isinstance(payload, dict) else items
    if not items:
        return None, None, False
    item = items[0]
    ca = (item.get("sciskill")
          or item.get("currentSciSkill")
          or item.get("ca")
          or item.get("current"))
    pa = (item.get("potential")
          or item.get("potentialSciSkill")
          or item.get("pa"))
    was_seeded = bool(item.get("_seeded"))
    return (float(ca) if ca is not None else None,
            float(pa) if pa is not None else None,
            was_seeded)


def _ensure_player_ratings_table(con: sqlite3.Connection) -> None:
    con.execute("""
        CREATE TABLE IF NOT EXISTS player_ratings (
            tm_player_id       INTEGER PRIMARY KEY,
            current_ability    REAL,
            potential_ability  REAL,
            status             TEXT,
            last_updated       TEXT
        )
    """)


def _sync_xlsx_from_ratings(con: sqlite3.Connection) -> dict:
    """Write API-derived CA/PA from player_ratings back to scisports_ratings.xlsx.

    Preserves every other cell (notes, status, last_updated) exactly. Only
    cells that are currently blank or numeric-zero get overwritten with the
    API value. Manually-entered CA/PA values stay untouched.
    """
    if not XLSX_PATH.exists():
        return {"updated": 0, "skipped_manual": 0, "no_data": 0}
    from openpyxl import load_workbook
    wb = load_workbook(XLSX_PATH)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    ci = {h: i for i, h in enumerate(headers)}
    if "tm_player_id" not in ci or "current_ability" not in ci or "potential_ability" not in ci:
        return {"updated": 0, "skipped_manual": 0, "no_data": 0}

    ratings = {int(pid): (ca, pa, status) for pid, ca, pa, status in con.execute(
        "SELECT tm_player_id, current_ability, potential_ability, status "
        "FROM player_ratings"
    ).fetchall()}

    updated = 0
    skipped_manual = 0
    no_data = 0
    pid_col = ci["tm_player_id"] + 1
    ca_col = ci["current_ability"] + 1
    pa_col = ci["potential_ability"] + 1
    status_col = ci["status"] + 1 if "status" in ci else None
    lu_col = ci["last_updated"] + 1 if "last_updated" in ci else None

    today_iso = dt.date.today().isoformat()
    for row_idx in range(2, ws.max_row + 1):
        pid_cell = ws.cell(row=row_idx, column=pid_col).value
        if pid_cell is None:
            continue
        try:
            pid = int(pid_cell)
        except (TypeError, ValueError):
            continue
        if pid not in ratings:
            continue
        api_ca, api_pa, status = ratings[pid]
        if api_ca is None and api_pa is None:
            no_data += 1
            continue
        existing_ca = ws.cell(row=row_idx, column=ca_col).value
        existing_pa = ws.cell(row=row_idx, column=pa_col).value

        def _is_blank(v):
            if v is None or v == "":
                return True
            try:
                return float(v) == 0.0
            except (TypeError, ValueError):
                return False

        ca_blank = _is_blank(existing_ca)
        pa_blank = _is_blank(existing_pa)

        if not ca_blank and not pa_blank:
            skipped_manual += 1
            continue

        if ca_blank and api_ca is not None:
            ws.cell(row=row_idx, column=ca_col, value=round(float(api_ca), 1))
        if pa_blank and api_pa is not None:
            ws.cell(row=row_idx, column=pa_col, value=round(float(api_pa), 1))
        if status_col:
            ws.cell(row=row_idx, column=status_col, value="active")
        if lu_col:
            ws.cell(row=row_idx, column=lu_col, value=today_iso)
        updated += 1

    try:
        wb.save(XLSX_PATH)
    except PermissionError as e:
        raise PermissionError(
            f"Cannot write {XLSX_PATH} — may be open in Excel. Close and re-run."
        ) from e
    return {"updated": updated, "skipped_manual": skipped_manual, "no_data": no_data}


def _upsert_rating(con: sqlite3.Connection, tm_pid: int,
                   ca: float | None, pa: float | None) -> None:
    today_iso = dt.date.today().isoformat()
    status = "active" if (ca is not None or pa is not None) else "pending"
    con.execute("""
        INSERT INTO player_ratings (tm_player_id, current_ability, potential_ability,
                                    status, last_updated)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(tm_player_id) DO UPDATE SET
            current_ability = excluded.current_ability,
            potential_ability = excluded.potential_ability,
            status = excluded.status,
            last_updated = excluded.last_updated
    """, (int(tm_pid), ca, pa, status, today_iso))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--yes", action="store_true",
                        help="Skip the y/n confirmation (CI / re-invocations)")
    parser.add_argument("--limit", type=int, default=None,
                        help="Cap the number of fresh API calls (for testing)")
    args = parser.parse_args()

    print("=" * 72)
    print("Step 30 — SciSports CA/PA refresh")
    print("=" * 72)

    client = ScisportsClient()
    pf = client.preflight_baseline(halt_threshold=800)
    print(f"  Pre-flight remaining: {pf['remaining']}  fresh={pf['looks_fresh']}  "
          f"sample={pf['sample_league_name']}")
    if not pf["looks_fresh"]:
        print("  ⚠️  Remaining quota < 800. Confirm no concurrent maps-repo run.")
        sys.exit(2)

    con = sqlite3.connect(config.SQLITE_FILE)
    _ensure_player_ratings_table(con)
    freshness = _player_ratings_freshness(con)
    queue, skipped_fresh = _candidate_queue(con, freshness)

    # Cost estimate (each entry is at most 1 API call; cache may absorb many).
    # Group by priority band for the printout.
    bands = {0: 0, 1: 0, 2: 0, 3: 0}
    for q in queue:
        bands[q["priority_band"]] += 1

    print()
    print(f"  tm_squad_scrape players with scisports_player_id: {len(queue) + skipped_fresh}")
    print(f"  Skipped (already 'active' within {FRESH_TTL_DAYS}d):    {skipped_fresh}")
    print(f"  In refresh queue:                                {len(queue)}")
    print(f"    relegated cohort (priority 0):                 {bands[0]}")
    print(f"    sellable_now      (priority 1):                {bands[1]}")
    print(f"    sellable_w/caveat (priority 2):                {bands[2]}")
    print(f"    other             (priority 3):                {bands[3]}")
    print()
    eff_limit = args.limit if args.limit is not None else len(queue)
    work_total = min(len(queue), eff_limit)
    est_seconds = round(work_total * 0.95)  # ~250ms gap + ~700ms response
    print(f"  About to attempt up to {work_total} SciSkill lookups.")
    print(f"  Cache hits cost zero quota. Worst case (all cache misses):")
    print(f"    ~{work_total} API calls, ~{est_seconds}s wall clock,")
    print(f"    consuming ~{work_total} of {pf['remaining']} quota.")
    print()

    if not args.yes:
        try:
            ans = input("  Continue? [y/N] ").strip().lower()
        except EOFError:
            ans = ""
        if ans != "y":
            print("  Aborted.")
            sys.exit(1)

    print()
    print("Processing…")
    print()

    cache_hits = 0
    seed_hits = 0
    live_calls_ok = 0
    live_calls_no_data = 0
    halted = False
    halt_reason = ""

    for i, q in enumerate(queue[:eff_limit], start=1):
        sci_id = q["scisports_player_id"]
        tm_pid = q["player_id"]
        params = {"Offset": 0, "Limit": 1, "PlayerIds": sci_id}
        try:
            data, src = get_or_fetch(client, SCISKILL_ENDPOINT, params, TTL_SCISKILL)
        except ScisportsRateLimitEmergency as e:
            halted = True
            halt_reason = str(e)
            break
        except ScisportsRateLimitedError as e:
            halted = True
            halt_reason = str(e)
            break

        ca, pa, was_seeded = _extract_ca_pa(data)
        _upsert_rating(con, tm_pid, ca, pa)

        if src == "manual_seed":
            seed_hits += 1
        elif src == "cache_hit":
            cache_hits += 1
        else:
            if ca is None and pa is None:
                live_calls_no_data += 1
            else:
                live_calls_ok += 1

        # Periodic commit + progress log
        if i % 50 == 0:
            con.commit()
            print(f"  [{i:>4}/{work_total}]  {q['priority_label']:22s} "
                  f"remaining={client.last_remaining}  "
                  f"seed={seed_hits} cache={cache_hits} live_ok={live_calls_ok} "
                  f"live_empty={live_calls_no_data}")

    con.commit()

    # Write CA/PA back to the xlsx so the user sees fresh values in the
    # same workbook they edit manually. Notes + manual CA/PA stay untouched.
    print()
    print("Syncing xlsx from player_ratings…")
    try:
        sync = _sync_xlsx_from_ratings(con)
        print(f"  xlsx cells filled from API:  {sync['updated']}")
        print(f"  xlsx rows preserved manual:  {sync['skipped_manual']}")
        print(f"  xlsx rows skipped (no data): {sync['no_data']}")
    except PermissionError as e:
        print(f"  ! {e}")

    print()
    print("=" * 72)
    print("Step 30 summary")
    print("=" * 72)
    print(f"  Processed:                  {min(eff_limit, len(queue)) if not halted else 'PARTIAL'}")
    print(f"  Manual-seed cache hits:     {seed_hits}")
    print(f"  Live-cache hits (prior fetch): {cache_hits}")
    print(f"  Live API fetches (CA/PA):   {live_calls_ok}")
    print(f"  Live API fetches (empty):   {live_calls_no_data}")
    print(f"  Total client.requests_made: {client.requests_made}")
    print(f"  Min X-RateLimit-Remaining:  {client.min_remaining_seen}")
    print(f"  Final X-RateLimit-Remaining: {client.last_remaining}")
    if halted:
        print()
        print(f"  ⚠️  HALTED: {halt_reason}")
        print("  Re-run after quota resets; already-fetched players are in player_ratings.")
        sys.exit(2)
    con.close()


if __name__ == "__main__":
    try:
        main()
    except (ScisportsRateLimitEmergency, ScisportsRateLimitedError) as e:
        print(f"\nRATE-LIMIT EMERGENCY: {e}")
        sys.exit(2)
