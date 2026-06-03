"""Phase 4 — Refresh data/scisports_ratings.xlsx as a downstream snapshot.

The xlsx is now an OUTPUT (a read-only-ish snapshot for human reference),
NOT a source of truth for seeding. SciSports API populates player_ratings
in SQLite; this script then renders that table to the xlsx.

Notes column is preserved from the prior xlsx where rows still exist.
Otherwise the xlsx is rebuilt from scratch each run.

A header row note clarifies the new contract.
"""
from __future__ import annotations
import datetime as dt
import sqlite3
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
import config

FILE_PATH = PROJECT_ROOT / "data" / "scisports_ratings.xlsx"

COLUMNS = [
    "tm_player_id", "player_name", "parent_club", "current_club",
    "position_bucket", "age", "is_on_loan",
    "current_ability", "potential_ability",
    "status", "source", "last_updated", "notes",
]
COLUMN_WIDTHS = {
    "tm_player_id": 14, "player_name": 28, "parent_club": 30,
    "current_club": 30, "position_bucket": 10, "age": 6,
    "is_on_loan": 12, "current_ability": 14, "potential_ability": 14,
    "status": 12, "source": 14, "last_updated": 14, "notes": 42,
}

NOTE_HEADER = (
    "Output snapshot — source of truth is player_ratings table populated "
    "from SciSports API. Edits here are NOT read back into the matcher."
)

STATUS_ORDER = {"pending": 0, "active": 1, "departed": 2, "killed": 3, "invalid": 4}


def read_notes_from_existing(file_path: Path) -> dict[int, str]:
    """Preserve manually-written notes from the prior xlsx."""
    if not file_path.exists():
        return {}
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {}
        # The prior xlsx had the snapshot note in row 1 if present (or maybe not).
        # Robust approach: find the header row by looking for "tm_player_id".
        header_row_idx = None
        for i, r in enumerate(rows):
            if r and "tm_player_id" in [str(c).strip().lower() if c else "" for c in r]:
                header_row_idx = i
                break
        if header_row_idx is None:
            return {}
        headers = [str(c).strip() if c else "" for c in rows[header_row_idx]]
        try:
            pid_idx = headers.index("tm_player_id")
            notes_idx = headers.index("notes")
        except ValueError:
            return {}
        out: dict[int, str] = {}
        for r in rows[header_row_idx + 1:]:
            if not r or r[pid_idx] is None:
                continue
            try:
                pid = int(r[pid_idx])
            except (TypeError, ValueError):
                continue
            note = r[notes_idx] if notes_idx < len(r) else None
            if note not in (None, ""):
                out[pid] = str(note)
        return out
    except Exception:
        return {}


def main() -> None:
    today = dt.date.today().isoformat()
    prior_notes = read_notes_from_existing(FILE_PATH)
    print(f"Preserved notes from prior xlsx: {len(prior_notes)}")

    con = sqlite3.connect(config.SQLITE_FILE)
    # Join player_ratings with player_universe context
    rows = con.execute("""
        SELECT pr.tm_player_id, pu.name, pu.parent_club, pu.current_club,
               pu.position_bucket, pu.age, pu.on_loan,
               pr.current_ability, pr.potential_ability,
               pr.status, pr.source, pr.last_updated
        FROM player_ratings pr
        LEFT JOIN player_universe pu ON pu.player_id = pr.tm_player_id
        ORDER BY pr.status, pu.name
    """).fetchall()
    con.close()
    print(f"player_ratings rows: {len(rows)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "scisports_ratings"

    # Snapshot note in row 1 (spans the table width)
    ws.cell(row=1, column=1, value=NOTE_HEADER).font = Font(italic=True, color="6B6B6B")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    ws.row_dimensions[1].height = 24

    # Header row in row 2
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3864")
    for col_idx, name in enumerate(COLUMNS, start=1):
        c = ws.cell(row=2, column=col_idx, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center")

    # Sort rows so pending → active → departed → killed → invalid for the worklist
    rows_sorted = sorted(
        rows,
        key=lambda r: (STATUS_ORDER.get((r[9] or "").lower(), 9),
                       (r[1] or "").lower()),
    )

    for ri, row in enumerate(rows_sorted, start=3):
        (pid, name, parent_club, current_club, position_bucket, age,
         on_loan, ca, pa, status, source, last_updated) = row
        values = {
            "tm_player_id": pid,
            "player_name": name or "",
            "parent_club": parent_club or "",
            "current_club": current_club or "",
            "position_bucket": position_bucket or "",
            "age": age,
            "is_on_loan": "TRUE" if on_loan else "FALSE",
            "current_ability": ca,
            "potential_ability": pa,
            "status": status,
            "source": source,
            "last_updated": last_updated or today,
            "notes": prior_notes.get(int(pid), ""),
        }
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            ws.cell(row=ri, column=col_idx, value=values[col_name])

    # Column widths + freeze headers
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(col_name, 16)
    ws.freeze_panes = "A3"

    try:
        wb.save(FILE_PATH)
    except PermissionError:
        sys.exit(f"Cannot write {FILE_PATH} — may be open in Excel.")

    print(f"Wrote {FILE_PATH}: {len(rows)} rows, notes preserved on "
          f"{sum(1 for r in rows_sorted if prior_notes.get(int(r[0])))} rows")


if __name__ == "__main__":
    main()
