"""
Step 26 — Generate / refresh `data/manual_wages.xlsx`.

A user-input checklist for wage data per sellable player. One row per player in
the Sellable Asset Ledger (Sheet 4 filter). Ryan fills in `wage_pw_eur` for the
players he wants real wage-feasibility signal on; the match engine (script 22)
reads back from this file and uses 1.0 / 0.0 / 0.7-fallback semantics.

Same pattern as `data/manual_flags.xlsx`:
  • Idempotent — preserves prior user inputs across runs.
  • Sorted: mapped-league players first (buyer_demand_mapped=YES), then by
    sellability_score DESC inside each group. Highest-priority entries surface
    at the top of the sheet so partial coverage stays useful.
  • Staging-file + count-reconciliation safety net.
  • Rolling `.bak` at start of every run.

Read-only columns (grey): player_id, name, age, position, current_club,
  parent_club, league, TM_value, sellability_score, contract_end, buyer_demand_mapped
User-input columns (yellow): wage_pw_eur, wage_source, notes
Auto-stamped: last_reviewed
"""

import shutil
import sqlite3
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import config

XLSX_PATH = Path("data/manual_wages.xlsx")
BAK_XLSX_PATH = Path("data/manual_wages.xlsx.bak")
STAGING_XLSX_PATH = Path("data/manual_wages.staging.xlsx")
LOG_DIR = Path("logs")

# Leagues with buyer-side demand mapping — sellers in these leagues are highest
# priority because their wages directly affect wage_feasibility for matches
# involving demand-side data.
MAPPED_LEAGUES = {"GB1", "GB2", "ES1", "IT1", "L1", "FR1", "FR2", "PO1", "NL1", "BE1"}

HEADERS = [
    "player_id",                # 1  RO
    "name",                     # 2  RO
    "age",                      # 3  RO
    "position",                 # 4  RO
    "current_club",             # 5  RO
    "parent_club",              # 6  RO
    "league",                   # 7  RO
    "TM_value_eur",             # 8  RO
    "sellability_score",        # 9  RO
    "contract_end",             # 10 RO
    "buyer_demand_mapped",      # 11 RO  yes/no
    "wage_pw_eur",              # 12 user input
    "wage_source",              # 13 user input
    "notes",                    # 14 user input
    "last_reviewed",            # 15 auto-stamped
]
WIDTHS = [10, 28, 5, 8, 28, 28, 7, 14, 12, 14, 11, 14, 18, 36, 14]

COL_WAGE_PW    = 12
COL_WAGE_SRC   = 13
COL_NOTES      = 14
COL_LAST_REV   = 15


def _load_existing() -> dict[int, dict]:
    """Re-read existing xlsx to preserve user entries."""
    if not XLSX_PATH.exists():
        return {}
    out: dict[int, dict] = {}
    wb = load_workbook(XLSX_PATH, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return out
    headers = [str(h) if h is not None else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}
    if "player_id" not in idx:
        return out
    for row in rows[1:]:
        if not row or row[idx["player_id"]] in (None, ""):
            continue
        try:
            pid = int(row[idx["player_id"]])
        except (TypeError, ValueError):
            continue
        out[pid] = {
            "wage_pw_eur":   row[idx.get("wage_pw_eur", -1)] if idx.get("wage_pw_eur", -1) >= 0 else None,
            "wage_source":   row[idx.get("wage_source", -1)] if idx.get("wage_source", -1) >= 0 else None,
            "notes":         row[idx.get("notes", -1)] if idx.get("notes", -1) >= 0 else None,
            "last_reviewed": row[idx.get("last_reviewed", -1)] if idx.get("last_reviewed", -1) >= 0 else None,
        }
    return out


def _build_rows(con: sqlite3.Connection, existing: dict[int, dict]) -> list[dict]:
    today_str = str(date.today())
    rows = con.execute("""
        SELECT pu.player_id, pu.name, pu.age, pu.position_bucket, pu.current_club,
               pu.parent_club, pu.league_id, pu.current_tm_value_eur,
               pu.sellability_score, pu.contract_end_date
        FROM player_universe pu
        WHERE (pu.right_priced=1 OR pu.finished_product=1
               OR pu.finished_product IS NULL OR pu.contract_leveraged=1)
    """).fetchall()

    out: list[dict] = []
    for r in rows:
        (pid, name, age, bucket, cc, pc, lid, tm, sell, ce) = r
        prev = existing.get(pid, {})
        new_row = {
            "player_id":            pid,
            "name":                 name,
            "age":                  age,
            "position":             bucket,
            "current_club":         cc,
            "parent_club":          pc,
            "league":               lid,
            "TM_value_eur":         tm,
            "sellability_score":    round(sell or 0.0, 1),
            "contract_end":         ce,
            "buyer_demand_mapped":  "yes" if lid in MAPPED_LEAGUES else "no",
            "wage_pw_eur":          prev.get("wage_pw_eur"),
            "wage_source":          prev.get("wage_source") or "",
            "notes":                prev.get("notes") or "",
            "last_reviewed":        prev.get("last_reviewed") or "",
        }
        # Auto-stamp last_reviewed when wage value changes
        if (prev.get("wage_pw_eur") != new_row["wage_pw_eur"]) and new_row["wage_pw_eur"]:
            new_row["last_reviewed"] = today_str
        out.append(new_row)

    # Sort: mapped-league=yes first, then sellability_score DESC
    out.sort(key=lambda r: (
        0 if r["buyer_demand_mapped"] == "yes" else 1,
        -(r["sellability_score"] or 0.0),
    ))
    return out


def write_xlsx(rows: list[dict], path: Path = XLSX_PATH) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "manual_wages"

    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F3864")
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = bold; c.fill = fill; c.alignment = centre
    ws.row_dimensions[1].height = 32

    # Header comments
    ws.cell(row=1, column=COL_WAGE_PW).comment = Comment(
        "Enter weekly wage in EUR. Leave blank to keep the 0.7 wage_feasibility "
        "fallback. The match engine reads this column on every run.", "build"
    )
    ws.cell(row=1, column=COL_WAGE_SRC).comment = Comment(
        "Free text — where you sourced this (capology, manual, agent, etc.).", "build"
    )

    ro_fill = PatternFill("solid", fgColor="F2F2F2")
    in_fill = PatternFill("solid", fgColor="FFF2CC")
    mapped_yes_fill = PatternFill("solid", fgColor="E2EFDA")  # green for mapped
    mapped_no_fill  = PatternFill("solid", fgColor="FCE4D6")  # orange for unmapped
    money = '"€"#,##0;[Red]"€"-#,##0'
    wrap = Alignment(wrap_text=True, vertical="top")

    for ri, r in enumerate(rows, start=2):
        ws.cell(row=ri, column=1,  value=r["player_id"]).fill = ro_fill
        ws.cell(row=ri, column=2,  value=r["name"]).fill = ro_fill
        ws.cell(row=ri, column=3,  value=r["age"]).fill = ro_fill
        ws.cell(row=ri, column=4,  value=r["position"]).fill = ro_fill
        ws.cell(row=ri, column=5,  value=r["current_club"]).fill = ro_fill
        ws.cell(row=ri, column=6,  value=r["parent_club"]).fill = ro_fill
        ws.cell(row=ri, column=7,  value=r["league"]).fill = ro_fill
        c = ws.cell(row=ri, column=8, value=r["TM_value_eur"])
        c.fill = ro_fill; c.number_format = money
        c = ws.cell(row=ri, column=9, value=r["sellability_score"])
        c.fill = ro_fill; c.number_format = "0.0"
        ws.cell(row=ri, column=10, value=r["contract_end"]).fill = ro_fill
        c = ws.cell(row=ri, column=11, value=r["buyer_demand_mapped"])
        c.fill = mapped_yes_fill if r["buyer_demand_mapped"] == "yes" else mapped_no_fill
        c.alignment = Alignment(horizontal="center")
        c = ws.cell(row=ri, column=COL_WAGE_PW, value=r["wage_pw_eur"])
        c.fill = in_fill; c.number_format = money
        ws.cell(row=ri, column=COL_WAGE_SRC, value=r["wage_source"]).fill = in_fill
        c = ws.cell(row=ri, column=COL_NOTES, value=r["notes"])
        c.fill = in_fill; c.alignment = wrap
        ws.cell(row=ri, column=COL_LAST_REV, value=r["last_reviewed"]).fill = ro_fill

    # Data validation: wage_pw_eur must be a positive integer (or blank)
    dv = DataValidation(type="whole", operator="greaterThan", formula1=0, allow_blank=True)
    dv.error = "Enter a positive whole number (weekly wage in EUR) or leave blank."
    dv.errorTitle = "Invalid wage"
    letter = get_column_letter(COL_WAGE_PW)
    dv.add(f"{letter}2:{letter}{len(rows) + 1}")
    ws.add_data_validation(dv)

    for ci, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "D2"  # freeze through column C (player_id, name, age)
    last_col = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _reconcile_counts(staging_path: Path, expected_player_count: int) -> list[dict]:
    """Lightweight count check: same player count as input, every player has a
    valid row, no duplicates."""
    wb = load_workbook(staging_path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [{"check": "any rows", "expected": expected_player_count, "actual": 0}]
    headers = [str(h) if h is not None else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}
    pids = [r[idx["player_id"]] for r in rows[1:] if r and r[idx["player_id"]] is not None]
    mismatches: list[dict] = []
    if len(pids) != expected_player_count:
        mismatches.append({
            "check": "row count vs sellable cohort",
            "expected": expected_player_count, "actual": len(pids),
        })
    if len(set(pids)) != len(pids):
        mismatches.append({
            "check": "duplicate player_ids",
            "expected": 0, "actual": len(pids) - len(set(pids)),
        })
    return mismatches


def _write_recon_log(diff: list[dict]) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = LOG_DIR / f"script_26_reconciliation_{ts}.log"
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("w", encoding="utf-8") as f:
        f.write("# Script 26 reconciliation failure — manual_wages.staging.xlsx\n")
        f.write(f"# Timestamp: {datetime.now().isoformat(timespec='seconds')}\n")
        f.write(f"# Existing {XLSX_PATH} was NOT overwritten.\n")
        f.write(f"# Rolling backup at {BAK_XLSX_PATH}; staging kept at {STAGING_XLSX_PATH}.\n\n")
        f.write(f"# {'check':40s}  {'expected':>10s}  {'actual':>10s}\n")
        for d in diff:
            f.write(f"  {d['check']:40s}  {d['expected']:>10}  {d['actual']:>10}\n")
    return log_path


def main() -> None:
    if not Path(config.SQLITE_FILE).exists():
        sys.exit(f"Missing {config.SQLITE_FILE} — run 02→09 first.")

    # Rolling .bak at start
    if XLSX_PATH.exists():
        shutil.copy2(XLSX_PATH, BAK_XLSX_PATH)

    existing = _load_existing()
    with sqlite3.connect(config.SQLITE_FILE) as con:
        rows = _build_rows(con, existing)
        write_xlsx(rows, STAGING_XLSX_PATH)
        mismatches = _reconcile_counts(STAGING_XLSX_PATH, len(rows))

    if mismatches:
        log = _write_recon_log(mismatches)
        bar = "=" * 78
        print()
        print(bar)
        print("\033[1;31m⚠  RECONCILIATION FAILED — staged manual_wages.xlsx diverges\033[0m")
        print(bar)
        for m in mismatches:
            print(f"   • {m['check']}: expected={m['expected']}, actual={m['actual']}")
        print(f"   Log: {log.resolve()}")
        print(f"   Live xlsx: {XLSX_PATH.resolve()}  (UNCHANGED)")
        print(f"   Staging:   {STAGING_XLSX_PATH.resolve()}  (kept for inspection)")
        print("   Inspect the log file, then re-run.")
        print(bar)
        sys.exit(2)

    STAGING_XLSX_PATH.replace(XLSX_PATH)

    # Summary
    n = len(rows)
    n_mapped = sum(1 for r in rows if r["buyer_demand_mapped"] == "yes")
    n_filled = sum(1 for r in rows if r["wage_pw_eur"])
    n_unmapped = n - n_mapped
    print(f"Wrote {XLSX_PATH.resolve()}")
    print(f"  {n} rows total — {n_mapped} in mapped leagues (priority), {n_unmapped} in unmapped leagues")
    print(f"  {n_filled} wages already filled in; {n - n_filled} blank → wage_feasibility falls back to 0.7")
    print()
    print("Top 20 priority rows (mapped league, sellability DESC):")
    print(f"  {'name':28s} {'pos':6s} {'lg':5s} {'sell':>5s}  {'TM':>5s}  {'wage_pw':>10s}")
    n_shown = 0
    for r in rows:
        if r["buyer_demand_mapped"] != "yes":
            continue
        wage = r["wage_pw_eur"]
        wage_str = f"€{wage:,}" if wage else "(blank)"
        print(f"  {(r['name'] or '')[:28]:28s} {r['position']:6s} {r['league']:5s} "
              f"{r['sellability_score']:>5.1f}  €{(r['TM_value_eur'] or 0)/1e6:>3.0f}m  {wage_str:>10s}")
        n_shown += 1
        if n_shown >= 20:
            break


if __name__ == "__main__":
    main()
