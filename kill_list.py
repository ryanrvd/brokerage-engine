"""Shared Kill List module — single source of truth.

Two consumers:
  • scripts/23_export_brokerage_sheet.py  — writes Sheet 1 with Kill List exclusions
                                            and maintains the Kill List tab itself.
  • app/db.py                              — reads live Kill List + agency rules
                                            to filter the Streamlit Targets view.

The Kill List has two sources of exclusion:
  1. MANUAL entries  — user-typed rows in BrokerageWorkbook.xlsx → Kill List tab.
                       Preserved verbatim across pipeline runs.
  2. AGENCY rules    — players whose `agency` field matches any string in
                       data/blocked_agencies.csv (case-insensitive substring).
                       Auto-derived on every run; not user-editable inside the tab.

A "source" column in the Kill List tab distinguishes them. When an agency rule
matches a player who's ALSO listed manually, manual wins (no duplicate row).

Fuzzy matching for manual entries:
  Pass 1 — exact normalised match (lowercase, diacritics stripped, punctuation removed)
  Pass 2 — token-subset (every kill-list token appears in the player's name —
           handles surname-only entries like "Tzolis")
  Pass 3 — SequenceMatcher ratio >= 0.85 fallback
"""

from __future__ import annotations

import csv
import sqlite3
import unicodedata
from difflib import SequenceMatcher
from pathlib import Path

WORKBOOK_PATH = Path("BrokerageWorkbook.xlsx")
KILL_LIST_SHEET = "Kill List"
BLOCKED_AGENCIES_CSV = Path("data/blocked_agencies.csv")
KILL_LIST_HEADERS = ["player", "reason why", "source"]
FUZZY_THRESHOLD = 0.85

# Default seeds — used when the Kill List tab doesn't exist yet (first run).
SEED_MANUAL_ENTRIES: list[tuple[str, str]] = [
    ("Zeno Van Den Bosch", "Going free — Bosman this summer"),
    ("Arouna Sangante",    "Going free — contract end June 2026"),
    ("José Ángel Carmona", "Signed exclusive mandate with Unique Sports Group"),
]

SEED_BLOCKED_AGENCIES: list[str] = [
    "CAA Base", "Gestifute", "Gol International", "Unique Sports Group",
]


# ─── Name normalisation ──────────────────────────────────────────────────────

def normalise_name(s: str | None) -> str:
    """Lowercase, strip diacritics, collapse whitespace, drop punctuation."""
    if not s:
        return ""
    nfkd = unicodedata.normalize("NFKD", str(s))
    ascii_only = "".join(c for c in nfkd if not unicodedata.combining(c))
    cleaned = "".join(c if c.isalnum() or c.isspace() else " " for c in ascii_only)
    return " ".join(cleaned.lower().split())


# ─── Blocked agencies CSV ────────────────────────────────────────────────────

def read_blocked_agencies(csv_path: Path = BLOCKED_AGENCIES_CSV) -> list[str]:
    """Returns a list of agency strings (one per row). Missing file → empty list."""
    if not csv_path.exists():
        return []
    out: list[str] = []
    with csv_path.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            ag = (row.get("agency") or "").strip()
            if ag:
                out.append(ag)
    return out


def seed_blocked_agencies_if_missing(csv_path: Path = BLOCKED_AGENCIES_CSV) -> None:
    if csv_path.exists():
        return
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["agency"])
        for ag in SEED_BLOCKED_AGENCIES:
            writer.writerow([ag])


# ─── Kill List xlsx I/O ──────────────────────────────────────────────────────

def read_kill_list_tab(workbook_path: Path = WORKBOOK_PATH) -> dict:
    """Returns {'manual': [(name, reason)], 'auto': [(name, reason, source)]}.

    Auto rows are read but discarded by callers — they're regenerated on every
    pipeline run. Manual rows are preserved verbatim. Missing tab → seed.
    """
    if not workbook_path.exists():
        return {"manual": list(SEED_MANUAL_ENTRIES), "auto": []}
    from openpyxl import load_workbook
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        if KILL_LIST_SHEET not in wb.sheetnames:
            return {"manual": list(SEED_MANUAL_ENTRIES), "auto": []}
        ws = wb[KILL_LIST_SHEET]
        manual: list[tuple[str, str]] = []
        auto: list[tuple[str, str, str]] = []
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return {"manual": list(SEED_MANUAL_ENTRIES), "auto": []}
    headers = [str(h) if h is not None else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}
    name_i   = idx.get("player", 0)
    reason_i = idx.get("reason why", 1)
    source_i = idx.get("source", -1)
    for row in rows[1:]:
        if not row or row[name_i] is None or str(row[name_i]).strip() == "":
            continue
        name   = str(row[name_i]).strip()
        reason = str(row[reason_i] or "").strip() if reason_i < len(row) else ""
        source = (str(row[source_i]).strip()
                  if source_i >= 0 and source_i < len(row) and row[source_i] is not None
                  else "manual")
        if source == "manual" or source == "":
            manual.append((name, reason))
        else:
            auto.append((name, reason, source))
    return {"manual": manual, "auto": auto}


# ─── Matching ────────────────────────────────────────────────────────────────

def match_manual_entries(
    manual_entries: list[tuple[str, str]],
    player_rows: list[dict],
) -> tuple[set[int], list[dict]]:
    """For each manual Kill List entry, find the matching player_id(s).

    Returns (excluded_player_ids, warnings).
    warnings: list of dicts {entry, matches, kind} where kind ∈ {'zero', 'ambiguous'}.
    Ambiguous entries are NOT applied — better to under-drop than over-drop.
    """
    players_norm = [
        (p["player_id"], p["player_name"], normalise_name(p["player_name"]))
        for p in player_rows
    ]
    excluded: set[int] = set()
    warnings: list[dict] = []

    for entry_name, _reason in manual_entries:
        e_norm = normalise_name(entry_name)
        if not e_norm:
            continue
        e_tokens = set(e_norm.split())

        # Pass 1: exact normalised match
        hits = [(pid, pn) for (pid, pn, p_norm) in players_norm if p_norm == e_norm]

        # Pass 2: token-subset match (kill entry's tokens are all present in player name)
        if not hits:
            hits = [
                (pid, pn) for (pid, pn, p_norm) in players_norm
                if e_tokens and e_tokens.issubset(set(p_norm.split()))
            ]

        # Pass 3: fuzzy ratio
        if not hits:
            for (pid, pn, p_norm) in players_norm:
                if SequenceMatcher(None, e_norm, p_norm).ratio() >= FUZZY_THRESHOLD:
                    hits.append((pid, pn))

        if len(hits) == 0:
            warnings.append({"entry": entry_name, "matches": [], "kind": "zero"})
        elif len(hits) > 1:
            warnings.append({"entry": entry_name, "matches": hits, "kind": "ambiguous"})
        else:
            excluded.add(hits[0][0])

    return excluded, warnings


def match_agency_rules(
    player_rows: list[dict],
    blocked_agencies: list[str],
) -> list[dict]:
    """Returns one dict per matching player:
       {player_id, player_name, agency, matched_rule}

    Word-boundary case-insensitive match: the rule must appear as a complete
    word (or word-sequence) inside the player's agency string. This catches
    'CAA Base' in 'CAA Base Ltd' and 'EPIC' in 'Epic Atubolu', but does NOT
    over-fire on 'SP' being a substring of 'Sports' (which the previous
    bidirectional plain-substring match incorrectly killed).
    """
    import re
    patterns = [
        (re.compile(r"\b" + re.escape(r) + r"\b", re.IGNORECASE), r)
        for r in blocked_agencies
    ]
    out: list[dict] = []
    for p in player_rows:
        agency = (p.get("agency") or "").strip()
        if not agency:
            continue
        for pat, rule_original in patterns:
            if pat.search(agency):
                out.append({
                    "player_id":    p["player_id"],
                    "player_name":  p["player_name"],
                    "agency":       agency,
                    "matched_rule": rule_original,
                })
                break
    return out


# ─── Combined computation: one entry point ──────────────────────────────────

def compute_kill_list_state(
    con: sqlite3.Connection,
    workbook_path: Path = WORKBOOK_PATH,
    blocked_agencies_csv: Path = BLOCKED_AGENCIES_CSV,
) -> dict:
    """One-stop computation used by both scripts/23 and app/db.

    Returns:
      {
        'manual_entries':   [(name, reason)] — preserve verbatim,
        'manual_excluded_ids': set[int] — players matched from manual entries,
        'auto_rows':        [{player_name, reason, source, player_id}, ...]
                            — to render in the Kill List tab AND in Excluded page,
        'auto_excluded_ids': set[int],
        'excluded_ids':     set[int] — union, deduped,
        'warnings':         list[{entry, matches, kind}] — manual match failures,
        'blocked_agencies': list[str] — currently in force,
      }
    """
    # Player universe restricted to the sellable cohort (= Sheet 4 filter)
    rows = list(con.execute("""
        SELECT player_id, name AS player_name, agency, current_club, parent_club,
               position_bucket, league_id, current_tm_value_eur, sellability_score
        FROM player_universe
        WHERE (right_priced=1 OR finished_product=1
               OR finished_product IS NULL OR contract_leveraged=1)
    """))
    cols = ["player_id", "player_name", "agency", "current_club", "parent_club",
            "position_bucket", "league_id", "current_tm_value_eur", "sellability_score"]
    player_rows = [dict(zip(cols, r)) for r in rows]

    state = read_kill_list_tab(workbook_path)
    manual_entries = state["manual"]

    manual_excluded_ids, warnings = match_manual_entries(manual_entries, player_rows)

    blocked_agencies = read_blocked_agencies(blocked_agencies_csv)
    agency_hits = match_agency_rules(player_rows, blocked_agencies)

    # Drop agency hits whose player_id is already in the manual exclusion set
    # (so we don't write a duplicate row for Carmona-type cases).
    auto_rows = []
    auto_excluded_ids: set[int] = set()
    for hit in agency_hits:
        if hit["player_id"] in manual_excluded_ids:
            continue
        auto_excluded_ids.add(hit["player_id"])
        auto_rows.append({
            "player_id":   hit["player_id"],
            "player_name": hit["player_name"],
            "reason":      f"Agency: {hit['agency']} (does not collaborate)",
            "source":      f"agency:{hit['matched_rule']}",
        })

    return {
        "manual_entries":      manual_entries,
        "manual_excluded_ids": manual_excluded_ids,
        "auto_rows":           auto_rows,
        "auto_excluded_ids":   auto_excluded_ids,
        "excluded_ids":        manual_excluded_ids | auto_excluded_ids,
        "warnings":            warnings,
        "blocked_agencies":    blocked_agencies,
        "player_rows":         player_rows,  # used by Excluded page for context
    }
