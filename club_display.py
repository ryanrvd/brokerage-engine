"""Club Display Names — single source of truth, shared by pipeline scripts and the Streamlit app.

Same architecture as kill_list.py:
  • Read-preserve-write: rows tagged 'manual' are preserved verbatim across
    pipeline runs; rows tagged 'auto' get regenerated each run.
  • Tab lives in BrokerageWorkbook.xlsx → "Club Display Names".
  • Pipeline scripts call read_display_map(workbook_path) at the top of each
    run; rendering code looks up club_id → display_name.

The auto-generation has two layers:
  1. EXPLICIT_NAMES — hand-curated mappings for clubs where the official name
     can't be reduced by pattern rules (long Dutch names, Greek transliterations,
     non-obvious short forms like "Bayer 04 Leverkusen" → "Bayer Leverkusen").
  2. _auto_simplify(name) — conservative suffix/prefix stripping for the rest.

Low-confidence rule: any auto-generated display_name that's >25 chars OR still
contains structural markers (e.g. "Football Club", "S.A.D.") after stripping
gets flagged "low confidence — review" in the notes column.
"""

from __future__ import annotations

import re
import sqlite3
from pathlib import Path

WORKBOOK_PATH = Path("BrokerageWorkbook.xlsx")
DISPLAY_TAB = "Club Display Names"
LOW_CONFIDENCE_NOTE = "low confidence — review"
LOW_CONFIDENCE_MAX_LEN = 25


# ─── Explicit mappings (football-domain knowledge) ───────────────────────────

EXPLICIT_NAMES: dict[str, str] = {
    # Spanish — Real stays per user spec, except Mallorca
    "Real Madrid Club de Fútbol":                            "Real Madrid",
    "Real Sociedad de Fútbol S.A.D.":                        "Real Sociedad",
    "Real Sociedad B":                                       "Real Sociedad B",
    "Real Betis Balompié S.A.D.":                            "Real Betis",
    "Real Oviedo S.A.D.":                                    "Real Oviedo",
    "Real Valladolid CF":                                    "Real Valladolid",
    "Real Club Deportivo Mallorca S.A.D.":                   "Mallorca",
    "Real Club Deportivo Mallorca":                          "Mallorca",
    "Club Atlético de Madrid S.A.D.":                        "Atlético Madrid",
    "Athletic Club Bilbao":                                  "Athletic Club",
    "Reial Club Deportiu Espanyol de Barcelona S.A.D.":      "Espanyol",
    "Reial Club Celta de Vigo":                              "Celta Vigo",
    "Getafe Club de Fútbol S. A. D. Team Dubai":             "Getafe",
    "Sevilla Fútbol Club S.A.D.":                            "Sevilla",
    "Girona Fútbol Club S. A. D.":                           "Girona",
    "Valencia Club de Fútbol":                               "Valencia",
    "Villarreal Club de Fútbol":                             "Villarreal",
    "Rayo Vallecano de Madrid S.A.D.":                       "Rayo Vallecano",
    "Cádiz Club de Fútbol":                                  "Cádiz",
    "Granada Club de Fútbol":                                "Granada",
    "Las Palmas":                                            "Las Palmas",
    "Unión Deportiva Almería S.A.D.":                        "Almería",
    "Club Deportivo Leganés":                                "Leganés",
    "Club Deportivo Alavés":                                 "Alavés",
    "Reial Club Deportiu Mallorca":                          "Mallorca",
    # Italian
    "Football Club Internazionale Milano S.p.A.":            "Inter Milan",
    "Associazione Calcio Milan":                             "AC Milan",
    "Associazione Sportiva Roma":                            "AS Roma",
    "Società Sportiva Calcio Napoli":                        "Napoli",
    "Società Sportiva Lazio S.p.A.":                         "Lazio",
    "Juventus Football Club":                                "Juventus",
    "Atalanta Bergamasca Calcio S.p.A.":                     "Atalanta",
    "Atalanta BC":                                           "Atalanta",
    "Bologna FC 1909":                                       "Bologna",
    "Genoa Cricket and Football Club":                       "Genoa",
    "ACF Fiorentina":                                        "Fiorentina",
    "Hellas Verona FC":                                      "Hellas Verona",
    "Torino FC":                                             "Torino",
    "Udinese Calcio":                                        "Udinese",
    "FC Empoli":                                             "Empoli",
    "Pisa Sporting Club":                                    "Pisa",
    "US Sassuolo Calcio":                                    "Sassuolo",
    "Unione Sportiva Cremonese S.p.A.":                      "Cremonese",
    "Unione Sportiva Lecce":                                 "Lecce",
    "Parma Calcio 1913":                                     "Parma",
    "Calcio Como":                                           "Como",
    "Venezia":                                               "Venezia",
    # German — keep "1. FC", "VfB", "VfL", "TSG", "HSV", "Bayer" per user spec
    "1. Fußballclub Union Berlin":                           "1. FC Union Berlin",
    "1. Fußball-Club Köln":                                  "1. FC Köln",
    "1. Fußball- und Sportverein Mainz 05":                  "Mainz 05",
    "FC Bayern München":                                     "Bayern Munich",
    "FC Bayern Munich":                                      "Bayern Munich",
    "FC Bayern Munich II":                                   "Bayern Munich II",
    "Verein für Bewegungsspiele Stuttgart 1893":             "VfB Stuttgart",
    "Verein für Leibesübungen Wolfsburg":                    "VfL Wolfsburg",
    "VfL Bochum":                                            "VfL Bochum",
    "Turn- und Sportgemeinschaft 1899 Hoffenheim Fußball-Spielbetriebs": "TSG Hoffenheim",
    "Borussia Verein für Leibesübungen 1900 Mönchengladbach":"Borussia Mönchengladbach",
    "Hamburger Sport Verein":                                "Hamburger SV",
    "Eintracht Frankfurt Fußball AG":                        "Eintracht Frankfurt",
    "Borussia Dortmund":                                     "Borussia Dortmund",
    "Bayer 04 Leverkusen Fußball":                           "Bayer Leverkusen",
    "RasenBallsport Leipzig":                                "RB Leipzig",
    "Sport-Club Freiburg":                                   "SC Freiburg",
    "Sportverein Werder Bremen von 1899":                    "Werder Bremen",
    "FC St. Pauli":                                          "St. Pauli",
    "Holstein Kiel":                                         "Holstein Kiel",
    "FC Augsburg":                                           "FC Augsburg",
    "Eintracht Braunschweig":                                "Eintracht Braunschweig",
    "Hertha BSC":                                            "Hertha BSC",
    "SpVgg Greuther Fürth":                                  "Greuther Fürth",
    "1. FC Heidenheim 1846":                                 "1. FC Heidenheim",
    "1. FC Nürnberg":                                        "1. FC Nürnberg",
    "1. FC Magdeburg":                                       "1. FC Magdeburg",
    "1. FC Kaiserslautern":                                  "1. FC Kaiserslautern",
    "Schalke 04":                                            "Schalke 04",
    "Fortuna Düsseldorf":                                    "Fortuna Düsseldorf",
    "Hannover 96":                                           "Hannover 96",
    "Karlsruher SC":                                         "Karlsruher SC",
    "Darmstadt 98":                                          "Darmstadt 98",
    "Preußen Münster":                                       "Preußen Münster",
    "SV Elversberg":                                         "SV Elversberg",
    "SC Paderborn 07":                                       "SC Paderborn",
    "Jahn Regensburg":                                       "Jahn Regensburg",
    "FC Erzgebirge Aue":                                     "Erzgebirge Aue",
    "Arminia Bielefeld":                                     "Arminia Bielefeld",
    # French
    "Paris Saint-Germain Football Club":                     "PSG",
    "Olympique de Marseille":                                "Marseille",
    "Olympique Gymnaste Club Nice Côte d'Azur":              "Nice",
    "Olympique Lyonnais":                                    "Lyon",
    "Association sportive de Monaco Football Club":          "AS Monaco",
    "Association de la Jeunesse auxerroise":                 "Auxerre",
    "Football Club Lorient-Bretagne Sud":                    "Lorient",
    "Lille Olympique Sporting Club":                         "Lille",
    "Lille OSC":                                             "Lille",
    "Stade Rennais FC":                                      "Rennes",
    "Football Club de Nantes":                               "Nantes",
    "Stade de Reims":                                        "Reims",
    "Toulouse Football Club":                                "Toulouse",
    "Racing Club de Strasbourg Alsace":                      "Strasbourg",
    "Le Havre Athletic Club":                                "Le Havre",
    "Montpellier Hérault Sport Club":                        "Montpellier",
    "Angers Sporting Club de l'Ouest":                       "Angers",
    "Stade Brestois 29":                                     "Brest",
    "Clermont Foot 63":                                      "Clermont",
    # Portuguese
    "Sport Lisboa e Benfica":                                "Benfica",
    "Sporting Clube de Portugal":                            "Sporting CP",
    "Sporting Clube de Braga":                               "Braga",
    "Futebol Clube do Porto":                                "Porto",
    "FC Porto":                                              "Porto",
    "Vitória Sport Clube":                                   "Vitória SC",
    "Vitória Sport Clube Guimarães":                         "Vitória SC",
    "Club Football Estrela da Amadora":                      "Estrela",
    "Futebol Clube Famalicão":                               "Famalicão",
    "Futebol Clube de Alverca":                              "Alverca",
    "Casa Pia Atlético Clube":                               "Casa Pia",
    "AVS Futebol SAD":                                       "AVS",
    "Grupo Desportivo Estoril Praia":                        "Estoril",
    "Rio Ave Futebol Clube":                                 "Rio Ave",
    "Moreirense Futebol Clube":                              "Moreirense",
    "Clube Desportivo Nacional":                             "Nacional",
    "Clube Desportivo Santa Clara":                          "Santa Clara",
    "Clube Desportivo Tondela":                              "Tondela",
    "Gil Vicente Futebol Clube":                             "Gil Vicente",
    "Futebol Clube de Arouca":                               "Arouca",
    # Dutch
    "AFC Ajax Amsterdam":                                    "Ajax",
    "Eindhovense Voetbalvereniging Philips Sport Vereniging":"PSV Eindhoven",
    "Feyenoord Rotterdam":                                   "Feyenoord",
    "Alkmaarse Zaanlandse Voetbalclub":                      "AZ Alkmaar",
    "AZ Alkmaar":                                            "AZ Alkmaar",
    "Football Club Twente":                                  "Twente",
    "FC Twente":                                             "Twente",
    "FC Utrecht":                                            "Utrecht",
    "Football Club Utrecht":                                 "Utrecht",
    "NEC Nijmegen":                                          "NEC Nijmegen",
    "Nooit Opgeven Altijd Doorzetten Aangenaam Door Vermaak En Nuttig Door Ontspanning Combinatie Breda": "NAC Breda",
    "NAC Breda":                                             "NAC Breda",
    "Prins Hendrik Ende Desespereert Nimmer Combinatie Zwolle": "PEC Zwolle",
    "PEC Zwolle":                                            "PEC Zwolle",
    "Football Club Groningen":                               "Groningen",
    "FC Groningen":                                          "Groningen",
    "Heracles Almelo":                                       "Heracles",
    "Go Ahead Eagles":                                       "Go Ahead Eagles",
    "SC Heerenveen":                                         "Heerenveen",
    "Excelsior Rotterdam":                                   "Excelsior",
    "Sparta Rotterdam":                                      "Sparta Rotterdam",
    "Fortuna Sittard":                                       "Fortuna Sittard",
    "FC Volendam":                                           "Volendam",
    "SC Telstar":                                            "Telstar",
    # Belgian
    "Club Brugge Koninklijke Voetbalvereniging":             "Club Brugge",
    "Cercle Brugge Koninklijke Sportvereniging":             "Cercle Brugge",
    "Royale Union Saint-Gilloise":                           "Union Saint-Gilloise",
    "Royal Antwerp Football Club":                           "Royal Antwerp",
    "Koninklijke Atletiek Associatie Gent":                  "KAA Gent",
    "RSC Anderlecht":                                        "Anderlecht",
    "KRC Genk":                                              "Genk",
    "Standard Liege":                                        "Standard Liège",
    "Standard de Liège":                                     "Standard Liège",
    "Yellow-Red Koninklijke Voetbalclub Mechelen":           "KV Mechelen",
    "Koninklijke Sint-Truidense Voetbalvereniging":          "Sint-Truiden",
    "KVC Westerlo":                                          "Westerlo",
    "Oud-Heverlee Leuven":                                   "OH Leuven",
    "Royal Charleroi SC":                                    "Charleroi",
    "Zulte Waregem":                                         "Zulte Waregem",
    "FC Verbroedering Denderhoutem Denderleeuw Eendracht Hekelgem": "Dender",
    "FCV Dender EH":                                         "Dender",
    "RAAL La Louvière":                                      "RAAL La Louvière",
    # Turkish
    "Beşiktaş Jimnastik Kulübü":                             "Beşiktaş",
    "Galatasaray Spor Kulübü":                               "Galatasaray",
    "Fenerbahçe Spor Kulübü":                                "Fenerbahçe",
    "Trabzonspor Kulübü":                                    "Trabzonspor",
    "Başakşehir Futbol Kulübü":                              "Başakşehir",
    "Adana Demirspor Kulübü":                                "Adana Demirspor",
    "Kayserispor Kulübü":                                    "Kayserispor",
    "Antalyaspor":                                           "Antalyaspor",
    "Konyaspor":                                             "Konyaspor",
    "Sivasspor":                                             "Sivasspor",
    "Alanyaspor":                                            "Alanyaspor",
    "Gaziantep FK":                                          "Gaziantep",
    "Fatih Karagümrük Sportif Faaliyetler San. Tic. A.Ş.":   "Fatih Karagümrük",
    # Greek
    "Olympiakos Syndesmos Filathlon Peiraios":               "Olympiakos",
    "Panthessalonikios Athlitikos Omilos Konstantinoupoliton":"PAOK",
    "Athlitiki Enosi Konstantinoupoleos":                    "AEK Athens",
    "Athlitiki Enosi Larisas":                               "AEK Larissa",
    "Panathinaikos Athlitikos Omilos":                       "Panathinaikos",
    "Aris Thessalonikis":                                    "Aris",
    "Panaitolikos GFS":                                      "Panetolikos",
    "Panetolikos Agrinio":                                   "Panetolikos",
    "Panserraikos Serres":                                   "Panserraikos",
    "OFI Kritis":                                            "OFI Crete",
    "Volos NFC":                                             "Volos",
    "Asteras Tripolis":                                      "Asteras Tripolis",
    "Lamia":                                                 "Lamia",
    # English
    "Brighton and Hove Albion Football Club":                "Brighton",
    "Association Football Club Bournemouth":                 "Bournemouth",
    "Wolverhampton Wanderers Football Club":                 "Wolves",
    "Tottenham Hotspur Football Club":                       "Tottenham",
    "Newcastle United Football Club":                        "Newcastle",
    "Sunderland Association Football Club":                  "Sunderland",
    "Leeds United Association Football Club":                "Leeds",
    "Liverpool FC":                                          "Liverpool",
    "Arsenal FC":                                            "Arsenal",
    "Chelsea FC":                                            "Chelsea",
    "Manchester United":                                     "Man United",
    "Manchester City Football Club":                         "Man City",
    "Manchester City FC":                                    "Man City",
    "West Ham United":                                       "West Ham",
    "Crystal Palace":                                        "Crystal Palace",
    "Brentford FC":                                          "Brentford",
    "Fulham FC":                                             "Fulham",
    "Everton FC":                                            "Everton",
    "Nottingham Forest":                                     "Nottingham Forest",
    "Aston Villa":                                           "Aston Villa",
    "Burnley FC":                                            "Burnley",
    "AFC Bournemouth":                                       "Bournemouth",
    # Saudi
    "Al-Hilal Saudi Football Club":                          "Al-Hilal",
    "Al-Nassr Football Club":                                "Al-Nassr",
    "Al-Ahli Saudi Football Club":                           "Al-Ahli",
    "Al-Ittihad Saudi Football Club":                        "Al-Ittihad",
    # MLS (limited interest; common forms)
    "Internacional de Toronto":                              "Toronto FC",
    "Club Internacional de Fútbol Miami":                    "Inter Miami",
    "Inter Miami CF":                                        "Inter Miami",
    "Atlanta United Football Club":                          "Atlanta United",
    "Los Angeles Football Club":                             "LAFC",
    "New York City Football Club":                           "NYCFC",
    "Real Salt Lake":                                        "Real Salt Lake",
    "Real Salt Lake B":                                      "Real Salt Lake B",
    "FC Cincinnati":                                          "FC Cincinnati",
    # Scottish
    "Celtic FC":                                             "Celtic",
    "Rangers FC":                                            "Rangers",
    "Heart of Midlothian":                                   "Hearts",
    "Hibernian FC":                                          "Hibernian",
    "Aberdeen Football Club":                                "Aberdeen",
    # Danish
    "Fodbold Club Midtjylland":                              "FC Midtjylland",
    "FC København":                                          "FC Copenhagen",
    "FC Kobenhavn":                                          "FC Copenhagen",
    "Brøndby IF":                                            "Brøndby",
    "Aarhus GF":                                             "AGF",
    "FC Nordsjælland":                                       "FC Nordsjælland",
    "Vejle Boldklub":                                        "Vejle",
}


# ─── Fallback auto-simplifier ────────────────────────────────────────────────

# Suffixes stripped from the end (lowercased for matching)
_SUFFIX_DROPS = [
    " s.a.d.", " s. a. d.", " s.a.", " s. a.", " s.p.a.", " s.p.a", " s.r.l.",
    " a.ş.", " a.s.", " ag", " b.v.", " ltda.", " ltd.", " limited",
    " football club", " association football club", " fußball ag", " fußball",
    " futebol clube", " sport clube", " spor kulübü", " jimnastik kulübü",
    " voetbalvereniging", " sportvereniging", " calcio",
]
_LEADING_DROPS = [
    "football club ", "association football club ", "club ",
]
_KEEP_PREFIX_PATTERNS = (
    re.compile(r"^1\. f[uü]ßball", re.IGNORECASE),
    re.compile(r"^real\s", re.IGNORECASE),
    re.compile(r"^atl[eé]tico\s", re.IGNORECASE),
    re.compile(r"^athletic\s", re.IGNORECASE),
)


def _auto_simplify(name: str) -> str:
    if not name:
        return name
    s = name.strip()
    # Don't strip leading anything if the name starts with a protected prefix
    is_protected = any(p.match(s) for p in _KEEP_PREFIX_PATTERNS)
    # Strip trailing forms (case-insensitive)
    changed = True
    while changed:
        changed = False
        s_lower = s.lower()
        for suf in _SUFFIX_DROPS:
            if s_lower.endswith(suf):
                s = s[: len(s) - len(suf)].rstrip(", .")
                changed = True
                break
    # Strip leading forms (case-insensitive, only if not protected)
    if not is_protected:
        s_lower = s.lower()
        for pre in _LEADING_DROPS:
            if s_lower.startswith(pre):
                s = s[len(pre):]
                break
    return s.strip()


def auto_display_for_name(official_name: str) -> tuple[str, bool]:
    """Returns (display_name, is_low_confidence).
    If the name is in EXPLICIT_NAMES, use that and is_low_confidence=False.
    Otherwise apply _auto_simplify and flag low-confidence if the result is
    still long or contains structural markers."""
    if official_name in EXPLICIT_NAMES:
        return EXPLICIT_NAMES[official_name], False
    simplified = _auto_simplify(official_name)
    low_conf_markers = (
        "football club", "voetbal", "spor", "calcio", "verein",
        "futebol", "deportiv", "fußball", "sportif", "spor kulübü",
        "s.a.d", "s.p.a", "s.a.", "association",
    )
    is_low = (
        len(simplified) > LOW_CONFIDENCE_MAX_LEN
        or any(m in simplified.lower() for m in low_conf_markers)
    )
    return simplified, is_low


# ─── Read / write the tab ────────────────────────────────────────────────────

def read_display_tab(workbook_path: Path = WORKBOOK_PATH) -> dict[int, dict]:
    """Returns {club_id: {display_name, auto_or_manual, notes}}.
    Empty dict if the tab doesn't exist yet."""
    if not workbook_path.exists():
        return {}
    from openpyxl import load_workbook
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        if DISPLAY_TAB not in wb.sheetnames:
            return {}
        ws = wb[DISPLAY_TAB]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return {}
    headers = [str(h) if h is not None else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}
    if "club_id" not in idx or "display_name" not in idx:
        return {}
    out: dict[int, dict] = {}
    for row in rows[1:]:
        if not row or row[idx["club_id"]] is None:
            continue
        try:
            cid = int(row[idx["club_id"]])
        except (TypeError, ValueError):
            continue
        out[cid] = {
            "display_name":   str(row[idx.get("display_name", -1)] or "").strip(),
            "auto_or_manual": str(row[idx.get("auto_or_manual", -1)] or "auto").strip().lower(),
            "notes":          str(row[idx.get("notes", -1)] or "").strip()
                              if idx.get("notes", -1) >= 0 and row[idx.get("notes", -1)] is not None
                              else "",
        }
    return out


def compute_display_state(
    con: sqlite3.Connection,
    workbook_path: Path = WORKBOOK_PATH,
) -> dict:
    """Build the full display-name state.

    Returns:
      {
        'rows':          list of {club_id, official_name, league_id, display_name, auto_or_manual, notes}
                         — every club, sorted by league then official_name,
        'display_map':   dict[club_id] = display_name (the lookup table everything uses),
      }

    For each club:
      • If a row exists with auto_or_manual='manual' → preserve display_name and notes.
      • Otherwise → regenerate from auto_display_for_name(official_name) (auto rows
        are refreshed every run so future rule improvements propagate).
      • Any club not previously in the tab → added as auto.
    """
    existing = read_display_tab(workbook_path)

    clubs = list(con.execute("""
        SELECT club_id, name AS official_name, league_id
        FROM club_pressure
        ORDER BY league_id, name
    """))

    rows: list[dict] = []
    display_map: dict[int, str] = {}

    for cid, official, league_id in clubs:
        prev = existing.get(int(cid))
        if prev and prev.get("auto_or_manual") == "manual":
            display_name = prev["display_name"] or official
            auto_or_manual = "manual"
            notes = prev.get("notes", "")
        else:
            display_name, is_low = auto_display_for_name(official)
            auto_or_manual = "auto"
            notes = LOW_CONFIDENCE_NOTE if is_low else ""

        rows.append({
            "club_id":         int(cid),
            "official_name":   official,
            "league_id":       league_id,
            "display_name":    display_name,
            "auto_or_manual":  auto_or_manual,
            "notes":           notes,
        })
        display_map[int(cid)] = display_name

    return {
        "rows":        rows,
        "display_map": display_map,
    }


def write_display_tab(wb, rows: list[dict], league_names: dict[str, str]) -> None:
    """Write the Club Display Names tab to an open openpyxl workbook.
    Insert as 'Club Display Names' (last position by default; caller can reorder)."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    if DISPLAY_TAB in wb.sheetnames:
        del wb[DISPLAY_TAB]
    ws = wb.create_sheet(DISPLAY_TAB)

    HEADERS = ["club_id", "official_name", "league", "display_name", "auto_or_manual", "notes"]
    WIDTHS  = [10, 50, 22, 28, 14, 28]

    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F3864")
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = bold; c.fill = fill; c.alignment = centre
    ws.row_dimensions[1].height = 28

    ro_fill = PatternFill("solid", fgColor="F2F2F2")
    in_fill = PatternFill("solid", fgColor="FFF2CC")
    low_fill = PatternFill("solid", fgColor="FCE4D6")
    wrap = Alignment(wrap_text=True, vertical="top")

    sorted_rows = sorted(rows, key=lambda r: (r["league_id"] or "", r["official_name"] or ""))
    for ri, r in enumerate(sorted_rows, start=2):
        ws.cell(row=ri, column=1, value=r["club_id"]).fill = ro_fill
        c2 = ws.cell(row=ri, column=2, value=r["official_name"])
        c2.fill = ro_fill; c2.alignment = wrap
        ws.cell(row=ri, column=3, value=league_names.get(r["league_id"], r["league_id"])).fill = ro_fill
        c4 = ws.cell(row=ri, column=4, value=r["display_name"])
        c4.fill = in_fill; c4.alignment = wrap
        c5 = ws.cell(row=ri, column=5, value=r["auto_or_manual"])
        c5.fill = in_fill; c5.alignment = Alignment(horizontal="center")
        c6 = ws.cell(row=ri, column=6, value=r["notes"])
        c6.fill = low_fill if r["notes"] else ro_fill
        c6.alignment = wrap

    # Data validation on auto_or_manual: must be exactly 'auto' or 'manual'
    dv = DataValidation(type="list", formula1='"auto,manual"', allow_blank=False)
    dv.error = "Use 'auto' or 'manual'."
    dv.errorTitle = "Invalid value"
    letter = get_column_letter(5)
    dv.add(f"{letter}2:{letter}{len(sorted_rows) + 1}")
    ws.add_data_validation(dv)

    for ci, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "D2"
    last_col = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{len(sorted_rows) + 1}"


# ─── Lightweight lookup for non-pipeline consumers (e.g. Streamlit) ──────────

def load_display_map(workbook_path: Path = WORKBOOK_PATH) -> dict[int, str]:
    """Returns {club_id: display_name}. Empty dict if tab not present."""
    existing = read_display_tab(workbook_path)
    return {cid: v["display_name"] for cid, v in existing.items() if v.get("display_name")}


def display_for(club_id: int | None, official_fallback: str | None,
                display_map: dict[int, str]) -> str:
    """Lookup helper for rendering. Falls back to official_fallback if no
    mapping exists for this club_id."""
    if club_id is None:
        return official_fallback or ""
    return display_map.get(int(club_id)) or (official_fallback or "")
