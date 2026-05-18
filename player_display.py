"""Player Display Names — single source of truth, shared by pipeline scripts and the Streamlit app.

Mirrors club_display.py's architecture:
  • Read-preserve-write: 'manual' rows are preserved verbatim across pipeline
    runs; 'auto' rows get regenerated each run.
  • Tab lives in BrokerageWorkbook.xlsx → "Player Display Names".
  • Pipeline scripts call read_display_map(workbook_path) at startup; rendering
    code looks up player_id → display_name.

The auto-generation has two layers:
  1. EXPLICIT_NAMES — hand-curated mappings encoding football-domain knowledge
     (nicknames like "Joe Gelhardt" / "Tommy Doyle", ordering quirks like
     "Abdul Fatawu", Spanish dual-surname drops like "Sergio Arribas Calvo" →
     "Sergio Arribas").
  2. _auto_simplify(name) — pattern rules for the rest:
     • 1-2 tokens → no change (already short)
     • Particle anywhere (Dutch "van/de/den", Portuguese "da/dos/de", etc.)
       → preserve full name (structural particle is part of the surname)
     • Compound first names (José Ángel / Juan Carlos / João Pedro) →
       preserve first two tokens + last
     • Default (3+ tokens, no particle, no compound first) → first + last
       (Anglo middle-name strip)

Low-confidence flag: any auto-generated display_name that's noticeably long or
contains unusual structural markers gets flagged "low confidence — review".
"""

from __future__ import annotations

import sqlite3
from pathlib import Path

WORKBOOK_PATH = Path("BrokerageWorkbook.xlsx")
DISPLAY_TAB = "Player Display Names"
LOW_CONFIDENCE_NOTE = "low confidence — review"
LOW_CONFIDENCE_MAX_TOKENS = 3   # auto display > 3 tokens → flag


# ─── Explicit mappings (football-domain knowledge) ───────────────────────────

EXPLICIT_NAMES: dict[str, str] = {
    # Anglo nicknames (BBC / Sky / Athletic universal usage)
    "Joseph Paul Gelhardt":         "Joe Gelhardt",
    "Thomas Glyn Doyle":            "Tommy Doyle",
    "Tommy Daniel John Conway":     "Tommy Conway",
    "Thomas Christopher Cannon":    "Tommy Cannon",
    "Memeh Caleb Okoli":            "Caleb Okoli",
    "Carl Andrew Rushworth":        "Carl Rushworth",
    "Cameron Desmond Archer":       "Cameron Archer",
    "William Anthony Patrick Saliba": "William Saliba",
    "Harrison James Burrows":       "Harrison Burrows",
    "Romain Joy Kouakou Esse":      "Romain Esse",
    "Aidan Zen Patrick Morris":     "Aidan Morris",
    "Shea Emmanuel Charles":        "Shea Charles",
    "Jordan Anthony James":         "Jordan James",
    "Taylor Jay Harwood-Bellis":    "Taylor Harwood-Bellis",
    "Gavin Okeroghene Bazunu":      "Gavin Bazunu",
    "Sydie Frederick Peck":         "Sydie Peck",
    "Daniel Marc Adu-Adjei":        "Daniel Adu-Adjei",
    "Charles Roger Hughes":         "Charlie Hughes",
    "Jay Steven Robinson":          "Jay Robinson",
    "Isaac Jude Price":             "Isaac Price",
    "Kasey Ian McAteer":            "Kasey McAteer",
    "Lewis Hall":                   "Lewis Hall",
    "Aaron James Ramsey":           "Aaron Ramsey",  # Aston Villa Aaron, not the older Welsh AR
    "Jack Edward Rudoni":           "Jack Rudoni",
    "Jake O'Brien":                 "Jake O'Brien",
    "Mathias Damm Kvistgaarden":    "Mathias Kvistgaarden",
    "Victor Bernth Flesner Kristiansen": "Victor Kristiansen",
    "Daniel Svensson":              "Daniel Svensson",
    "Mio Backhaus":                 "Mio Backhaus",
    # Ordering convention
    "Issahaku Abdul Fatawu":        "Abdul Fatawu",
    # Spanish dual-surname drops
    "Sergio Arribas Calvo":         "Sergio Arribas",
    "Juanlu Sánchez":               "Juanlu Sánchez",
    # Compound first names (explicit so the algorithm doesn't trip)
    "José Ángel Carmona":           "José Ángel Carmona",
    # Multi-part Brazilian / Portuguese
    "Gustavo Nunes Fernandes Gomes": "Gustavo Nunes",
    "Lucas Da Cunha":               "Lucas Da Cunha",
    "Malick Junior Yalcouyé":       "Malick Yalcouyé",
    # Dutch particle names — keep intact (algorithm catches but listing for clarity)
    "Zeno Van Den Bosch":           "Zeno Van Den Bosch",
    "Sepp van den Berg":            "Sepp van den Berg",
    # Hyphenated firsts — keep
    "Jan-Carlo Simić":              "Jan-Carlo Simić",
    "Francis-Ikechukwu Onyeka":     "Francis-Ikechukwu Onyeka",
    # Hyphenated last — keep
    "Jaden Philogene-Bidace":       "Jaden Philogene-Bidace",
    # Greek (keep full first names per spec)
    "Konstantinos Tzolakis":        "Konstantinos Tzolakis",
    "Konstantinos Koulierakis":     "Konstantinos Koulierakis",
    "Giannis Konstantelias":        "Giannis Konstantelias",
    "Christos Mouzakitis":          "Christos Mouzakitis",
    "Christos Tzolis":              "Christos Tzolis",
    # Two-token cases listed defensively
    "Takefusa Kubo":                "Takefusa Kubo",
    "Bilal El Khannouss":           "Bilal El Khannouss",
    # German full names — already short
    "Patrick Wimmer":               "Patrick Wimmer",
    "Maximilian Beier":             "Maximilian Beier",
    "Paul Nebel":                   "Paul Nebel",
    # French
    "Maghnes Akliouche":            "Maghnes Akliouche",
    "Arthur Atta":                  "Arthur Atta",
    "Ngal'ayel Mukau":              "Ngal'ayel Mukau",
    "Lucas Frédéric Stassin":       "Lucas Stassin",
    # Specific edge cases observed in data
    "Cajetan Benjamin Lenz":        "Cajetan Lenz",
    "Issa Doumbia":                 "Issa Doumbia",
}


# ─── Helper constants for the fallback simplifier ────────────────────────────

# Surname particles that signal "this is a multi-part surname, don't strip middle".
# Lowercased; comparison is case-insensitive.
_PARTICLES = {
    "van", "von", "der", "den", "de", "du", "da", "do", "dos", "del", "della",
    "di", "le", "la", "el", "al", "ten", "ter",
}

# Compound first names — when these appear as the first two tokens, treat them
# as the first name (don't strip the second token as a "middle name").
_COMPOUND_FIRSTS = {
    "josé ángel", "juan carlos", "juan pablo", "juan manuel", "juan maría",
    "juan antonio", "pedro pablo", "josé luis", "josé antonio", "josé maría",
    "marco antonio", "marco aurelio", "luis miguel", "luis enrique",
    "joão pedro", "joão félix", "joão maria", "joão paulo",
    "jean carlos", "jean pierre", "jean michel", "jean marc",
    "víctor manuel", "iván david",
}


def _auto_simplify(name: str) -> tuple[str, bool]:
    """Returns (display_name, is_low_confidence)."""
    if not name or not name.strip():
        return name, False
    tokens = name.split()

    # Single or two-token names: already short, no change
    if len(tokens) <= 2:
        return name, False

    # If any token (case-insensitive) is a surname-particle, preserve full name
    if any(t.lower() in _PARTICLES for t in tokens):
        # Likely Dutch / Portuguese / Spanish / etc. compound surname; the
        # structural particle is part of the surname.
        return name, False

    # Compound first name? Keep first two tokens + last
    first_two_lower = " ".join(tokens[:2]).lower()
    if first_two_lower in _COMPOUND_FIRSTS:
        if len(tokens) == 3:
            return name, False  # Already "Compound First + surname"
        return f"{tokens[0]} {tokens[1]} {tokens[-1]}", False

    # Default Anglo rule: 3+ tokens → first + last (strip middles)
    simplified = f"{tokens[0]} {tokens[-1]}"
    # Low-confidence if the simplified version still feels suspicious
    # (extremely long surname or unusual character density). Conservative: never flag a 2-token output.
    return simplified, False


def auto_display_for_name(official_name: str) -> tuple[str, bool]:
    """Returns (display_name, is_low_confidence).
    EXPLICIT_NAMES wins. Otherwise apply _auto_simplify."""
    if official_name in EXPLICIT_NAMES:
        return EXPLICIT_NAMES[official_name], False
    simplified, is_low = _auto_simplify(official_name)
    # Flag low-confidence: name still has 3+ tokens after simplification (means
    # particles or compound first detected, and the player isn't in EXPLICIT_NAMES
    # for human review).
    if len(simplified.split()) > LOW_CONFIDENCE_MAX_TOKENS:
        is_low = True
    return simplified, is_low


# ─── Read / write the tab ────────────────────────────────────────────────────

def read_display_tab(workbook_path: Path = WORKBOOK_PATH) -> dict[int, dict]:
    """Returns {player_id: {display_name, auto_or_manual, notes}}."""
    if not workbook_path.exists():
        return {}
    from openpyxl import load_workbook
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        if DISPLAY_TAB not in wb.sheetnames:
            return {}
        ws = wb[DISPLAY_TAB]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return {}
    headers = [str(h) if h is not None else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}
    if "player_id" not in idx or "display_name" not in idx:
        return {}
    out: dict[int, dict] = {}
    for row in rows[1:]:
        if not row or row[idx["player_id"]] is None:
            continue
        try:
            pid = int(row[idx["player_id"]])
        except (TypeError, ValueError):
            continue
        out[pid] = {
            "display_name":   str(row[idx.get("display_name", -1)] or "").strip(),
            "auto_or_manual": str(row[idx.get("auto_or_manual", -1)] or "auto").strip().lower(),
            "notes":          str(row[idx.get("notes", -1)] or "").strip()
                              if idx.get("notes", -1) >= 0 and row[idx.get("notes", -1)] is not None
                              else "",
        }
    return out


def compute_display_state(
    con: sqlite3.Connection,
    workbook_path: Path = WORKBOOK_PATH,
    club_display_map: dict[int, str] | None = None,
) -> dict:
    """Build the full player-display state.

    Returns:
      {
        'rows':          list of dicts (one per player) with display_name, etc.,
        'display_map':   dict[player_id] = display_name,
      }
    """
    existing = read_display_tab(workbook_path)

    players = list(con.execute("""
        SELECT pu.player_id, pu.name AS official_name, pu.league_id,
               pu.current_club, pu.current_club_id, pu.sellability_score
        FROM player_universe pu
        ORDER BY pu.league_id, pu.name
    """))

    rows: list[dict] = []
    display_map: dict[int, str] = {}

    for pid, official, league_id, current_club_official, current_club_id, sellability in players:
        prev = existing.get(int(pid))
        if prev and prev.get("auto_or_manual") == "manual":
            display_name = prev["display_name"] or official
            auto_or_manual = "manual"
            notes = prev.get("notes", "")
        else:
            display_name, is_low = auto_display_for_name(official)
            auto_or_manual = "auto"
            notes = LOW_CONFIDENCE_NOTE if is_low else ""

        # Current club for context — use Club Display Names if available
        current_club_display = current_club_official
        if club_display_map and current_club_id is not None:
            current_club_display = club_display_map.get(int(current_club_id), current_club_official) or current_club_official

        rows.append({
            "player_id":            int(pid),
            "official_name":        official,
            "current_club_display": current_club_display,
            "league_id":            league_id,
            "display_name":         display_name,
            "auto_or_manual":       auto_or_manual,
            "notes":                notes,
            "sellability_score":    sellability or 0.0,
        })
        display_map[int(pid)] = display_name

    return {
        "rows":        rows,
        "display_map": display_map,
    }


def write_display_tab(wb, rows: list[dict], league_names: dict[str, str]) -> None:
    """Write the Player Display Names tab to an openpyxl workbook."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    if DISPLAY_TAB in wb.sheetnames:
        del wb[DISPLAY_TAB]
    ws = wb.create_sheet(DISPLAY_TAB)

    HEADERS = ["player_id", "official_name", "current_club", "league",
               "display_name", "auto_or_manual", "notes"]
    WIDTHS  = [10, 32, 28, 22, 28, 14, 28]

    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F3864")
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = bold; c.fill = fill; c.alignment = centre
    ws.row_dimensions[1].height = 28

    ro_fill = PatternFill("solid", fgColor="F2F2F2")
    in_fill = PatternFill("solid", fgColor="FFF2CC")
    low_fill = PatternFill("solid", fgColor="FCE4D6")
    wrap = Alignment(wrap_text=True, vertical="top")

    # Sort by league, then display_name
    sorted_rows = sorted(rows, key=lambda r: (r["league_id"] or "", r["display_name"] or ""))
    for ri, r in enumerate(sorted_rows, start=2):
        ws.cell(row=ri, column=1, value=r["player_id"]).fill = ro_fill
        c2 = ws.cell(row=ri, column=2, value=r["official_name"])
        c2.fill = ro_fill; c2.alignment = wrap
        ws.cell(row=ri, column=3, value=r["current_club_display"]).fill = ro_fill
        ws.cell(row=ri, column=4, value=league_names.get(r["league_id"], r["league_id"])).fill = ro_fill
        c5 = ws.cell(row=ri, column=5, value=r["display_name"])
        c5.fill = in_fill; c5.alignment = wrap
        c6 = ws.cell(row=ri, column=6, value=r["auto_or_manual"])
        c6.fill = in_fill; c6.alignment = Alignment(horizontal="center")
        c7 = ws.cell(row=ri, column=7, value=r["notes"])
        c7.fill = low_fill if r["notes"] else ro_fill
        c7.alignment = wrap

    dv = DataValidation(type="list", formula1='"auto,manual"', allow_blank=False)
    dv.error = "Use 'auto' or 'manual'."
    dv.errorTitle = "Invalid value"
    letter = get_column_letter(6)
    dv.add(f"{letter}2:{letter}{len(sorted_rows) + 1}")
    ws.add_data_validation(dv)

    for ci, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "E2"
    last_col = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{len(sorted_rows) + 1}"


# ─── Lightweight lookup for non-pipeline consumers ──────────────────────────

def load_display_map(workbook_path: Path = WORKBOOK_PATH) -> dict[int, str]:
    existing = read_display_tab(workbook_path)
    return {pid: v["display_name"] for pid, v in existing.items() if v.get("display_name")}


def display_for(player_id: int | None, official_fallback: str | None,
                display_map: dict[int, str]) -> str:
    if player_id is None:
        return official_fallback or ""
    return display_map.get(int(player_id)) or (official_fallback or "")


def apply_to_text_list(text: str | None, name_map: dict[str, str]) -> str:
    """For text fields like top_3_likely_to_move = 'Name; Name; Name', swap each
    name segment to its display form. Looks up by official_name → display_name.
    Returns original text if no map entries match."""
    if not text:
        return text or ""
    parts = [p.strip() for p in str(text).split(";")]
    swapped = [name_map.get(p, p) for p in parts]
    return "; ".join(swapped)
